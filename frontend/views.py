from django.shortcuts import render
import firebase_admin
#simport pandas as pd
from firebase_admin import credentials, initialize_app
from firebase_admin import firestore
from firebase import firebase
from django.contrib.auth.decorators import login_required
from django.contrib.auth.hashers import make_password
from django.shortcuts import render, redirect
from django.contrib.auth import authenticate, login, logout
from django.contrib import messages
from api.models import User
from .models import users
from django.db.models import Q
from django.contrib.auth import get_user_model
from frontend.models import bites
import logging
from .forms import DocumentForm
from .forms import *
from django.http import JsonResponse
from .models import collection
from .models import *
from datetime import datetime
from .models import tags
from django.http import HttpResponseForbidden
from functools import wraps
from django.shortcuts import render, get_object_or_404
from django.core.paginator import Paginator
import openpyxl
from io import BytesIO
import pandas as pd
from django.core.files.storage import default_storage
from django.conf import settings
from pathlib import Path
from django.http import HttpResponseForbidden, HttpResponse
import logging
import pytz



db = firestore.client()
firebase_app = firebase.FirebaseApplication('https://techcare-diabetes.firebaseio.com', None)


def admin_required(view_func):
    @wraps(view_func)
    def _wrapped_view(request, *args, **kwargs):
        if not request.user.is_superuser:
            return HttpResponseForbidden("Unauthorized") # Redirect to a permission denied page or login page
        return view_func(request, *args, **kwargs)
    return _wrapped_view

config = {
    'apiKey': 'AIzaSyDYH31LLGfe492t4xklzTeZrIy_Rs-Om3M',
    'authDomain': 'techcare-diabetes.firebaseapp.com',
    'projectId': 'techcare-diabetes',
    'storageBucket': 'techcare-diabetes.appspot.com',
    'messagingSenderId': '1086234468488',
    'appId': '1:1086234468488:web:d32c01a322df2e0e76ff3d',
    'measurementId': 'G-VWVCF3QKTZ'
}

def index(request):
    return render(request, 'frontend/index.html'),

def logout_page(request):
    logout(request)
    return redirect('login')

def index(request):
    return render(request, 'frontend/index.html'),

def logout_page(request):
    logout(request)
    return redirect('login')


db = firestore.client()

def get_document_reference(path):
            if path:
        # Ensure the path does not start with a '/'
                if path.startswith('/'):
                  path = path[1:]
                path_elements = path.split('/')
        # Check if the path has exactly 2 elements
                if len(path_elements) == 2:
            # Return a DocumentReference object instead of a string
                    return db.collection(path_elements[0]).document(path_elements[1])
                else:
                 raise ValueError(f"Invalid path: {path}")
            return None


def login_page(request):
    page = 'login'

    if request.user.is_authenticated:
        if request.user.is_patient():
            return redirect('patient-homepage')
        else:
            return redirect('doctor-homepage')

    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')

        try:
            user = User.objects.filter(username=username).first()
        except:
            messages.error(request, 'User does not exist')

        if request.user is not None and request.user.is_authenticated:
            messages.error(request, 'Your account is still pending approval.')
        else:
            user = authenticate(request, username=username, password=password)

            if user is not None:
                if user.is_patient():
                    # Log in the patient and redirect to the patient homepage
                    login(request, user)
                    return redirect('patient-homepage')
                elif user.is_Doctor():
                    # Log in the doctor and redirect to the doctor if approved
                    login(request, user)
                    return redirect('doctor-homepage')
            else:
                # Display error message if authentication fails
                messages.error(request, 'Invalid username or password')

    context = {'page': page}
    return render(request, 'login_signup.html', context)

   # def trivia_template_view(request):
    
    #  return render(request, 'patient/trivia.html')



@login_required(login_url='/login')
def logout_user(request):
    logout(request)
    return redirect('login')


def patient_signup(request):
    page = 'patientSignup'
    form = PatientForm()

    if request.method == 'POST':
        form = PatientForm(request.POST)
        if form.is_valid():
            form.instance.password = make_password(form.cleaned_data['password'])
            form.save()
            messages.success(request, 'Account created successfully')
            return redirect('login')

    context = {'page': page,
               'form': form, }
    return render(request, 'frontend/patient_signup.html', context)


def doctor_signup(request):
    page = 'doctorSignup'
    form = DoctorForm()

    if request.method == 'POST':
        form = DoctorForm(request.POST)
        if form.is_valid():
            form.instance.password = make_password(form.cleaned_data['password'])
            form.save(commit=False)
            form.save()
            messages.success(request, 'Your account has been created and is pending approval.')
            return redirect('login')
        else:
            messages.error(request, 'error occurred')

    context = {'page': page,
               'form': form, }
    return render(request, 'frontend/doctor_signup.html', context)


@login_required(login_url='/login')
def patient_homepage(request):
    return render(request, 'frontend/patient/patientHomepage.html')


@login_required(login_url='/login')
def doctor_homepage(request):
    return render(request, 'frontend/doctor/doctorHomepage.html')


@login_required(login_url='/login')
def calendar(request):
    return render(request, 'frontend/calendar.html')


@login_required(login_url='/login')
def patients_list(request):
    collection = db.collection("users")
    documents = collection.stream()
    document_data = [{'id': doc.id, 'data': doc.to_dict()} for doc in documents]

    query = request.GET.get('q')

    if query:
        filtered_data = []
        for entry in document_data:
            if (query.lower() in entry['data'].get('display_name', '').lower()) or \
               (query.lower() in entry['data'].get('email', '').lower()) or \
               (query.lower() in entry['data'].get('gender', '').lower()) or \
               (query.lower() in str(entry['data'].get('height', '')).lower()) or \
               (query.lower() in str(entry['data'].get('weight', '')).lower()) or \
               (query.lower() in entry['data'].get('yearOfDiagnosis', '').lower()):
                filtered_data.append(entry)
    else:
        filtered_data = document_data

    # Pagination setup
    paginator = Paginator(filtered_data, 20)  # Show 20 patients per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, 'frontend/doctor/patients.html', {'page_obj': page_obj, 'query': query})

def calculate_bmi(weight, height):
    # Convert height from cm to meters
    height_meters = height / 100.0

    # Calculate BMI
    bmi = weight / (height_meters * height_meters)

    # Round BMI to one decimal place
    return round(bmi, 1)

from datetime import datetime, timedelta
from math import sqrt
def patients_detail(request, document_name):
    db = firestore.Client()

  
    
    # Fetch user document
    user_collection = db.collection("users")
    user_document_ref = user_collection.document(document_name)
    user_document = user_document_ref.get()

    if not user_document.exists:
        return render(request, 'frontend/techcare_data/document_not_found.html')

    document_data = user_document.to_dict()

    print(document_data)
    biomarkers_collection = db.collection("biomarkers")
    biomarkers_query = biomarkers_collection.where('user', '==', user_document.reference).stream()
    weights = []
    labels = []
        # Collect weights and corresponding timestamps
    for biomarker in biomarkers_query:
        data = biomarker.to_dict()
        if 'weight' in data and 'time' in data:  # Ensure 'time' and 'weight' are both present
            weights.append(data['weight'])
            labels.append(data['time'])  # Assuming 'time' is a timestamp field

    # Pair weights with corresponding labels (timestamps)
    paired_weights_labels = list(zip(weights, labels))

    # Sort the pairs by the 'labels' (time field)
    sorted_paired_weights_labels = sorted(paired_weights_labels, key=lambda x: x[1])  # Sort by time

    # Unzip the sorted pairs back into weights and labels
    weights, labels = zip(*sorted_paired_weights_labels)

    # If you need lists, convert back from tuples:
    weights = list(weights)
    labels = list(labels)

    weight = document_data.get('weight')
    height = document_data.get('height')
    
    if weight is not None and height is not None:
        # Calculate BMI
        bmi = calculate_bmi(weight, height)
    else:
        bmi = None  # or some default value or handle error
     # Extract current weight from Firestore
    current_weight = document_data.get("weight", None)

    if current_weight is None:
        return render(request, 'frontend/techcare_data/document_not_found.html')

    # Retrieve or initialize the weight history from the session
    weight_history = request.session.get('weight_history', [])

    # Append the current weight if it's not the same as the last one in the history
    if len(weight_history) == 0 or current_weight != weight_history[-1]:
        weight_history.append(current_weight)

    # Store the updated weight history back into the session
    request.session['weight_history'] = weight_history

    # Create labels (index values for each entry in history)
    labels = list(range(1, len(weight_history) + 1))


    # Fetch readBites data for the user
    readbites_collection = db.collection("readBites")
    readbites_query = readbites_collection.where('user', '==', f'/users/{document_name}').stream()
    readbites_data = []
    for doc in readbites_query:
        readbites_data.append(doc.to_dict())

    readstories_collection = db.collection("readStories")
    readstories_query = readstories_collection.where('user', '==', f'/users/{document_name}').stream()
    readstories_data = []
    for doc in readstories_query:
        readstories_data.append(doc.to_dict())

    suggestedActivities_collection = db.collection("suggestedActivities")
    suggestedActivities_query = suggestedActivities_collection.where('user', '==', f'/users/{document_name}').stream()
    suggestedActivities_data = []
    for doc in suggestedActivities_query:
        suggestedActivities_data.append(doc.to_dict())

    suggestedBites_collection = db.collection("suggestedBites")
    suggestedBites_query = suggestedBites_collection.where('user', '==', f'/users/{document_name}').stream()
    suggestedBites_data = []
    for doc in suggestedBites_query:
        suggestedBites_data.append(doc.to_dict())

    suggestedInAppLinks_collection = db.collection("suggestedInAppLinks")
    suggestedInAppLinks_query = suggestedInAppLinks_collection.where('user', '==', f'/users/{document_name}').stream()
    suggestedInAppLinks_data = []
    for doc in suggestedInAppLinks_query:
        suggestedInAppLinks_data.append(doc.to_dict())

    suggestedJournals_collection = db.collection("suggestedJournals")
    suggestedJournals_query = suggestedJournals_collection.where('user', '==', f'/users/{document_name}').stream()
    suggestedJournals_data = []
    for doc in suggestedJournals_query:
        suggestedJournals_data.append(doc.to_dict())

    suggestedSelfAwarnessBites_collection = db.collection("suggestedSelfAwarnessBites")
    suggestedSelfAwarnessBites_query = suggestedSelfAwarnessBites_collection.where('user', '==', f'/users/{document_name}').stream()
    suggestedSelfAwarnessBites_data = []
    for doc in suggestedSelfAwarnessBites_query:
        suggestedSelfAwarnessBites_data.append(doc.to_dict())

    suggestedWildCards_collection = db.collection("suggestedWildCards")
    suggestedWildCards_query = suggestedWildCards_collection.where('user', '==', f'/users/{document_name}').stream()
    suggestedWildCards_data = []
    for doc in suggestedWildCards_query:
        suggestedWildCards_data.append(doc.to_dict())

    selfLadder_collection = db.collection("selfLadder")
    selfLadder_query = selfLadder_collection.where('userID', '==', f'/users/{document_name}').stream()
    selfLadder_data = []
    for doc in selfLadder_query:
        selfLadder_data.append(doc.to_dict())

    psychomarkers_collection = db.collection("psychomarkers")
    psychomarkers_query = psychomarkers_collection.where('user', '==', user_document.reference).stream()
    psychomarkers_data = []
    for doc in psychomarkers_query:
        psychomarkers_data.append(doc.to_dict())

    inquiry_collection = db.collection("inquiry")
    inquiry_query = inquiry_collection.where('user', '==', f'/users/{document_name}').stream()
    inquiry_data = []
    for doc in inquiry_query:
        inquiry_data.append(doc.to_dict())

    biomarkers_collection = db.collection("biomarkers")
    biomarkers_query = biomarkers_collection.where('user', '==', user_document.reference).stream()
    biomarkers_data=[]
 
    for doc in biomarkers_query:
        biomarkers_data.append(doc.to_dict())

    dailyBloodGlucoseAverage_collection = db.collection("dailyBloodGlucoseAverage")
    dailyBloodGlucoseAverage_query = dailyBloodGlucoseAverage_collection.where('user', '==', f'/users/{document_name}').stream()
    dailyBloodGlucoseAverage_data = []
    for doc in dailyBloodGlucoseAverage_query:
        dailyBloodGlucoseAverage_data.append(doc.to_dict())

    feelings_collection = db.collection("feelings")
    feelings_query = feelings_collection.where('user', '==', f'/users/{document_name}').stream()
    feelings_data = []
    for doc in feelings_query:
        feelings_data.append(doc.to_dict())

        biomarkers_data = []

    # Fetch biomarkers data for the user
    biomarkers_collection = db.collection("biomarkers")
    biomarkers_forchart_query = biomarkers_collection.where('user', '==', user_document.reference).stream()
    chartData=[]

    for doc in biomarkers_forchart_query:
        chartData.append(doc.to_dict())
     # Paginate psychomarkers_data

    print(biomarkers_forchart_query)

    paginator_psychomarkers = Paginator(psychomarkers_data, 10)  # Show 10 psychomarkers per page
    page_number_psychomarkers = request.GET.get('page_psychomarkers')
    page_psychomarkers = paginator_psychomarkers.get_page(page_number_psychomarkers)

    # Paginate biomarkers_data
    paginator_biomarkers = Paginator(biomarkers_data, 10)  # Show 10 biomarkers per page
    page_number_biomarkers = request.GET.get('page_biomarkers')
    page_biomarkers = paginator_biomarkers.get_page(page_number_biomarkers)

    now = datetime.now(pytz.utc)
    cutoff_time = now - timedelta(days=1)
    # Query the biomarkers collection
  

    # Query the biomarkers collection
    biomarkers_collection = db.collection("biomarkers")
    biomarkers_query = biomarkers_collection.where('user', '==', user_document.reference).stream()

    # Initialize data structures for chart and pagination
    data_for_chartjs = {
        'labels': [],
        'datasets': [{
            'label': 'Blood Glucose Levels',
            'data': [],
            'borderColor': 'rgb(75, 192, 192)',
            'fill': False
         }, {
            'label': 'Daily Activity',
            'data': [],
            'borderColor': 'rgb(255, 99, 132)',
            'fill': False
        }, {
            'label': 'Daily Carbs',
            'data': [],
            'borderColor': 'rgb(255, 99, 132)',
            'fill': False
        }],
        'min_timestamp': None,  # To store min timestamp
        'max_timestamp': None 
    }

    data_for_chartjs_weekly = {
        'labels': [],
        'datasets': [{
            'label': 'Blood Glucose Levels',
            'data': [],
            'borderColor': 'rgb(75, 192, 192)',
            'fill': False
         }]
    }

    biomarkers_data = []

    # Debugging: Print cutoff time
    print("Cutoff time:", cutoff_time)


    cutoff_time_weekly = now - timedelta(days=7)
    
    sorted_chartData = sorted(chartData, key=lambda x: x['time'].astimezone(pytz.utc))

    # Check if query is returning data
    found_data = False
    for doc in sorted_chartData:
        found_data = True
       

        if 'time' in doc and ('bloodGlucose' in doc or 'dailyActivity' in doc or 'dailyCarbs'):
            try:
                timestamp = doc['time'].astimezone(pytz.utc)
                if timestamp > cutoff_time:
                    time_str = timestamp.isoformat()
                    data_for_chartjs['labels'].append(time_str)
                    if 'bloodGlucose' in doc:
                        data_for_chartjs['datasets'][0]['data'].append(doc['bloodGlucose'])
                    if 'dailyActivity' in doc:
                        data_for_chartjs['datasets'][1]['data'].append(doc['dailyActivity'])
                    if 'dailyCarbs' in doc:
                        data_for_chartjs['datasets'][2]['data'].append(doc['dailyCarbs'])

                    # Set min and max timestamps
                    if data_for_chartjs['min_timestamp'] is None or timestamp < data_for_chartjs['min_timestamp']:
                        data_for_chartjs['min_timestamp'] = timestamp
                    if data_for_chartjs['max_timestamp'] is None or timestamp > data_for_chartjs['max_timestamp']:
                        data_for_chartjs['max_timestamp'] = timestamp
                        
            except Exception as e:
                print("Error processing timestamp:", e)
                continue

    if not found_data:
        print("No documents found matching the query.")

    print("data for chartr",data_for_chartjs)

  
    for doc in sorted_chartData:
       
        if 'time' in doc and 'bloodGlucose' in doc :
            try:
                timestamp = doc['time'].astimezone(pytz.utc)
                if timestamp > cutoff_time_weekly:
                    time_str = timestamp.isoformat()
                    data_for_chartjs_weekly['labels'].append(time_str)
                    if 'bloodGlucose' in doc:
                        data_for_chartjs_weekly['datasets'][0]['data'].append(doc['bloodGlucose'])
            except Exception as e:
                print("Error processing timestamp:", e)
                continue

    if not found_data:
        print("No documents found matching the query.")
    biomarkers_collection = db.collection("biomarkers")
    biomarkers_with_bloodGlucose_query = biomarkers_collection.where('user', '==', user_document.reference).stream()
    bloodGlucose_biomarkers_data = []
    
    for doc in biomarkers_with_bloodGlucose_query:
        bloodGlucose_biomarkers_data.append(doc.to_dict())
        
        

    fasting_and_before_meal = []
    after_meal = []    
    
    # Loop through the data and categorize based on bloodGlucoseType
    for doc in bloodGlucose_biomarkers_data:
        if 'bloodGlucose' in doc and 'bloodGlucoseType' in doc:
            if doc['bloodGlucoseType'] in ['صيامي', 'قبل الوجبة']:  
                fasting_and_before_meal.append(doc['bloodGlucose'])
            elif doc['bloodGlucoseType'] == 'بعد الوجبة':  
                after_meal.append(doc['bloodGlucose'])

    # Calculate averages
    if fasting_and_before_meal:
        avg_fasting_before_meal = sum(fasting_and_before_meal) / len(fasting_and_before_meal)
    else:
        avg_fasting_before_meal = None

    if after_meal:
        avg_after_meal = sum(after_meal) / len(after_meal)
    else:
        avg_after_meal = None

    avg_fasting_before_meal = round(sum(fasting_and_before_meal) / len(fasting_and_before_meal), 2) if fasting_and_before_meal else None
    avg_after_meal = round(sum(after_meal) / len(after_meal), 2) if after_meal else None
    # Calculate highest and lowest blood glucose readings
    all_readings = fasting_and_before_meal + after_meal
    if all_readings:
        highest_blood_glucose = max(all_readings)
        lowest_blood_glucose = min(all_readings)
    else:
        highest_blood_glucose = None
        lowest_blood_glucose = None

    
    
    bloodGlucose =[]   
    for doc in bloodGlucose_biomarkers_data:
        if 'bloodGlucose' in doc and 'bloodGlucoseType' in doc:
            bloodGlucose.append(doc)
    
    
    number_of_readings=len(bloodGlucose)
    
    # Initialize counters for each range
    very_low_count = 0
    low_count = 0
    target_count = 0
    high_count = 0
    very_high_count = 0

    # Categorize blood glucose values
    for doc in bloodGlucose:
        value = doc['bloodGlucose']
        if value < 54:
            very_low_count += 1
        elif 54 <= value < 70:
            low_count += 1
        elif 70 <= value < 180:
            target_count += 1
        elif 180 <= value < 250:
            high_count += 1
        else:  # value >= 250
            very_high_count += 1
            
    # Calculate percentages for each range
    very_low_percentage = (very_low_count / number_of_readings * 100) if number_of_readings > 0 else 0
    low_percentage = (low_count / number_of_readings * 100) if number_of_readings > 0 else 0
    target_percentage = (target_count / number_of_readings * 100) if number_of_readings > 0 else 0
    high_percentage = (high_count / number_of_readings * 100) if number_of_readings > 0 else 0
    very_high_percentage = (very_high_count / number_of_readings * 100) if number_of_readings > 0 else 0

    # Round the percentages
    very_low_percentage = round(very_low_percentage, 2)
    low_percentage = round(low_percentage, 2)
    target_percentage = round(target_percentage, 2)
    high_percentage = round(high_percentage, 2)
    very_high_percentage = round(very_high_percentage, 2)
    
    
    bloodGlucose_values = []
    for doc in bloodGlucose_biomarkers_data:
        if 'bloodGlucose' in doc and 'bloodGlucoseType' in doc:
            bloodGlucose_values.append(doc['bloodGlucose']) 

    number_of_readings = len(bloodGlucose_values)
    if number_of_readings > 0:
        sum_blood_glucose = sum(bloodGlucose_values)
        average_blood_glucose = sum_blood_glucose / number_of_readings
    else:
        average_blood_glucose = 0  

    average_blood_glucose=round(average_blood_glucose,2)
        
        
   
    if number_of_readings > 0:
   
        timestamps = [doc['time'].astimezone(pytz.utc) for doc in bloodGlucose if 'time' in doc]
        
        if timestamps:
            first_reading = min(timestamps)
            last_reading = max(timestamps)

            # Calculate the number of days between first and last reading
            days_span = (last_reading - first_reading).days + 1  # Add 1 to include the last day

            # Calculate average readings per day
            average_readings_per_day = number_of_readings / days_span
            average_readings_per_day=round(average_readings_per_day,2)
            
        else:
            
            average_readings_per_day = None
    else:
        average_readings_per_day = None

    
    
    
    all_blood_glucose = []
    for doc in bloodGlucose_biomarkers_data:
        if 'bloodGlucose' in doc:
            all_blood_glucose.append(doc['bloodGlucose'])
            
    all_cv = calculate_cv(all_blood_glucose)
    
    context = {
    'document_data': document_data,
    'readbites_data': readbites_data, 
    'readstories_data': readstories_data, 
    'suggestedActivities_data': suggestedActivities_data, 
    'suggestedBites_data': suggestedBites_data, 
    'suggestedInAppLinks_data': suggestedInAppLinks_data, 
    'suggestedJournals_data': suggestedJournals_data, 
    'suggestedSelfAwarnessBites_data': suggestedSelfAwarnessBites_data, 
    'suggestedWildCards_data': suggestedWildCards_data,
    'selfLadder_data': selfLadder_data,
    'psychomarkers_data': psychomarkers_data,
    'page_psychomarkers': page_psychomarkers,
    'inquiry_data': inquiry_data,
    'page_biomarkers': page_biomarkers,
    'biomarkers_data': biomarkers_data,
    'dailyBloodGlucoseAverage_data': dailyBloodGlucoseAverage_data,
    'data_for_chartjs': data_for_chartjs,
    'data_for_chartjs_weekly': data_for_chartjs_weekly,
    'labels': labels,
    'weights': weights,
    'bmi': bmi,
    'number_of_readings':number_of_readings,
    'average_readings_per_day': average_readings_per_day,
    'avg_fasting_before_meal': avg_fasting_before_meal,
    'avg_after_meal': avg_after_meal,
    'highest_blood_glucose': highest_blood_glucose,
    'lowest_blood_glucose': lowest_blood_glucose,
    'variation': all_cv,
    'average_blood_glucose':average_blood_glucose,
    'very_low_percentage': very_low_percentage,
    'low_percentage': low_percentage,
    'target_percentage': target_percentage,
    'high_percentage': high_percentage,
    'very_high_percentage': very_high_percentage,
    }
    
   
    return render(request, 'frontend/techcare_data/patientssdocument.html', context)

import statistics

# Function to calculate Coefficient of Variation (CV)
def calculate_cv(data):
    if len(data) < 2 :
        return None  # Return None if no data

    mean = statistics.mean(data)
    std_dev = statistics.stdev(data)  # Standard deviation
    cv = (std_dev / mean) * 100 if mean != 0 else None  # CV in percentage

    return round(cv, 1)

@login_required(login_url='/login')
def doctor_chat_page(request):
    return render(request, 'frontend/doctor/doctor-chat.html')


    
@login_required(login_url='/login')
def mind_activities_view(request):
    return render(request,'frontend/patient/activities.html')

@login_required(login_url='/login')
def quiz_question_view(request):
    return render(request,'frontend/patient/quiz.html')



def update_bites_with_ngrams():
    bites_collection = db.collection("bites")
    bites_documents = bites_collection.stream()

    for doc in bites_documents:
        data = doc.to_dict()
        title = data.get('title', '')
        if title:  # Only process if title is not empty
            title_ngrams = generate_ngrams(title, 3)  # Adjust the value of n as needed
            doc.reference.update({'ngrams': title_ngrams})



def bites_view(request):
    db = firestore.client()

    # Fetching bites data
    collection = db.collection("bites")
    documents = collection.stream()
    document_data = [{'name': doc.id, 'data': doc.to_dict()} for doc in documents]

    # Fetching tags data
    collection = db.collection("tags")
    documents = collection.stream()
    tags_data = [{'name': doc.id, 'data': doc.to_dict()} for doc in documents]

    

    form = BitesForm() 

    return render(request, 'frontend/techcare_data/bites_table.html', {
        'form': form, 
        'tags_data': tags_data,
        'document_data': document_data,
    })


def selfawarenessScenarios_view(request):
    db = firestore.client()
    collection = db.collection("selfawarenessScenarios")
    documents = collection.stream()
    document_data = [{'name': doc.id, 'data': doc.to_dict()} for doc in documents]

    query = request.GET.get('q')  # Get search query from request
    if query:
        # Perform search based on 'title' field (modify as per your Firestore structure)
        filtered_documents = []
        for doc in document_data:
            title = doc['data'].get('storyTitle', '').lower()
            if query.lower() in title:
                filtered_documents.append(doc)
        document_data = filtered_documents

    
    


    collection = db.collection("activities")
    documents = collection.stream()
    activity_data = [{'name': doc.id, 'data': doc.to_dict()} for doc in documents]

    collection = db.collection("selfAwarnessBites")
    documents = collection.stream()
    selfAwarnessBites_data = [{'name': doc.id, 'data': doc.to_dict()} for doc in documents]

    collection = db.collection("inAppLinks")
    documents = collection.stream()
    inAppLinks_data = [{'name': doc.id, 'data': doc.to_dict()} for doc in documents]

    collection = db.collection("journalPrompt")
    documents = collection.stream()
    journalPrompt_data = [{'name': doc.id, 'data': doc.to_dict()} for doc in documents]

    collection = db.collection("bites")
    documents = collection.stream()
    bites_data = [{'name': doc.id, 'data': doc.to_dict()} for doc in documents]

    collection = db.collection("wildCard")
    documents = collection.stream()
    wildCard_data = [{'name': doc.id, 'data': doc.to_dict()} for doc in documents]


    form = SelfAwarenessScenariosForm()

    paginator = Paginator(document_data, 20)  # Show 20 bites per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    context = {'activity_data': activity_data, 'selfAwarnessBites_data': selfAwarnessBites_data, 'inAppLinks_data': inAppLinks_data, 'journalPrompt_data': journalPrompt_data, 'bites_data': bites_data, 'wildCard_data': wildCard_data, 'page_obj': page_obj ,'form':form, 'query':query,   }


    return render(request, 'frontend/techcare_data/selfawarenessScenarios_table.html', context)

def bites_view(request):
    db = firestore.client()

    # Fetching bites data
    collection = db.collection("bites")
    documents = collection.stream()
    document_data = [{'name': doc.id, 'data': doc.to_dict()} for doc in documents]
    
    query = request.GET.get('q')  # Get search query from request
    if query:
        # Perform search based on 'title' field (modify as per your Firestore structure)
        filtered_documents = []
        for doc in document_data:
            title = doc['data'].get('title', '').lower()
            if query.lower() in title:
                filtered_documents.append(doc)
        document_data = filtered_documents

    # Fetching tags data
    collection = db.collection("tags")
    documents = collection.stream()
    tags_data = [{'name': doc.id, 'data': doc.to_dict()} for doc in documents]

    # Fetching categories data
    collection = db.collection("categories")
    documents = collection.stream()
    categories_data = [{'name': doc.id, 'data': doc.to_dict()} for doc in documents]

    # Fetching users data
    collection = db.collection("users")
    documents = collection.stream()
    users_data = [{'name': doc.id, 'data': doc.to_dict()} for doc in documents]

    # Pagination for bites data
    paginator = Paginator(document_data, 20)  # Show 20 bites per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    form = BitesForm()  # Assuming BitesForm is defined in forms.py

    return render(request, 'frontend/techcare_data/bites_table.html', {
        'form': form,
        'tags_data': tags_data,
        'categories_data': categories_data,
        'users_data': users_data,
        'page_obj': page_obj,  # Pass the paginated bites data to the template
        'query': query  # Pass the search query to the template
    })





def bitesdocument_detail(request, document_name):
    db = firestore.Client()
    collection = db.collection("bites")
    document_ref = collection.document(document_name)
    document_data = document_ref.get().to_dict()

    if not document_data:

        return render(request, 'frontend/techcare_data/document_not_found.html')

    return render(request, 'frontend/techcare_data/bitesdocument.html', {'document_data': document_data})



def create_bite(request):
    if request.method == 'POST':
        form = BitesForm(request.POST)
        db = firestore.client()
        
        if form.is_valid():
            data = {
                'title': form.cleaned_data['title'],
                'tags': request.POST.get('tags'),
                'difficulty': form.cleaned_data['difficulty'],
                'categories': request.POST.get('categories'),
                'content': form.cleaned_data['content'],
            }
            db.collection("bites").document().set(data)  
            messages.success(request, 'Successfully created Bites.')  
            return redirect('bites_view') 
        else:
            messages.error(request, 'Error creating Bites. Please check your input.')
            return redirect('bites_view') 



def get_all_collections(request):
    db = firestore.client()
    all_collections = [collection.id for collection in db.collections()]
    return {'collections': all_collections}

def handle_form_submission(request):
    try:
        if request.method == 'POST':
            selected_collection = request.POST.get('collection_dropdown')
            collection_views = {
                'bites': 'bites_view',
                'activities': 'activities_view',
                'badges': 'badges_view',
                'biomarkers': 'biomarkers_view',
                'assessmentQuestion': 'assessmentQuestion_view',
                'categories': 'categories_view',
                'feelings': 'feelings_view',
                'inAppLinks': 'inAppLinks_view',
                'inquiry': 'inquiry_view',
                'items': 'items_view',
                'journal': 'journal_view',
                'journalPrompt': 'journalPrompt_view',
                'majorAssessment': 'majorAssessment_view',
                'psychomarkers': 'psychomarkers_view',
                'scenarios': 'scenarios_view',
                'shortBite': 'shortBite_view',
                'tags': 'tags_view',
                'trivia': 'trivia_view',
                'users': 'users_view',
                'selfawarenessScenarios': 'selfawarenessScenarios_view',
                'assets': 'assets_view',
                'notifications':'notifications_view',
                'nutrition': 'nutrition_view',
                'readBites': 'readBites_view',
                'readStories': 'readStories_view',
                'selfAwarnessBites': 'selfAwarnessBites_view',
                'selfawareness_collection': 'selfawareness_collection_view',
                'suggestedActivities': 'suggestedActivities_view',
                'suggestedBites': 'suggestedBites_view',
                'suggestedInAppLinks': 'suggestedInAppLinks_view',
                'suggestedJournals': 'suggestedJournals_view',
                'suggestedSelfAwarnessBites': 'suggestedSelfAwarnessBites_view',
                'suggestedWildCards': 'suggestedWildCards_view',
                'testTrivia': 'testTrivia_view',
                'wildCard': 'wildCard_view',
                'selfLadder': 'selfladder_view',

            }

            if selected_collection in collection_views:
                # If yes, redirect to the corresponding view
                return redirect(collection_views[selected_collection])

    except Exception as e:
        # Handle the exception or log the error
        print(f"An error occurred: {e}")

    # Handle other cases or render the page
    return render(request, 'frontend/techcare_data/collections.html')






def activities_view(request):

    # Fetch activities data
    activities_collection = db.collection("activities")
    activities_documents = activities_collection.stream()
    activities_data = [{'name': doc.id, 'data': doc.to_dict()} for doc in activities_documents]

    query = request.GET.get('q')  # Get search query from request
    if query:
        # Perform search based on 'title' field (modify as per your Firestore structure)
        filtered_documents = []
        for doc in activities_data:
            title = doc['data'].get('title', '').lower()
            if query.lower() in title:
                filtered_documents.append(doc)
        activities_data = filtered_documents

    # Paginate activities data
    activities_paginator = Paginator(activities_data, 20)  # Show 10 activities per page
    activities_page_number = request.GET.get('activities_page')
    activities_page_obj = activities_paginator.get_page(activities_page_number)

    form = ActivitiesForm()

    # Fetch tags data
    tags_collection = db.collection("tags")
    tags_documents = tags_collection.stream()
    tags_data = [{'name': doc.id, 'data': doc.to_dict()} for doc in tags_documents]

    # Paginate tags data
    tags_paginator = Paginator(tags_data, 20)  # Show 10 tags per page
    tags_page_number = request.GET.get('tags_page')
    tags_page_obj = tags_paginator.get_page(tags_page_number)

    return render(request, 'frontend/techcare_data/activities_table.html', {
        'form': form,
        'tags_data': tags_page_obj,
        'activities_page_obj': activities_page_obj,
        'query':query
    })



def assets_view(request):
    db = firestore.Client()
    collection = db.collection("assets")
    documents = collection.stream()
    document_data = [{'name': doc.id, 'data': doc.to_dict()} for doc in documents]
    form= AssetsForm()
    query = request.GET.get('q')  # Get search query from request
    if query:
        # Perform search based on 'title' field (modify as per your Firestore structure)
        filtered_documents = []
        for doc in document_data:
            name = doc['data'].get('name', '').lower()
            label = doc['data'].get('label', '').lower()
            
            if query.lower() in name or query.lower() in label:
                filtered_documents.append(doc)
        document_data = filtered_documents

    paginator = Paginator(document_data, 20)  # 20 items per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)    

    major_collection = db.collection("majorAssessment")
    major_documents = major_collection.stream()
    major_data = [{'name': doc.id, 'data': doc.to_dict()} for doc in major_documents]

    major_paginator = Paginator(major_data, 20)  # Show 10 major assessments per page
    major_page_number = request.GET.get('major_page')
    major_page_obj = major_paginator.get_page(major_page_number)

    

    return render(request, 'frontend/techcare_data/assets_table.html', {'page_obj': page_obj, 'query':query,'form':form, 'majorAssessment':major_page_obj})


def nutrition_view(request):
    db = firestore.Client()
    
    # Fetching nutrition data
    collection = db.collection("nutrition")
    documents = collection.stream()
    document_data = [{'name': doc.id, 'data': doc.to_dict()} for doc in documents]
    
    query = request.GET.get('q')  # Get search query from request
    if query:
        # Perform search based on 'unit', 'name_en', or 'name_ar' fields
        query_lower = query.lower()
        filtered_documents = []
        for doc in document_data:
            data = doc['data']
            unit = str(data.get('unit', '')).lower()
            name_en = str(data.get('name_en', '')).lower()
            name_ar = str(data.get('name_ar', '')).lower()
            if query_lower in unit or query_lower in name_en or query_lower in name_ar:
                filtered_documents.append(doc)
        document_data = filtered_documents

    # Implementing pagination for nutrition data
    paginator = Paginator(document_data, 20)  # Show 20 nutrition entries per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    form = NutritionForm()
    return render(request, 'frontend/techcare_data/nutrition_table.html', {
        'form': form, 
        'page_obj': page_obj,
        'query': query
    })


def notifications_view(request):
    db = firestore.Client()
    
    # Fetching nutrition data
    collection = db.collection("notifications")
    documents = collection.stream()
    document_data = [{'name': doc.id, 'data': doc.to_dict()} for doc in documents]
    
    query = request.GET.get('q')  # Get search query from request
    if query:
        # Perform search based on 'unit', 'name_en', or 'name_ar' fields
        query_lower = query.lower()
        filtered_documents = []
        for doc in document_data:
            data = doc['data']
            body = str(data.get('body', '')).lower()
            title = str(data.get('title', '')).lower()
            time = str(data.get('time', '')).lower()
            if query_lower in body or query_lower in title or query_lower in time:
                filtered_documents.append(doc)
        document_data = filtered_documents

    # Implementing pagination for nutrition data
    paginator = Paginator(document_data, 20)  # Show 20 nutrition entries per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    form = NotificationsForm()
    return render(request, 'frontend/techcare_data/notifications_table.html', {
        'form': form, 
        'page_obj': page_obj,
        'query': query
    })


def readBites_view(request):
    db = firestore.Client()
    
    # Fetching readBites data
    collection = db.collection("readBites")
    documents = collection.stream()
    document_data = [{'name': doc.id, 'data': doc.to_dict()} for doc in documents]
    
    query = request.GET.get('q')  # Get search query from request
    if query:
        # Perform search based on 'title' field (modify as per your Firestore structure)
        filtered_documents = []
        for doc in document_data:
            title = doc['data'].get('bite', '').lower()
            if query.lower() in title:
                filtered_documents.append(doc)
        document_data = filtered_documents

    # Fetching users data
    collection = db.collection("users")
    documents = collection.stream()
    users_data = [{'name': doc.id, 'data': doc.to_dict()} for doc in documents]

    # Fetching bites data
    collection = db.collection("bites")
    documents = collection.stream()
    bites_data = [{'name': doc.id, 'data': doc.to_dict()} for doc in documents]
    
    # Implementing pagination for readBites data
    paginator = Paginator(document_data, 20)  # Show 20 readBites entries per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    return render(request, 'frontend/techcare_data/readBites_table.html', {
        'document_data': document_data,
        'users_data': users_data,
        'bites_data': bites_data,
        'page_obj': page_obj,  # Pass page_obj to template for pagination
    })

def readStories_view(request):
    db = firestore.Client()
    
    # Fetching readStories data
    collection = db.collection("readStories")
    documents = collection.stream()
    document_data = [{'name': doc.id, 'data': doc.to_dict()} for doc in documents]
    
    # Fetching users data
    collection = db.collection("users")
    documents = collection.stream()
    users_data = [{'name': doc.id, 'data': doc.to_dict()} for doc in documents]

    # Fetching selfawarenessScenarios data
    collection = db.collection("selfawarenessScenarios")
    documents = collection.stream()
    selfawarenessScenarios_data = [{'name': doc.id, 'data': doc.to_dict()} for doc in documents]
    
    # Implementing pagination for readStories data
    paginator = Paginator(document_data, 20)  # Show 20 readStories entries per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    return render(request, 'frontend/techcare_data/readStories_table.html', {
        'document_data': document_data,
        'users_data': users_data,
        'selfawarenessScenarios_data': selfawarenessScenarios_data,
        'page_obj': page_obj,  # Pass page_obj to template for pagination
    })

def selfAwarnessBites_view(request):
    db = firestore.client()
    collection = db.collection("selfAwarnessBites")
    documents = collection.stream()
    document_data = [{'name': doc.id, 'data': doc.to_dict()} for doc in documents]

    query = request.GET.get('q')  # Get search query from request
    if query:
        # Perform search based on 'unit', 'name_en', or 'name_ar' fields
        query_lower = query.lower()
        filtered_documents = []
        for doc in document_data:
            data = doc['data']
            selfAwarnessBite = data.get('selfawarenessBiteText', '').lower()
            tags = data.get('tags', '').lower()
          
            if query_lower in selfAwarnessBite or query_lower in tags :
                filtered_documents.append(doc)
        document_data = filtered_documents
    form=selfAwarnessBitesForm()
    paginator = Paginator(document_data, 20)  
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, 'frontend/techcare_data/selfAwarnessBites_table.html', {'page_obj': page_obj,'query':query ,'form':form   })

def selfawareness_collection_view(request):
    db = firestore.client()
    collection = db.collection("selfawareness_collection")
    documents = collection.stream()
    document_data = [{'name': doc.id, 'data': doc.to_dict()} for doc in documents]
    
    query = request.GET.get('q')  # Get search query from request
    if query:
        # Perform search based on 'title' field (modify as per your Firestore structure)
        filtered_documents = []
        for doc in document_data:
            title = doc['data'].get('selfawareness_bite_text', '').lower()
            if query.lower() in title:
                filtered_documents.append(doc)
        document_data = filtered_documents

    paginator = Paginator(document_data, 20)  
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, 'frontend/techcare_data/selfawareness_collection_table.html', {'page_obj': page_obj,'query':query})

def suggestedActivities_view(request):
    db = firestore.client()
    collection = db.collection("suggestedActivities")
    documents = collection.stream()
    document_data = [{'name': doc.id, 'data': doc.to_dict()} for doc in documents]

    query = request.GET.get('q')  # Get search query from request
    if query:
        # Perform search based on 'title' field (modify as per your Firestore structure)
        filtered_documents = []
        for doc in document_data:
            title = doc['data'].get('type', '').lower()
            if query.lower() in title:
                filtered_documents.append(doc)
        document_data = filtered_documents

    collection = db.collection("users")
    documents = collection.stream()
    users_data = [{'name': doc.id, 'data': doc.to_dict()} for doc in documents]

    acitvities_collection = db.collection("activities")
    activites_documents = acitvities_collection.stream()
    activities_data = [{'name': doc.id, 'data': doc.to_dict()} for doc in activites_documents]

    paginator = Paginator(document_data, 20)  
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, 'frontend/techcare_data/suggestedActivities_table.html', {'page_obj': page_obj, 'users_data': users_data, 'activities_data': activities_data, 'query':query})

def suggestedBites_view(request):
    db = firestore.client()
    collection = db.collection("suggestedBites")
    documents = collection.stream()
    document_data = [{'name': doc.id, 'data': doc.to_dict()} for doc in documents]

    query = request.GET.get('q')  # Get search query from request
    if query:
        # Perform search based on 'title' field (modify as per your Firestore structure)
        filtered_documents = []
        for doc in document_data:
            title = doc['data'].get('name', '').lower()
            if query.lower() in title:
                filtered_documents.append(doc)
        document_data = filtered_documents

    collection = db.collection("users")
    documents = collection.stream()
    users_data = [{'name': doc.id, 'data': doc.to_dict()} for doc in documents]
    collection = db.collection("bites")
    documents = collection.stream()
    bites_data = [{'name': doc.id, 'data': doc.to_dict()} for doc in documents]
    collection = db.collection("selfAwarnessBites")
    documents = collection.stream()
    selfAwarnessBites_data = [{'name': doc.id, 'data': doc.to_dict()} for doc in documents]

    paginator = Paginator(document_data, 20)  
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)


    return render(request, 'frontend/techcare_data/suggestedBites_table.html', {'page_obj': page_obj,'document_data': document_data, 'users_data': users_data, 'bites_data': bites_data, 'selfAwarnessBites_data': selfAwarnessBites_data, 'query':query })





def suggestedInAppLinks_view(request):
    db = firestore.client()
    collection = db.collection("suggestedInAppLinks")
    documents = collection.stream()
    document_data = [{'name': doc.id, 'data': doc.to_dict()}
                     for doc in documents]

    # Search functionality
    query = request.GET.get('q')  # Get search query from request
    if query:
        # Perform search based on 'title' field (modify as per your Firestore structure)
        filtered_documents = []
        for doc in document_data:
            title = doc['data'].get('name', '').lower()
            if query.lower() in title:
                filtered_documents.append(doc)
        document_data = filtered_documents

    # Resolve references in document_data
    resolved_document_data = []
    for doc in document_data:
        doc_data = doc['data']
        # Example: resolving 'user' reference
        user_ref = doc_data.get('user')
        in_app_link_ref = doc_data.get('inAppLink')
        
        user_data = None
        in_app_link_data = None
        if isinstance(user_ref, firestore.DocumentReference):
            user_doc = user_ref.get()
            if user_doc.exists:
                user_data = user_doc.to_dict()
        
        if isinstance(in_app_link_ref, firestore.DocumentReference):
            in_app_link_doc = in_app_link_ref.get()
            if in_app_link_doc.exists:
                in_app_link_data = in_app_link_doc.to_dict()
        
        resolved_doc = {
            'name': doc['name'],
            'data': doc_data,
            'resolved_user': user_data,
            'resolved_inAppLink': in_app_link_data,
        }
        resolved_document_data.append(resolved_doc)
    
    # Fetch users data
    users_collection = db.collection("users")
    users_documents = users_collection.stream()
    users_data = [{'name': doc.id, 'data': doc.to_dict()} for doc in users_documents]
    
    # Fetch inAppLinks data
    in_app_links_collection = db.collection("inAppLinks")
    in_app_links_documents = in_app_links_collection.stream()
    inAppLinks_data = [{'name': doc.id, 'data': doc.to_dict()} for doc in in_app_links_documents]

    # Pagination
    paginator = Paginator(resolved_document_data, 20)  
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, 'frontend/techcare_data/suggestedInAppLinks_table.html', {
        'page_obj': page_obj, 
        'users_data': users_data, 
        'inAppLinks_data': inAppLinks_data, 
        'query': query
    })

# def suggestedInAppLinks_view(request):
#     db = firestore.client()
#     collection = db.collection("suggestedInAppLinks")
#     documents = collection.stream()
#     document_data = [{'name': doc.id, 'data': doc.to_dict()} for doc in documents]

#     # Search functionality
#     query = request.GET.get('q')  # Get search query from request
#     if query:
#         # Perform search based on 'title' field (modify as per your Firestore structure)
#         filtered_documents = []
#         for doc in document_data:
#             title = doc['data'].get('name', '').lower()
#             if query.lower() in title:
#                 filtered_documents.append(doc)
#         document_data = filtered_documents

#     # Resolve references in document_data
#     resolved_document_data = []
#     for doc in document_data:
#         doc_data = doc['data']
#         # Example: resolving 'user' reference
#         user_ref = doc_data.get('user')
#         in_app_link_ref = doc_data.get('inAppLink')
        
#         user_data = None
#         in_app_link_data = None
#         if isinstance(user_ref, firestore.DocumentReference):
#             user_doc = user_ref.get()
#             if user_doc.exists:
#                 user_data = user_doc.to_dict()
        
#         if isinstance(in_app_link_ref, firestore.DocumentReference):
#             in_app_link_doc = in_app_link_ref.get()
#             if in_app_link_doc.exists:
#                 in_app_link_data = in_app_link_doc.to_dict()
        
#         resolved_doc = {
#             'name': doc['name'],
#             'data': doc_data,
#             'resolved_user': user_data,
#             'resolved_inAppLink': in_app_link_data,
#         }
#         resolved_document_data.append(resolved_doc)
    
#     # Fetch users data
#     users_collection = db.collection("users")
#     users_documents = users_collection.stream()
#     users_data = [{'name': doc.id, 'data': doc.to_dict()} for doc in users_documents]
    
#     # Fetch inAppLinks data
#     in_app_links_collection = db.collection("inAppLinks")
#     in_app_links_documents = in_app_links_collection.stream()
#     inAppLinks_data = [{'name': doc.id, 'data': doc.to_dict()} for doc in in_app_links_documents]

#     # Pagination
#     paginator = Paginator(resolved_document_data, 20)  
#     page_number = request.GET.get('page')
#     page_obj = paginator.get_page(page_number)

#     return render(request, 'frontend/techcare_data/suggestedInAppLinks_table.html', {
#         'page_obj': page_obj, 
#         'users_data': users_data, 
#         'inAppLinks_data': inAppLinks_data, 
#         'query': query
#     })

def suggestedJournals_view(request):
    db = firestore.client()
    collection = db.collection("suggestedJournals")
    documents = collection.stream()
    document_data = [{'name': doc.id, 'data': doc.to_dict()} for doc in documents]

    query = request.GET.get('q')  # Get search query from request
    if query:
        # Perform search based on 'title' field (modify as per your Firestore structure)
        filtered_documents = []
        for doc in document_data:
            title = doc['data'].get('name', '').lower()
            if query.lower() in title:
                filtered_documents.append(doc)
        document_data = filtered_documents

    collection = db.collection("users")
    documents = collection.stream()
    users_data = [{'name': doc.id, 'data': doc.to_dict()} for doc in documents]
    collection = db.collection("journalPrompt")
    documents = collection.stream()
    journalPrompt_data = [{'name': doc.id, 'data': doc.to_dict()} for doc in documents]
    paginator = Paginator(document_data, 20)  
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, 'frontend/techcare_data/suggestedJournals_table.html', {'page_obj': page_obj, 'users_data': users_data, 'journalPrompt_data': journalPrompt_data, 'query':query})

def suggestedSelfAwarnessBites_view(request):
    db = firestore.client()
    collection = db.collection("suggestedSelfAwarnessBites")
    documents = collection.stream()
    document_data = [{'name': doc.id, 'data': doc.to_dict()} for doc in documents]

    query = request.GET.get('q')  # Get search query from request
    if query:
        # Perform search based on 'title' field (modify as per your Firestore structure)
        filtered_documents = []
        for doc in document_data:
            title = doc['data'].get('selfAwarnessBite', '').lower()
            if query.lower() in title:
                filtered_documents.append(doc)
        document_data = filtered_documents

    collection = db.collection("users")
    documents = collection.stream()
    users_data = [{'name': doc.id, 'data': doc.to_dict()} for doc in documents]
    collection = db.collection("selfAwarnessBites")
    documents = collection.stream()
    selfAwarnessBites_data = [{'name': doc.id, 'data': doc.to_dict()} for doc in documents]
    paginator = Paginator(document_data, 20)  
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, 'frontend/techcare_data/suggestedSelfAwarnessBites_table.html', {'page_obj': page_obj, 'users_data': users_data, 'selfAwarnessBites_data': selfAwarnessBites_data, 'query':query})
def suggestedWildCards_view(request):
    db = firestore.client()
    collection = db.collection("suggestedWildCards")
    documents = collection.stream()
    document_data = [{'name': doc.id, 'data': doc.to_dict()} for doc in documents]

    query = request.GET.get('q')  # Get search query from request
    if query:
        # Perform search based on 'title' field (modify as per your Firestore structure)
        filtered_documents = []
        for doc in document_data:
            title = doc['data'].get('name', '').lower()
            if query.lower() in title:
                filtered_documents.append(doc)
        document_data = filtered_documents

    collection = db.collection("users")
    documents = collection.stream()
    users_data = [{'name': doc.id, 'data': doc.to_dict()} for doc in documents]
    collection = db.collection("wildCard")
    documents = collection.stream()
    wildCard_data = [{'name': doc.id, 'data': doc.to_dict()} for doc in documents]

    paginator = Paginator(document_data, 20)  
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, 'frontend/techcare_data/suggestedWildCards_table.html', {'page_obj': page_obj, 'users_data': users_data, 'wildCard_data': wildCard_data, 'query':query})

def selfladder_view(request):
    db = firestore.client()
    collection = db.collection("selfLadder")
    documents = collection.stream()
    document_data = [{'name': doc.id, 'data': doc.to_dict()} for doc in documents]

    query = request.GET.get('q')  # Get search query from request
    if query:
        # Perform search based on 'title' field (modify as per your Firestore structure)
        filtered_documents = []
        for doc in document_data:
            title = doc['data'].get('type', '').lower()
            if query.lower() in title:
                filtered_documents.append(doc)
        document_data = filtered_documents
    form=SelfLadderForm()
    collection = db.collection("users")
    documents = collection.stream()
    users_data = [{'name': doc.id, 'data': doc.to_dict()} for doc in documents]

    paginator = Paginator(document_data, 20)  # Show 20 readStories entries per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
   
    return render(request, 'frontend/techcare_data/selfladder_table.html', {'page_obj': page_obj, 'users_data': users_data, 'query':query,'form':form})

def testTrivia_view(request):
    db = firestore.client()
    collection = db.collection("testTrivia")
    documents = collection.stream()
    document_data = [{'name': doc.id, 'data': doc.to_dict()} for doc in documents]

    query = request.GET.get('q')  # Get search query from request
    if query:
        # Perform search based on 'title' field (modify as per your Firestore structure)
        filtered_documents = []
        for doc in document_data:
            title = doc['data'].get('question', '').lower()
            if query.lower() in title:
                filtered_documents.append(doc)
        document_data = filtered_documents


    paginator = Paginator(document_data, 20)  # Show 20 readStories entries per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
 
    return render(request, 'frontend/techcare_data/testTrivia_table.html', {'page_obj': page_obj, 'query':query})

def wildCard_view(request):
    db = firestore.client()
    collection = db.collection("wildCard")
    documents = collection.stream()
    document_data = [{'name': doc.id, 'data': doc.to_dict()} for doc in documents]


    query = request.GET.get('q')  # Get search query from request
    if query:
        # Perform search based on 'title' field (modify as per your Firestore structure)
        filtered_documents = []
        for doc in document_data:
            title = doc['data'].get('content', '').lower()
            if query.lower() in title:
                filtered_documents.append(doc)
        document_data = filtered_documents
    form=wildCardForm()
    paginator = Paginator(document_data, 20)  # Show 20 readStories entries per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    return render(request, 'frontend/techcare_data/wildCard_table.html', {'page_obj': page_obj,'query':query, "form":form})


def activitiesdocument_detail(request, document_name):
    db = firestore.Client()
    collection = db.collection("activities")
    document_ref = collection.document(document_name)
    document_data = document_ref.get().to_dict()

    if not document_data:
        return render(request, 'frontend/techcare_data/document_not_found.html')

    return render(request, 'frontend/techcare_data/activitiesdocument.html', {'document_data': document_data})

def badges_view(request):
    db = firestore.client()
    collection = db.collection("badges")
    documents = collection.stream()
    document_data = [{'name': doc.id, 'data': doc.to_dict()} for doc in documents]
    
    query = request.GET.get('q')  # Get search query from request
    if query:
        # Perform search based on 'title' field (modify as per your Firestore structure)
        filtered_documents = []
        for doc in document_data:
            title = doc['data'].get('title', '').lower()
            if query.lower() in title:
                filtered_documents.append(doc)
        document_data = filtered_documents

    paginator = Paginator(document_data, 20)  # Show 20 badges per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    form = BadgesForm()  # Assuming BadgesForm is defined in forms.py
    
    return render(request, 'frontend/techcare_data/badges_table.html', {'form': form, 'page_obj': page_obj,'query':query})

def badgesdocument_detail(request, document_name):
    db = firestore.Client()
    collection = db.collection("badges")
    document_ref = collection.document(document_name)
    document_data = document_ref.get().to_dict()

    if not document_data:

        return render(request, 'frontend/techcare_data/document_not_found.html')

    return render(request, 'frontend/techcare_data/badgesdocument.html', {'document_data': document_data})

def biomarkers_view(request):

    # Fetch biomarkers data
    biomarkers_collection = db.collection("biomarkers")
    biomarkers_documents = biomarkers_collection.stream()
    biomarkers_data = [{'name': doc.id, 'data': doc.to_dict()} for doc in biomarkers_documents]

    query = request.GET.get('q')  # Get search query from request
    if query:
        # Perform search based on 'title' field (modify as per your Firestore structure)
        filtered_documents = []
        for doc in biomarkers_data:
            title = str(doc['data'].get('bloodGlucose', ''))
            if query in title:
                filtered_documents.append(doc)
        biomarkers_data = filtered_documents

    # Paginate biomarkers data
    biomarkers_paginator = Paginator(biomarkers_data, 20)  # Show 10 biomarkers per page
    biomarkers_page_number = request.GET.get('biomarkers_page')
    biomarkers_page_obj = biomarkers_paginator.get_page(biomarkers_page_number)

    form = BiomarkersForm()

    # Fetch users data
    users_collection = db.collection("users")
    users_documents = users_collection.stream()
    users_data = [{'name': doc.id, 'data': doc.to_dict()} for doc in users_documents]

    # Paginate users data
    users_paginator = Paginator(users_data, 20)  # Show 10 users per page
    users_page_number = request.GET.get('users_page')
    users_page_obj = users_paginator.get_page(users_page_number)

    return render(request, 'frontend/techcare_data/biomarkers_table.html', {
        'form': form,
        'biomarkers_page_obj': biomarkers_page_obj,
        'users_data': users_page_obj
    })

def biomarkersdocument_detail(request, document_name):
    db = firestore.Client()
    collection = db.collection("biomarkers")
    document_ref = collection.document(document_name)
    document_data = document_ref.get().to_dict()

    if not document_data:

        return render(request, 'frontend/techcare_data/document_not_found.html')

    return render(request, 'frontend/techcare_data/biomarkersdocument.html', {'document_data': document_data})





def assessmentQuestion_view(request):
    db = firestore.Client()

    # Fetch assessmentQuestion data
    assessment_collection = db.collection("assessmentQuestion")
    assessment_documents = assessment_collection.stream()
    assessment_data = [{'name': doc.id, 'data': doc.to_dict()} for doc in assessment_documents]
    query = request.GET.get('q')  # Get search query from request
    if query:
        # Perform search based on 'title' field (modify as per your Firestore structure)
        filtered_documents = []
        for doc in assessment_data:
            title = doc['data'].get('question', '').lower()
            if query.lower() in title:
                filtered_documents.append(doc)
        assessment_data = filtered_documents


    # Paginate assessmentQuestion data
    assessment_paginator = Paginator(assessment_data, 20)  # Show 10 assessment questions per page
    assessment_page_number = request.GET.get('assessment_page')
    assessment_page_obj = assessment_paginator.get_page(assessment_page_number)

    form = AssessmentQuestionForm()

    # Fetch majorAssessment data
    major_collection = db.collection("majorAssessment")
    major_documents = major_collection.stream()
    major_data = [{'name': doc.id, 'data': doc.to_dict()} for doc in major_documents]

    # Paginate majorAssessment data
    major_paginator = Paginator(major_data, 20)  # Show 10 major assessments per page
    major_page_number = request.GET.get('major_page')
    major_page_obj = major_paginator.get_page(major_page_number)

    return render(request, 'frontend/techcare_data/assessmentQuestion_table.html', {
        'form': form,
        'majorAssessment': major_page_obj,
        'assessment_page_obj': assessment_page_obj,
        'query': query
    })



def assessmentQuestiondocument_detail(request, document_name):
    db = firestore.Client()
    collection = db.collection("assessmentQuestion")
    document_ref = collection.document(document_name)
    document_data = document_ref.get().to_dict()

    if not document_data:

        return render(request, 'frontend/techcare_data/document_not_found.html')

    return render(request, 'frontend/techcare_data/assessmentQuestiondocument.html', {'document_data': document_data})



def categories_view(request):
 
    db = firestore.Client()
    collection = db.collection("categories")
    documents = collection.stream()
    document_data = [{'name': doc.id, 'data': doc.to_dict()} for doc in documents]
    
    query = request.GET.get('q')  # Get search query from request
    if query:
        # Perform search based on 'title' field (modify as per your Firestore structure)
        filtered_documents = []
        for doc in document_data:
            title = doc['data'].get('title', '').lower()
            if query.lower() in title:
                filtered_documents.append(doc)
        document_data = filtered_documents
    # Pagination
    paginator = Paginator(document_data, 20)  # Show 20 categories per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    form = CategoriesForm()
    return render(request, 'frontend/techcare_data/categories_table.html', {
        'page_obj': page_obj, 
        'form': form,
        'query': query
        
    })

def categoriesdocument_detail(request, document_name):
    db = firestore.Client()
    collection = db.collection("categories")
    document_ref = collection.document(document_name)
    document_data = document_ref.get().to_dict()

    if not document_data:

        return render(request, 'frontend/techcare_data/document_not_found.html')

    return render(request, 'frontend/techcare_data/categoriesdocument.html', {'document_data': document_data})

def feelings_view(request):
    db = firestore.Client()
    
    # Retrieve all documents from the "feelings" collection
    collection = db.collection("feelings")
    documents = collection.stream()
    document_data = [{'name': doc.id, 'data': doc.to_dict()} for doc in documents]
    
    # Retrieve all documents from the "user" collection
    collection2 = db.collection("user")
    documents2 = collection2.stream()
    feelings = [{'name': doc.id, 'data': doc.to_dict()} for doc in documents2]

    query = request.GET.get('q')  # Get search query from request
    if query:
        # Perform search based on 'joy' field being True or False
        filtered_documents = []
        query_lower = query.lower()
        for doc in document_data:
            joy = doc['data'].get('joy', None)
            if joy is not None:
                joy_str = 'true' if joy else 'false'
                if query_lower in joy_str:
                    filtered_documents.append(doc)
        document_data = filtered_documents

    # Pagination for feelings
    paginator = Paginator(document_data, 20)  # Show 20 feelings per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    form = FeelingsForm()
    return render(request, 'frontend/techcare_data/feelings_table.html', {
        'form': form,
        'feelings': feelings,
        'page_obj': page_obj,
        'query': query
    })

def feelingsdocument_detail(request, document_name):
    db = firestore.Client()
    collection = db.collection("feelings")
    document_ref = collection.document(document_name)
    document_data = document_ref.get().to_dict()

    if not document_data:

        return render(request, 'frontend/techcare_data/document_not_found.html')

    return render(request, 'frontend/techcare_data/feelingsdocument.html', {'document_data': document_data})



def inAppLinks_view(request):
    
    db = firestore.Client()
    collection = db.collection("inAppLinks")
    documents = collection.stream()
    document_data = [{'name': doc.id, 'data': doc.to_dict()} for doc in documents]

    query = request.GET.get('q')  # Get search query from request
    if query:
        # Perform search based on 'title' field (modify as per your Firestore structure)
        filtered_documents = []
        for doc in document_data:
            title = doc['data'].get('title', '').lower()
            if query.lower() in title:
                filtered_documents.append(doc)
        document_data = filtered_documents

    # Pagination for inAppLinks
    paginator = Paginator(document_data, 20)  # Show 20 inAppLinks per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    form = InAppLinksForm()
    return render(request, 'frontend/techcare_data/inAppLinks_table.html', {
        'form': form, 
        'page_obj': page_obj,
        'query': query
    })

def inAppLinksdocument_detail(request, document_name):
    db = firestore.Client()
    collection = db.collection("inAppLinks")
    document_ref = collection.document(document_name)
    document_data = document_ref.get().to_dict()

    if not document_data:

        return render(request, 'frontend/techcare_data/document_not_found.html')

    return render(request, 'frontend/techcare_data/inAppLinksdocument.html', {'document_data': document_data})

def inquiry_view(request):
    
    db = firestore.Client()
    
    # Fetching inquiry data
    collection = db.collection("inquiry").order_by("time",direction=firestore.Query.DESCENDING)
    documents = collection.stream()
    document_data = [{'name': doc.id, 'data': doc.to_dict()} for doc in documents]
    
    query = request.GET.get('q')  # Get search query from request
    if query:
        # Perform search based on 'title' field (modify as per your Firestore structure)
        filtered_documents = []
        for doc in document_data:
            title = str(doc['data'].get('question', '')).lower()
            if query.lower() in title:
                filtered_documents.append(doc)
        document_data = filtered_documents
    # Fetching users data
    collection = db.collection("users")
    documents = collection.stream()
    users_data = [{'name': doc.id, 'data': doc.to_dict()} for doc in documents]
    
    # Implementing pagination for inquiry data
    paginator = Paginator(document_data, 20)  # Show 20 inquiries per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    form = InquiryForm()
    return render(request, 'frontend/techcare_data/inquiry_table.html', {
        'form': form, 
        'page_obj': page_obj,
        'users_data': users_data,
        'query':query

    })

def inquirydocument_detail(request, document_name):
    db = firestore.Client()
    collection = db.collection("inquiry")
    document_ref = collection.document(document_name)
    document_data = document_ref.get().to_dict()

    if not document_data:
        return render(request, 'frontend/techcare_data/document_not_found.html')

    # Fetching users data
    collection = db.collection("users")
    documents = collection.stream()
    users_data = [{'name': doc.id, 'data': doc.to_dict()} for doc in documents]

    return render(request, 'frontend/techcare_data/inquirydocument.html', {
        'document_data': document_data,
        'document_name': document_name,
        'users_data': users_data
    })

def items_view(request):
 
    db = firestore.Client()
    
    # Fetching items data
    collection = db.collection("items")
    documents = collection.stream()
    document_data = [{'name': doc.id, 'data': doc.to_dict()} for doc in documents]
    
    query = request.GET.get('q')  # Get search query from request
    if query:
        # Perform search based on 'title' field (modify as per your Firestore structure)
        filtered_documents = []
        for doc in document_data:
            title = doc['data'].get('name', '').lower()
            if query.lower() in title:
                filtered_documents.append(doc)
        document_data = filtered_documents
    
    # Implementing pagination for items data
    paginator = Paginator(document_data, 20)  # Show 20 items per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    form = ItemsForm()
    return render(request, 'frontend/techcare_data/items_table.html', {
        'form': form, 
        'page_obj': page_obj,
        'query':query
    })

def itemsdocument_detail(request, document_name):
    db = firestore.Client()
    collection = db.collection("items")
    document_ref = collection.document(document_name)
    document_data = document_ref.get().to_dict()

    if not document_data:

        return render(request, 'frontend/techcare_data/document_not_found.html')

    return render(request, 'frontend/techcare_data/itemsdocument.html', {'document_data': document_data})





def journal_view(request):
    db = firestore.Client()

    # Initial fetch of all documents (or based on your initial requirements)
    collection = db.collection("journal")
    documents = collection.stream()
    document_data = [{'name': doc.id, 'data': doc.to_dict()} for doc in documents]

    # Search functionality
    query = request.GET.get('q')  # Get search query from request
    if query:
        # Perform search based on 'title' field (modify as per your Firestore structure)
        filtered_documents = []
        for doc in document_data:
            title = doc['data'].get('title', '').lower()
            if query.lower() in title:
                filtered_documents.append(doc)
        document_data = filtered_documents

    # Implementing pagination for search results
    paginator = Paginator(document_data, 20)  # Show 20 journal entries per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    form = JournalForm()
    return render(request, 'frontend/techcare_data/journal_table.html', {
        'form': form,
        'page_obj': page_obj,
        'query': query  # Pass query back to template for displaying search query
    })

def journaldocument_detail(request, document_name):
    db = firestore.Client()
    collection = db.collection("journal")
    document_ref = collection.document(document_name)
    document_data = document_ref.get().to_dict()

    if not document_data:

        return render(request, 'frontend/techcare_data/document_not_found.html')

    return render(request, 'frontend/techcare_data/journaldocument.html', {'document_data': document_data})

def journalPrompt_view(request):
   
    db = firestore.Client()
    
    # Fetching journalPrompt data
    collection = db.collection("journalPrompt")
    documents = collection.stream()
    document_data = [{'name': doc.id, 'data': doc.to_dict()} for doc in documents]
    
    query = request.GET.get('q')  # Get search query from request
    if query:
        # Perform search based on 'title' field (modify as per your Firestore structure)
        filtered_documents = []
        for doc in document_data:
            title = doc['data'].get('title', '').lower()
            if query.lower() in title:
                filtered_documents.append(doc)
        document_data = filtered_documents
        
    # Implementing pagination for journalPrompt data
    paginator = Paginator(document_data, 20)  # Show 20 journalPrompt entries per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    form = JournalForm()
    return render(request, 'frontend/techcare_data/journalPrompt_table.html', {
        'form': form, 
        'page_obj': page_obj,
        'query': query
    })
    
def journalPromptdocument_detail(request, document_name):
    db = firestore.Client()
    collection = db.collection("journalPrompt")
    document_ref = collection.document(document_name)
    document_data = document_ref.get().to_dict()

    if not document_data:

        return render(request, 'frontend/techcare_data/document_not_found.html')

    return render(request, 'frontend/techcare_data/journalPromptdocument.html', {'document_data': document_data})

def majorAssessment_view(request):

    db = firestore.Client()
    
    # Fetching majorAssessment data
    collection = db.collection("majorAssessment")
    documents = collection.stream()
    document_data = [{'name': doc.id, 'data': doc.to_dict()} for doc in documents]
    
    query = request.GET.get('q')  # Get search query from request
    if query:
        # Perform search based on 'title' field (modify as per your Firestore structure)
        filtered_documents = []
        for doc in document_data:
            title = doc['data'].get('title', '').lower()
            if query.lower() in title:
                filtered_documents.append(doc)
        document_data = filtered_documents

    # Implementing pagination for majorAssessment data
    paginator = Paginator(document_data, 20)  # Show 20 majorAssessment entries per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    form = MajorAssessmentForm()
    return render(request, 'frontend/techcare_data/majorAssessment_table.html', {
        'form': form, 
        'page_obj': page_obj,
        'query': query
    })

def majorAssessmentdocument_detail(request, document_name):
    db = firestore.Client()
    collection = db.collection("majorAssessment")
    document_ref = collection.document(document_name)
    document_data = document_ref.get().to_dict()

    if not document_data:

        return render(request, 'frontend/techcare_data/document_not_found.html')

    return render(request, 'frontend/techcare_data/majorAssessmentdocument.html', {'document_data': document_data})



def psychomarkers_view(request):
  
    db = firestore.Client()

    # Fetching psychomarkers data
    collection = db.collection("psychomarkers")
    documents = collection.stream()
    document_data = [{'name': doc.id, 'data': doc.to_dict()} for doc in documents]

    query = request.GET.get('q')  # Get search query from request
    if query:
        # Perform search based on 'depression' field
        filtered_documents = []
        for doc in document_data:
            depression = str(doc['data'].get('depression', '')).lower()
            if query.lower() in depression:
                filtered_documents.append(doc)
        document_data = filtered_documents

    # Implementing pagination for psychomarkers data
    paginator = Paginator(document_data, 20)  # Show 20 psychomarkers entries per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # Fetching users data
    collection = db.collection("users")
    documents = collection.stream()
    users_data = [{'name': doc.id, 'data': doc.to_dict()} for doc in documents]

    form = PsychomarkersForm()
    return render(request, 'frontend/techcare_data/psychomarkers_table.html', {
        'form': form,
        'page_obj': page_obj,
        'users_data': users_data,
        'query': query  # Pass the query back to the template for displaying in the search input

    })

def psychomarkersdocument_detail(request, document_name):
    db = firestore.Client()
    collection = db.collection("psychomarkers")
    document_ref = collection.document(document_name)
    document_data = document_ref.get().to_dict()

    if not document_data:

        return render(request, 'frontend/techcare_data/document_not_found.html')

    return render(request, 'frontend/techcare_data/psychomarkersdocument.html', {'document_data': document_data})

def scenarios_view(request):
   
    db = firestore.Client()
    collection = db.collection("scenarios")
    documents = collection.stream()
    document_data = [{'name': doc.id, 'data': doc.to_dict()} for doc in documents]
    form = ScenariosForm()

    query = request.GET.get('q')  # Get search query from request
    if query:
        # Perform search based on 'title' field (modify as per your Firestore structure)
        filtered_documents = []
        for doc in document_data:
            title = doc['data'].get('title', '').lower()
            if query.lower() in title:
                filtered_documents.append(doc)
        document_data = filtered_documents

    collection = db.collection("majorAssessment")
    documents = collection.stream()
    majorAssessment_data = [{'name': doc.id, 'data': doc.to_dict()} for doc in documents]

    collection = db.collection("activities")
    documents = collection.stream()
    activities_data = [{'name': doc.id, 'data': doc.to_dict()} for doc in documents]

    collection = db.collection("suggestedBites")
    documents = collection.stream()
    suggestedBites_data = [{'name': doc.id, 'data': doc.to_dict()} for doc in documents]

    collection = db.collection("suggestedJournals")
    documents = collection.stream()
    suggestedJournals_data = [{'name': doc.id, 'data': doc.to_dict()} for doc in documents]

    paginator = Paginator(document_data, 20)  # Show 20 scenarios entries per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    collection = db.collection("bites")
    documents = collection.stream()
    bites_data = [{'name': doc.id, 'data': doc.to_dict()} for doc in documents]

    return render(request, 'frontend/techcare_data/scenarios_table.html', {'page_obj': page_obj , 'form': form, 'majorAssessment_data': majorAssessment_data, 'activities_data': activities_data,'suggestedBites_data':suggestedBites_data,'suggestedJournals_data':suggestedJournals_data, 'bites_data':bites_data, 'query': query})

def scenariosdocument_detail(request, document_name):
    db = firestore.Client()
    collection = db.collection("scenarios")
    document_ref = collection.document(document_name)
    document_data = document_ref.get().to_dict()

    if not document_data:

        return render(request, 'frontend/techcare_data/document_not_found.html')

    return render(request, 'frontend/techcare_data/scenariosdocument.html', {'document_data': document_data})




def shortBite_view(request):
   
    db = firestore.Client()
    collection = db.collection("shortBite")
    documents = collection.stream()
    document_data = [{'name': doc.id, 'data': doc.to_dict()} for doc in documents]
    form = ShortBiteForm()

    query = request.GET.get('q')  # Get search query from request
    if query:
        # Perform search based on 'title' field (modify as per your Firestore structure)
        filtered_documents = []
        for doc in document_data:
            title = doc['data'].get('title', '').lower()
            if query.lower() in title:
                filtered_documents.append(doc)
        document_data = filtered_documents

    paginator = Paginator(document_data, 20)  # Show 20 scenarios entries per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, 'frontend/techcare_data/shortBite_table.html', {'page_obj': page_obj, 'form': form ,'query':query})
    
def shortBitedocument_detail(request, document_name):
    db = firestore.Client()
    collection = db.collection("shortBite")
    document_ref = collection.document(document_name)
    document_data = document_ref.get().to_dict()

    if not document_data:

        return render(request, 'frontend/techcare_data/document_not_found.html')

    return render(request, 'frontend/techcare_data/shortBitedocument.html', {'document_data': document_data})

def tags_view(request):
   
    db = firestore.Client()
    collection = db.collection("tags")
    documents = collection.stream()
    document_data = [{'name': doc.id, 'data': doc.to_dict()} for doc in documents]
    form = TagsForm()

    query = request.GET.get('q')  # Get search query from request
    if query:
        # Perform search based on 'title' field (modify as per your Firestore structure)
        filtered_documents = []
        for doc in document_data:
            title = doc['data'].get('title', '').lower()
            if query.lower() in title:
                filtered_documents.append(doc)
        document_data = filtered_documents

    paginator = Paginator(document_data, 20)  # Show 20 scenarios entries per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    return render(request, 'frontend/techcare_data/tags_table.html', {'page_obj': page_obj, 'form': form,'query':query})

def tagsdocument_detail(request, document_name):
    db = firestore.Client()
    collection = db.collection("tags")
    document_ref = collection.document(document_name)
    document_data = document_ref.get().to_dict()

    if not document_data:

        return render(request, 'frontend/techcare_data/document_not_found.html')

    return render(request, 'frontend/techcare_data/tagsdocument.html', {'document_data': document_data})


def trivia_view(request):
  
    db = firestore.Client()
    collection = db.collection("trivia")
    documents = collection.stream()
    document_data = [{'name': doc.id, 'data': doc.to_dict()} for doc in documents]
    form = TriviaForm()

    query = request.GET.get('q')  # Get search query from request
    if query:
        # Perform search based on 'title' field (modify as per your Firestore structure)
        filtered_documents = []
        for doc in document_data:
            title = doc['data'].get('question', '').lower()
            if query.lower() in title:
                filtered_documents.append(doc)
        document_data = filtered_documents


    paginator = Paginator(document_data, 20)  # Show 20 scenarios entries per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    return render(request, 'frontend/techcare_data/trivia_table.html', {'page_obj': page_obj, 'form': form,'query':query})

def triviadocument_detail(request, document_name):
    db = firestore.Client()
    collection = db.collection("trivia")
    document_ref = collection.document(document_name)
    document_data = document_ref.get().to_dict()

    if not document_data:

        return render(request, 'frontend/techcare_data/document_not_found.html')

    return render(request, 'frontend/techcare_data/triviadocument.html', {'document_data': document_data})


def users_view(request):
    
    db = firestore.Client()
    collection = db.collection("users")
    documents = collection.stream()
    document_data = [{'name': doc.id, 'data': doc.to_dict()} for doc in documents]
    form = UsersForm()

    query = request.GET.get('q')  # Get search query from request
    if query:
        # Perform search based on 'title' field (modify as per your Firestore structure)
        filtered_documents = []
        for doc in document_data:
            title = doc['data'].get('email', '').lower()
            if query.lower() in title:
                filtered_documents.append(doc)
        document_data = filtered_documents

    paginator = Paginator(document_data, 20)  # Show 20 scenarios entries per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    return render(request, 'frontend/techcare_data/users_table.html', {'page_obj': page_obj, 'form': form, 'query':query})


def usersdocument_detail(request, document_name):
    db = firestore.Client()
    collection = db.collection("users")
    document_ref = collection.document(document_name)
    document_data = document_ref.get().to_dict()

    if not document_data:

        return render(request, 'frontend/techcare_data/document_not_found.html')

    return render(request, 'frontend/techcare_data/usersdocument.html', {'document_data': document_data})




def sidebar(request):
    firestore_collections = get_firestore_collections()
    return render(request, 'sidebar.html', {'firestore_collections': firestore_collections}) 

def create_tags(request):
    if request.method == 'POST':
        form = TagsForm(request.POST)
        if form.is_valid():
            data = {
                'title': form.cleaned_data['title'],
                'description': form.cleaned_data['description'],   
            }
            db.collection("tags").document().set(data)  
            messages.success(request, 'Success Creating Tags.')
            return redirect('tags_view')
        else:
             messages.error(request, 'Error creating Tags. Please check your input.')
             return redirect('tags_view')

def create_activities(request):
    if request.method == 'POST':
        form = ActivitiesForm(request.POST)

        db = firestore.client()
        
        

        tags_path=request.POST.get('tags')

        tags_ref=get_document_reference(tags_path)
        if form.is_valid():
            data = {
                'tags': tags_ref,
                'description': form.cleaned_data['description'],
                'title': form.cleaned_data['title'],
                'duration': form.cleaned_data['duration'],
            }
            
            db.collection("activities").document().set(data)
            messages.success(request, 'Success Creating Activity.')
            return redirect('activities_view')
        else:
             messages.error(request, 'Error creating activity. Please check your input.')
             return redirect('activities_view')
        

def create_assessmentQuestion(request):
    if request.method == 'POST':
        form = AssessmentQuestionForm(request.POST)
        if form.is_valid():
            data = {
                'majorAssessment': request.POST.get('majorAssessment'),   
                'max': form.cleaned_data['max'],
                'order': form.cleaned_data['order'],  
                'points': form.cleaned_data['points'],  
                'question': form.cleaned_data['question'],  
            }
            db = firestore.client()
            db.collection("assessmentQuestion").document().set(data)
            messages.success(request, 'Successfully created Assessment Question.')  
            return redirect('assessmentQuestion_view') 
        else:
            messages.error(request, 'Error creating Assessment Question. Please check your input.')
            return redirect('assessmentQuestion_view') 

import csv
import io
def create_assets(request):
    if request.method == 'POST':
        form = AssetsForm(request.POST, request.FILES)
        if form.is_valid():
            # Save form data to Firestore
            data = {
                'assetsType': form.cleaned_data['assetsType'],
                'label': form.cleaned_data['label'],
                'name': form.cleaned_data['name'],
                'path': form.cleaned_data['path'],
            }
            db = firestore.client()
            db.collection("assets").document().set(data)
            messages.success(request, 'Successfully created asset.')

            # Process the uploaded CSV file
            csv_file = request.FILES.get('csv_file')
            if csv_file:
                data_set = csv_file.read().decode('UTF-8')
                io_string = io.StringIO(data_set)
                reader = csv.reader(io_string, delimiter=',', quotechar='|')
                
                for row in reader:
                    # Skip the header row if present
                    if reader.line_num == 1:
                        continue
                    
                    assets_data = {
                        'assetsType': row[0],
                        'name': row[1],
                        'label': row[2],
                        'path': row[3],
                    }
                    db.collection("assets").document().set(assets_data)
                messages.success(request, 'CSV file processed successfully.')

            return redirect('assets_view')
        else:
            messages.error(request, 'Error creating asset. Please check your input.')
            return redirect('assets_view')
    else:
        form = AssetsForm()

    return render(request, 'frontend/techcare_data/assets_form.html', {'form': form})

def create_badges(request):
       if request.method == 'POST':
         form = BadgesForm(request.POST)
         if form.is_valid():
            data = {
                'title': form.cleaned_data['title'],   
            }
            db.collection("badges").document().set(data)  
            messages.success(request, 'Successfully created Badges View.')  
            return redirect('badges_view') 
         else:
            messages.error(request, 'Error creating Badges View. Please check your input.')
            return redirect('badges_view') 

def create_biomarkers(request):
    if request.method == 'POST':
        form = BiomarkersForm(request.POST)
        if form.is_valid():
            data = {
                'dailyActivity': form.cleaned_data['dailyActivity'],
                'dailyCarbs': form.cleaned_data['dailyCarbs'],
                'sleepQuality': form.cleaned_data['sleepQuality'],
                'time': form.cleaned_data['time'],
                'user': request.POST.get('user'),
                'weeklyActivity': form.cleaned_data['weeklyActivity'],
                'weight': form.cleaned_data['weight'],
                'FBS': form.cleaned_data['FBS'],
                'HBA1c': form.cleaned_data['HBA1c'],
                'bloodGlucose': form.cleaned_data['bloodGlucose'],
                'bloodGlucoseType': form.cleaned_data['bloodGlucoseType'],
            }
            db.collection("biomarkers").document().set(data)
            messages.success(request, 'Successfully created Biomarkers View.')  
            return redirect('biomarkers_view') 
        else:
            messages.error(request, 'Error creating Biomarkers View. Please check your input.')
            return redirect('biomarkers_view') 




   
def create_categories(request):
    if request.method == 'POST':
        form = CategoriesForm(request.POST)
        if form.is_valid():
            data = {
                'description': form.cleaned_data['description'], 
                'title': form.cleaned_data['title'],   
            }
            db.collection("categories").document().set(data)  
            messages.success(request, 'Successfully created Categories.')
            return redirect('categories_view') 
        else:
            messages.error(request, 'Error creating Categories. Please check your input.')
            return redirect('categories_view') 

   
def create_feelings(request):
       if request.method == 'POST':
         form = FeelingsForm(request.POST)
         if form.is_valid():
            data = {
                'anger': form.cleaned_data['anger'], 
                'fear': form.cleaned_data['fear'], 
                'happiness': form.cleaned_data['happiness'], 
                'joy': form.cleaned_data['joy'],   
                'love': form.cleaned_data['love'], 
               'sadness': form.cleaned_data['sadness'], 
               'shame': form.cleaned_data['shame'], 
                'strength': form.cleaned_data['strength'],   
                # 'time': form.cleaned_data['time'], 

            }
            db.collection("feelings").document().set(data)  
            messages.success(request, 'Successfully created Feelings.')
            return redirect('feelings_view') 
         else:
            messages.error(request, 'Error creating Feelings. Please check your input.')
            return redirect('feelings_view') 

def create_inAppLinks(request):
    if request.method == 'POST':
        form = InAppLinksForm(request.POST)
        if form.is_valid():
            data = {
                'title': form.cleaned_data['title'],
                'description': form.cleaned_data['description'],
                'order': form.cleaned_data['order'],
                'type': form.cleaned_data['type'],
                
            }
            db.collection("inAppLinks").document().set(data)
            messages.success(request, 'Successfully created inAppLinks.')  
            return redirect('inAppLinks_view') 
        else:
            messages.error(request, 'Error creating inAppLinks. Please check your input.')
            return redirect('inAppLinks_view')

        
    return render(request, 'frontend/techcare_data/create_inAppLinks.html', {'form': form})
from django.contrib.auth import get_user_model

from django.contrib.auth import get_user_model

def create_inquiry(request):
    if request.method == 'POST':
        form = InquiryForm(request.POST)
        if form.is_valid():

            data = {
                'question': form.cleaned_data['question'],
                'answer': form.cleaned_data['answer'],
                'topic': form.cleaned_data['topic'],
                'user': request.POST.get('user_id'),
                'time': timezone.now(),
            }

            # Save the data to Firestore
            db.collection("inquiry").document().set(data)  

            messages.success(request, 'Successfully created Inquiry.') 
            return redirect('inquiry_view')
        else:
            messages.error(request, 'Error creating Inquiry. Please check your input.')
            return redirect('inquiry_view')



def create_journal(request):
    if request.method == 'POST':
        form = JournalForm(request.POST)
        if form.is_valid():
            data = {
                'title': form.cleaned_data['title'],
                'description': form.cleaned_data['description'],
              
            }
            db.collection("journal").document().set(data)  
            messages.success(request, 'Successfully created Journal.') 
            return redirect('journal_view')
        else:
            messages.error(request, 'Error creating Journal. Please check your input.')
            return redirect('journal_view')


def create_journalPrompt(request):
    if request.method == 'POST':
        form = JournalPromptForm(request.POST)
        if form.is_valid():
            data = {
                'title': form.cleaned_data['title'],
              
            }
            db.collection("journalPrompt").document().set(data)  
            messages.success(request, 'Successfully created JournalPrompt.') 
            return redirect('journalPrompt_view')
        else:
            messages.error(request, 'Error creating JournalPrompt. Please check your input.')
            return redirect('journalPrompt_view')



def create_majorAssessment(request):
    if request.method == 'POST':
        form = MajorAssessmentForm(request.POST)
        if form.is_valid():
            data = {
                'title': form.cleaned_data['title'],
                'numberOfQuestions': form.cleaned_data['numberOfQuestions'],
                'description': form.cleaned_data['description'],
                'order': form.cleaned_data['order'],  # Removed the space before 'order'
            }
            db.collection("majorAssessment").document().set(data)  
            messages.success(request, 'Successfully created MajorAssessment.') 
            return redirect('majorAssessment_view')
        else:
            messages.error(request, 'Error creating MajorAssessment. Please check your input.')
            return redirect('majorAssessment_view')

def create_notifications(request):
    if request.method == 'POST':
        form = NotificationsForm(request.POST)
        if form.is_valid():
            data = {
                'title': form.cleaned_data['title'],
                'body': form.cleaned_data['body'],
                'time': form.cleaned_data['time'],
            }
            db.collection("notifications").document().set(data)  
            messages.success(request, 'Successfully created notification.') 
            return redirect('notifications_view')
        else:
            messages.error(request, 'Error creating notification. Please check your input.')
            return redirect('notifications_view')

def create_psychomarkers(request):
    if request.method == 'POST':
        form = PsychomarkersForm(request.POST)
        if form.is_valid():
            # user_id = form.cleaned_data['user'].id
            data = {
                'time': datetime.now(),
                'user': request.POST.get('user_id'),  # Assuming 'user' is a ForeignKey in your model, use the user_id
                'depression': form.cleaned_data['depression'],
            }
            db.collection("psychomarkers").document().set(data)
            messages.success(request, 'Successfully created Psychomarkers.') 
            return redirect('psychomarkers_view')
        else:
            messages.error(request, 'Error creating Psychomarkers. Please check your input.')
            return redirect('psychomarkers_view')

def create_scenarios(request):
    if request.method == 'POST':
        form = ScenariosForm(request.POST)
        if form.is_valid():
            data = {
                'PositiveCorrectionAlternative': form.cleaned_data['PositiveCorrectionAlternative'],
                'actionreply': form.cleaned_data['actionreply'],
                'correction': form.cleaned_data['correction'],
                'positiveActionReply': form.cleaned_data['positiveActionReply'],
                'title': form.cleaned_data['title'],
            }
            db.collection("scenarios").document().set(data)
            messages.success(request, 'Successfully created Scenarios.') 
            return redirect('scenarios_view')
        else:
            messages.error(request, 'Error creating Scenarios. Please check your input.')
            return redirect('scenarios_view')

def create_selfAwarenessScenarios(request):
    if request.method == 'POST':
        form = SelfAwarenessScenariosForm(request.POST)
        if form.is_valid():
            data = {
                'correction1from0to2': form.cleaned_data['correction1from0to2'],
                'correction2from3to5': form.cleaned_data['correction2from3to5'],
                'interactiveStatement': form.cleaned_data['interactiveStatement'],
                'recommendation1': form.cleaned_data['recommendation1'],
                'recommendation2': form.cleaned_data['recommendation2'],
                'scenarioID': form.cleaned_data['scenarioID'],
                'story': form.cleaned_data['story'],
                'storyTitle': form.cleaned_data['storyTitle'],

            }
            db.collection("selfawarenessScenarios").document().set(data)
            messages.success(request, 'Successfully created selfawarenessScenario.') 
            return redirect('selfawarenessScenarios_view')
        else:
            messages.error(request, 'Error creating Scenarios. Please check your input.')
            return redirect('selfawarenessScenarios_view')
        

def create_selfLadder(request):
    if request.method == 'POST':
        form = SelfLadderForm(request.POST)
        if form.is_valid():
            data = {
                'user': request.POST.get('user'),   
                'type': form.cleaned_data['type'],
                'time': form.cleaned_data['time'],
                

            }
            db.collection("selfLadder").document().set(data)
            messages.success(request, 'Successfully created selfLadder.') 
            return redirect('selfladder_view')
        else:
            messages.error(request, 'Error creating Scenarios. Please check your input.')
            return redirect('selfladder_view')
        

def create_selfAwarnessBites(request):
    if request.method == 'POST':
        form = selfAwarnessBitesForm(request.POST)
        if form.is_valid():
            data = {
                
                'scenarioID': form.cleaned_data['scenarioID'],
                'tags': form.cleaned_data['tags'],
                'selfawarenessBiteTitle': form.cleaned_data['selfawarenessBiteTitle'],
                'selfawarenessBiteText': form.cleaned_data['selfawarenessBiteText'],
                
            }
            db.collection("selfAwarnessBites").document().set(data)
            messages.success(request, 'Successfully created selfAwarnessBites.') 
            return redirect('selfAwarnessBites_view')
        else:
            messages.error(request, 'Error creating selfAwarnessBites. Please check your input.')
            return redirect('selfAwarnessBites_view')        
        
def create_wildCard(request):
    if request.method == 'POST':
        form = wildCardForm(request.POST)
        if form.is_valid():
            data = {
                
                'content': form.cleaned_data['content'],
                
                
            }
            db.collection("wildCard").document().set(data)
            messages.success(request, 'Successfully created wildCard.') 
            return redirect('wildCard_view')
        else:
            messages.error(request, 'Error creating wildCard. Please check your input.')
            return redirect('wildCard_view')        
                
    
def create_shortBite(request):
    if request.method == 'POST':
        form = ShortBiteForm(request.POST)
        if form.is_valid():
            data = {
                'order': form.cleaned_data['order'],
                'title': form.cleaned_data['title'],
            }
            db.collection("shortBite").document().set(data)
            messages.success(request, 'Successfully ShortBite Scenarios.') 
            return redirect('shortBite_view')
        else:
            messages.error(request, 'Error creating ShortBite. Please check your input.')
            return redirect('shortBite_view')


def create_trivia(request):
    if request.method == 'POST':
        form = TriviaForm(request.POST)
        if form.is_valid():
            data = {
                'CBT_points': form.cleaned_data['CBT_points'],
                'answer_1': form.cleaned_data['answer_1'],
                'answer_2': form.cleaned_data['answer_2'],
                'answer_3': form.cleaned_data['answer_3'],
                'correct_answer': form.cleaned_data['correct_answer'],
                'description': form.cleaned_data['description'],
                'explanation': form.cleaned_data['explanation'],
                'image': form.cleaned_data['image'],
                'learning_points': form.cleaned_data['learning_points'],
                'order': form.cleaned_data['order'],
                'question': form.cleaned_data['question'],
                'topic': form.cleaned_data['topic'],
                
            }
            db.collection("trivia").document().set(data)
            messages.success(request, 'Successfully Trivia Scenarios.') 
            return redirect('trivia_view')
        else:
            messages.error(request, 'Error creating Trivia. Please check your input.')
            return redirect('trivia_view')


from datetime import datetime

def create_users(request):
    if request.method == 'POST':
        form = UsersForm(request.POST)
        if form.is_valid():
            data = {
                'uid': form.cleaned_data['uid'],
                'gender': form.cleaned_data['gender'],
                'email': form.cleaned_data['email'],
                'display_name': form.cleaned_data['display_name'],
                'created_time': datetime.now(),  
            }
            db.collection("users").document().set(data)
            messages.success(request, 'Successfully Create Nutrition.') 
            return redirect('users_view')
        else:
            messages.error(request, 'Error creating Nutrition. Please check your input.')
            return redirect('users_view')
        
def create_nutrition(request):
    if request.method == 'POST':
        form = NutritionForm(request.POST)
        if form.is_valid():
            data = {
                'carbContent': request.POST.get('carbContent'),
                'name_ar': request.POST.get('name_ar'),
                'name_en': request.POST.get('name_en'),
                'portion': request.POST.get('portion'),
                'fatContent':request.POST.get('fatContent'),
                'proteinContent': request.POST.get('proteinContent'),
                'totalCalories': request.POST.get('totalCalories'),
                'weight': request.POST.get('weight'),
                'unit':request.POST.get('unit'),
                'food_category':request.POST.get('food_category'),
                'Class':request.POST.get('Class'),

                               
            }
            db.collection("nutrition").document().set(data)
            messages.success(request, 'Successfully Users Scenarios.') 
            return redirect('nutrition_view')
        else:
            messages.error(request, 'Error creating Nutrition. Please check your input.')
            return redirect('nutrition_view')        

        


# def create_items(request):
#     if request.method == 'POST':
#         form = ItemsForm(request.POST)
#         if form.is_valid():
#             item_instance = items(
#                 data_title=form.cleaned_data['data_title'],
#                 name=form.cleaned_data['name'],
#                 data_categories=form.cleaned_data['data_categories'],
               
#             )
#             item_instance.save()
            
#             return redirect('items_view')
#     else:
#         form = ItemsForm()

#     return render(request, 'frontend/techcare_data/create_items.html', {'form': form})
def create_items(request):
    if request.method == 'POST':
        form = ItemsForm(request.POST)
        if form.is_valid():
            # Extract form data
            name = form.cleaned_data['data']
            data_title = form.cleaned_data['title']
            data_categories = form.cleaned_data['categories']

            # Create a dictionary for the data
            data = {
                'name': name,
                'data': {
                    'title': data_title,
                    'categories': data_categories,
                }
            }

            # Add the data to the Firestore collection
            db.collection("items").add(data)
            messages.success(request, 'Successfully created Items.')  
            return redirect('items_view') 
        else:
            messages.error(request, 'Error creating Items. Please check your input.')
            return redirect('items_view') 
    return render(request, 'frontend/techcare_data/activitiesdocument.html', {'document_data': document_data})

        
def selfAwarnessBites_delete(request, document_name):
    db = firestore.Client()
    collection = db.collection("selfAwarnessBites")
    document_ref = collection.document(document_name)
    document = document_ref.get()

    if not document.exists:
        return render(request, 'frontend/techcare_data/document_not_found.html')

    document_ref.delete()

    messages.success(request, 'selfAwarnessBites deleted successfully!')
    return redirect('selfAwarnessBites_view')


def nutrition_delete(request, document_name):
    db = firestore.Client()
    collection = db.collection("nutrition")
    document_ref = collection.document(document_name)
    document = document_ref.get()

    if not document.exists:
        return render(request, 'frontend/techcare_data/document_not_found.html')

    document_ref.delete()

    messages.success(request, 'nutrition deleted successfully!')
    return redirect('nutrition_view')




def activities_delete(request, document_name):
    db = firestore.Client()
    collection = db.collection("activities")
    document_ref = collection.document(document_name)
    document = document_ref.get()

    if not document.exists:
        return render(request, 'frontend/techcare_data/document_not_found.html')

    document_ref.delete()

    messages.success(request, 'Activity deleted successfully!')
    return redirect('activities_view')
    
def assessmentQuestion_delete(request, document_name):
    db = firestore.Client()
    collection = db.collection("assessmentQuestion")
    document_ref = collection.document(document_name)
    document = document_ref.get()

    if not document.exists:
        return render(request, 'frontend/techcare_data/document_not_found.html')

    document_ref.delete()

    messages.success(request, 'AssessmentQuestion deleted successfully!')
    return redirect('assessmentQuestion_view')


def assets_delete(request, document_name):
    db = firestore.Client()
    collection = db.collection("assets")
    document_ref = collection.document(document_name)
    document = document_ref.get()

    if not document.exists:
        return render(request, 'frontend/techcare_data/document_not_found.html')

    document_ref.delete()

    messages.success(request, 'assets deleted successfully!')
    return redirect('assets_view')

def badges_delete(request, document_name):
    db = firestore.Client()
    collection = db.collection("badges")
    document_ref = collection.document(document_name)
    document = document_ref.get()

    if not document.exists:
        return render(request, 'frontend/techcare_data/document_not_found.html')

    document_ref.delete()

    messages.success(request, 'Badges deleted successfully!')
    return redirect('badges_view')


def biomarkers_delete(request, document_name):
    db = firestore.Client()
    collection = db.collection("biomarkers_delete")
    document_ref = collection.document(document_name)
    document = document_ref.get()

    if not document.exists:
        return render(request, 'frontend/techcare_data/document_not_found.html')

    document_ref.delete()

    messages.success(request, 'Biomarkers deleted successfully!')
    return redirect('biomarkers_view')

def bites_delete(request, document_name):
    db = firestore.Client()
    collection = db.collection("bites")
    document_ref = collection.document(document_name)
    document = document_ref.get()

    if not document.exists:
        return render(request, 'frontend/techcare_data/document_not_found.html')

    document_ref.delete()

    messages.success(request, 'Bites deleted successfully!')
    return redirect('bites_view')

def selfawarenessScenarios_delete(request, document_name):
    db = firestore.Client()
    collection = db.collection("selfawarenessScenarios")
    document_ref = collection.document(document_name)
    document = document_ref.get()

    if not document.exists:
        return render(request, 'frontend/techcare_data/document_not_found.html')

    document_ref.delete()

    messages.success(request, 'selfawarenessScenarios_delete deleted successfully!')
    return redirect('selfawarenessScenarios_view')

def selfLadder_delete(request, document_name):
    db = firestore.Client()
    collection = db.collection("selfLadder")
    document_ref = collection.document(document_name)
    document = document_ref.get()

    if not document.exists:
        return render(request, 'frontend/techcare_data/document_not_found.html')

    document_ref.delete()

    messages.success(request, 'selfLadder deleted successfully!')
    return redirect('selfladder_view')


def wildCard_delete(request, document_name):
    db = firestore.Client()
    collection = db.collection("wildCard")
    document_ref = collection.document(document_name)
    document = document_ref.get()

    if not document.exists:
        return render(request, 'frontend/techcare_data/document_not_found.html')

    document_ref.delete()

    messages.success(request, 'wildCard deleted successfully!')
    return redirect('wildCard_view')

def selfawarenessBites_delete(request, document_name):
    db = firestore.Client()
    collection = db.collection("selfAwarnessBites")
    document_ref = collection.document(document_name)
    document = document_ref.get()

    if not document.exists:
        return render(request, 'frontend/techcare_data/document_not_found.html')

    document_ref.delete()

    messages.success(request, 'selfawarenessBites deleted successfully!')
    return redirect('selfAwarnessBites_view')

def selfawarenessCollection_delete(request, document_name):
    db = firestore.Client()
    collection = db.collection("selfawareness_collection")
    document_ref = collection.document(document_name)
    document = document_ref.get()

    if not document.exists:
        return render(request, 'frontend/techcare_data/document_not_found.html')

    document_ref.delete()

    messages.success(request, 'selfawarenessCollection deleted successfully!')
    return redirect('selfawareness_collection_view')

def categories_delete(request, document_name):
    db = firestore.Client()
    collection = db.collection("categories")
    document_ref = collection.document(document_name)
    document = document_ref.get()

    if not document.exists:
        return render(request, 'frontend/techcare_data/document_not_found.html')

    document_ref.delete()

    messages.success(request, 'Categories deleted successfully!')
    return redirect('categories_view')

def feelings_delete(request, document_name):
    db = firestore.Client()
    collection = db.collection("feelings")
    document_ref = collection.document(document_name)
    document = document_ref.get()

    if not document.exists:
        return render(request, 'frontend/techcare_data/document_not_found.html')

    document_ref.delete()

    messages.success(request, 'Feelings deleted successfully!')
    return redirect('feelings_view')

def inAppLinks_delete(request, document_name):
    db = firestore.Client()
    collection = db.collection("inAppLinks")
    document_ref = collection.document(document_name)
    document = document_ref.get()

    if not document.exists:
        return render(request, 'frontend/techcare_data/document_not_found.html')

    document_ref.delete()

    messages.success(request, 'InAppLinks deleted successfully!')
    return redirect('inAppLinks_view')

def inquiry_delete(request, document_name):
    db = firestore.Client()
    collection = db.collection("inquiry")
    document_ref = collection.document(document_name)
    document = document_ref.get()

    if not document.exists:
        return render(request, 'frontend/techcare_data/document_not_found.html')

    document_ref.delete()

    messages.success(request, 'Inquiry deleted successfully!')
    return redirect('inquiry_view')

def items_delete(request, document_name):
    db = firestore.Client()
    collection = db.collection("items")
    document_ref = collection.document(document_name)
    document = document_ref.get()

    if not document.exists:
        return render(request, 'frontend/techcare_data/document_not_found.html')

    document_ref.delete()

    messages.success(request, 'Items deleted successfully!')
    return redirect('items_view')

def journal_delete(request, document_name):
    db = firestore.Client()
    collection = db.collection("journal")
    document_ref = collection.document(document_name)
    document = document_ref.get()

    if not document.exists:
        return render(request, 'frontend/techcare_data/document_not_found.html')

    document_ref.delete()

    messages.success(request, 'Journal deleted successfully!')
    return redirect('journal_view')

def journalPrompt_delete(request, document_name):
    db = firestore.Client()
    collection = db.collection("journalPrompt")
    document_ref = collection.document(document_name)
    document = document_ref.get()

    if not document.exists:
        return render(request, 'frontend/techcare_data/document_not_found.html')

    document_ref.delete()

    messages.success(request, 'JournalPrompt deleted successfully!')
    return redirect('journalPrompt_view')

def majorAssessment_delete(request, document_name):
    db = firestore.Client()
    collection = db.collection("majorAssessment")
    document_ref = collection.document(document_name)
    document = document_ref.get()

    if not document.exists:
        return render(request, 'frontend/techcare_data/document_not_found.html')

    document_ref.delete()

    messages.success(request, 'MajorAssessment deleted successfully!')
    return redirect('majorAssessment_view')

def psychomarkers_delete(request, document_name):
    db = firestore.Client()
    collection = db.collection("psychomarkers")
    document_ref = collection.document(document_name)
    document = document_ref.get()

    if not document.exists:
        return render(request, 'frontend/techcare_data/document_not_found.html')

    document_ref.delete()

    messages.success(request, 'Psychomarkers deleted successfully!')
    return redirect('psychomarkers_view')

def scenarios_delete(request, document_name):
    db = firestore.Client()
    collection = db.collection("scenarios")
    document_ref = collection.document(document_name)
    document = document_ref.get()

    if not document.exists:
        return render(request, 'frontend/techcare_data/document_not_found.html')

    document_ref.delete()

    messages.success(request, 'Scenarios deleted successfully!')
    return redirect('scenarios_view')

def shortBite_delete(request, document_name):
    db = firestore.Client()
    collection = db.collection("shortBite")
    document_ref = collection.document(document_name)
    document = document_ref.get()

    if not document.exists:
        return render(request, 'frontend/techcare_data/document_not_found.html')

    document_ref.delete()

    messages.success(request, 'ShortBite deleted successfully!')
    return redirect('shortBite_view')

def tags_delete(request, document_name):
    db = firestore.Client()
    collection = db.collection("tags")
    document_ref = collection.document(document_name)
    document = document_ref.get()

    if not document.exists:
        return render(request, 'frontend/techcare_data/document_not_found.html')

    document_ref.delete()

    messages.success(request, 'Tags deleted successfully!')
    return redirect('tags_view')

def trivia_delete(request, document_name):
    db = firestore.Client()
    collection = db.collection("trivia")
    document_ref = collection.document(document_name)
    document = document_ref.get()

    if not document.exists:
        return render(request, 'frontend/techcare_data/document_not_found.html')

    document_ref.delete()

    messages.success(request, 'Trivia deleted successfully!')
    return redirect('trivia_view')

def users_delete(request, document_name):
    db = firestore.Client()
    collection = db.collection("users")
    document_ref = collection.document(document_name)
    document = document_ref.get()

    if not document.exists:
        return render(request, 'frontend/techcare_data/document_not_found.html')

    document_ref.delete()

    messages.success(request, 'User deleted successfully!')
    return redirect('users_view')

# def _delete(request, document_name):
#     db = firestore.Client()
#     collection = db.collection("")
#     document_ref = collection.document(document_name)
#     document = document_ref.get()

#     if not document.exists:
#         return render(request, 'frontend/techcare_data/document_not_found.html')

#     document_ref.delete()

#     messages.success(request, ' deleted successfully!')
#     return redirect('_view')


def update_activities(request, document_name):
    if request.method == 'POST':
        form = ActivitiesForm(request.POST)
        entry_id = document_name

        db = firestore.client()
        
        
        duration=int(request.POST.get('duration'))
        tags_path=request.POST.get('tags')

        tags_ref=get_document_reference(tags_path)
        data = {
                'tags': tags_ref,
                'description': request.POST.get('description'),
                'title': request.POST.get('title'),
                'duration': duration,
                'type': request.POST.get('type'),
                'track': request.POST.get('track'),
                'audiotrackId': request.POST.get('audiotrackId'),
                'audiotrackTitle': request.POST.get('audiotrackTitle'),
                'label': request.POST.get('label'),
               
            }
       
        db.collection("activities").document(entry_id).update(data)
        messages.success(request, 'Activity updated successfully.')
        return redirect('activities_view')
    else:
        messages.error(request, 'Error updating activity. Please check your input.')
        return redirect('activities_view')
    
def update_assessmentQuestion(request, document_name):
    if request.method == 'POST':
        form = AssessmentQuestionForm(request.POST)
        entry_id = document_name
        data = {
                'majorAssessment': request.POST.get('majorAssessment'),
                'max': request.POST.get('max'),
                'order': request.POST.get('order'),
                'points': request.POST.get('points'),
                'question': request.POST.get('question'),
            }
        db = firestore.client()
        db.collection("assessmentQuestion").document(entry_id).update(data)
        messages.success(request, 'assessmentQuestion updated successfully.')
        return redirect('assessmentQuestion_view')
    else:
        messages.error(request, 'Error updating assessmentQuestion. Please check your input.')
        return redirect('assessmentQuestion_view')
    
def update_badges(request, document_name):
    if request.method == 'POST':
        form = BadgesForm(request.POST)
        entry_id = document_name
        data = {
                'title': request.POST.get('title'),
            }
        db = firestore.client()
        db.collection("badges").document(entry_id).update(data)
        messages.success(request, 'badges updated successfully.')
        return redirect('badges_view')
    else:
        messages.error(request, 'Error updating badges. Please check your input.')
        return redirect('badges_view')
    
def update_biomarkers(request, document_name):
    if request.method == 'POST':
        form = BiomarkersForm(request.POST)
        entry_id = document_name
        db = firestore.client()

        user_path=request.POST.get('user')
        user_ref=get_document_reference(user_path)

        data = {
                'bloodGlucose': request.POST.get('bloodGlucose'),
                'bloodGlucoseType': request.POST.get('bloodGlucoseType'),
                'dailyActivity': request.POST.get('dailyActivity'),
                'dailyCarbs': request.POST.get('dailyCarbs'),
                'weeklyActivity': request.POST.get('weeklyActivity'),
                'sleepQuality': request.POST.get('sleepQuality'),
                'time': request.POST.get('time'),
                'user': user_ref,
                'weight': request.POST.get('weight'),
                'FBS': request.POST.get('FBS'),
                'HBA1c': request.POST.get('HBA1c'),
            }
        db = firestore.client()
        db.collection("biomarkers").document(entry_id).update(data)
        messages.success(request, ' biomarkers updated successfully.')
        return redirect('biomarkers_view')
    else:
        messages.error(request, 'Error updating  biomarkers . Please check your input.')
        return redirect('biomarkers_view')
    
def update_bites(request, document_name):
    if request.method == 'POST':
        form = BitesForm(request.POST)
        entry_id = document_name
        db = firestore.client()
    
        
        thumbs_up_users_path= request.POST.get('thumbs_up_users')
        thumbs_down_users_path= request.POST.get('thumbs_down_users')
        categories_path= request.POST.get('categories')
        tags_path=request.POST.get('tags')
         
     
        thumbs_up_users_ref=get_document_reference(thumbs_up_users_path)
        thumbs_down_users_ref=get_document_reference(thumbs_down_users_path)
        categories_ref=get_document_reference(categories_path)
        tags_ref=get_document_reference(tags_path)


        data = {
                'image': request.POST.get('image'),
                'order': request.POST.get('order'),
                'Learning_ponits': request.POST.get('Learning_ponits'),
                'CBT_points': request.POST.get('CBT_points'),
                'next': request.POST.get('next'),
                'scenarioID': request.POST.get('scenarioID'),
                'thumbs_up_users': thumbs_up_users_ref,
                'thumbs_down_users':thumbs_down_users_ref,
                'categories': categories_ref,
                'content': request.POST.get('content'),
                'difficulty': request.POST.get('difficulty'),
                'tags':tags_ref,
                'title': request.POST.get('title'),
            }
        db.collection("bites").document(entry_id).update(data)
        messages.success(request, 'bites updated successfully.')
        return redirect('bites_view')
    else:
        messages.error(request, 'Error updating bites  . Please check your input.')
        return redirect('bites_view')
    

def update_categories(request, document_name):
    if request.method == 'POST':
        form = CategoriesForm(request.POST)
        entry_id = document_name
        data = {
                'title': request.POST.get('title'),
                'description': request.POST.get('description'),
            }
        db = firestore.client()
        db.collection("categories").document(entry_id).update(data)
        messages.success(request, ' categories updated successfully.')
        return redirect('categories_view')
    else:
        messages.error(request, 'Error updating  categories . Please check your input.')
        return redirect('categories_view')
    

    
def update_feelings(request, document_name):
    if request.method == 'POST':
        form = FeelingsForm(request.POST)
        entry_id = document_name
        data = {
                'anger': request.POST.get('anger'),
                'fear': request.POST.get('fear'),
                'happiness': request.POST.get('happiness'),
                'joy': request.POST.get('joy'),
                'love': request.POST.get('love'),
                'sadness': request.POST.get('sadness'),
                'shame': request.POST.get('shame'),
                'strength': request.POST.get('strength'),
                'user': request.POST.get('user'),
                'time': request.POST.get('time'),
            }
        db = firestore.client()
        db.collection("feelings").document(entry_id).update(data)
        messages.success(request, 'feelings updated successfully.')
        return redirect('feelings_view')
    else:
        messages.error(request, 'Error updating feelings . Please check your input.')
        return redirect('feelings_view')
def update_inAppLinks(request, document_name):
    if request.method == 'POST':
        form = InAppLinksForm(request.POST)
        entry_id = document_name
        data = {
                'description': request.POST.get('description'),
                'title': request.POST.get('title'),
                'order': request.POST.get('order'),
                'type': request.POST.get('type'),
                'image': request.POST.get('image'),
                'link': request.POST.get('link'),
            }
        db = firestore.client()
        db.collection("inAppLinks").document(entry_id).update(data)
        messages.success(request, 'inAppLinks updated successfully.')
        return redirect('inAppLinks_view')
    else:
        messages.error(request, 'Error updating inAppLinks . Please check your input.')
        return redirect('inAppLinks_view')
    
def update_inquiry(request, document_name):
    if request.method == 'POST':
        form = InquiryForm(request.POST)
        entry_id = document_name
        db = firestore.client()

        
        
        user_path=request.POST.get('user')
        user_ref=get_document_reference(user_path)

        data = {
                'answer': request.POST.get('answer'),
                'question': request.POST.get('question'),
                'time': timezone.now(),
                'topic': request.POST.get('topic'),
                'user': user_ref,
            }
        db.collection("inquiry").document(entry_id).update(data)
        messages.success(request, 'inquiry updated successfully.')
        source = request.POST.get('source')
        if source == 'document_detail':
            return redirect('inquirydocument', document_name=document_name)
        else:
            return redirect('inquiry_view')
    else:
        messages.error(request, 'Error updating inquiry. Please check your input.')
        return redirect('inquiry_view')
    
def update_items(request, document_name):
    if request.method == 'POST':
        form = ItemsForm(request.POST)
        entry_id = document_name
        data = {
                'title': request.POST.get('title'),
                'data': request.POST.get('data'),
                'categories': request.POST.get('categories'),
            }
        db = firestore.client()
        db.collection("items").document(entry_id).update(data)
        messages.success(request, 'items updated successfully.')
        return redirect('items_view')
    else:
        messages.error(request, 'Error updating items . Please check your input.')
        return redirect('items_view')
    
def update_journal(request, document_name):
    if request.method == 'POST':
        form = JournalForm(request.POST)
        entry_id = document_name
        data = {
                'description': request.POST.get('description'),
                'title': request.POST.get('title'),
            }
        db = firestore.client()
        db.collection("journal").document(entry_id).update(data)
        messages.success(request, 'journal updated successfully.')
        return redirect('journal_view')
    else:
        messages.error(request, 'Error updating journal . Please check your input.')
        return redirect('journal_view')
    
def update_journalPrompt(request, document_name):
    if request.method == 'POST':
        form = JournalPromptForm(request.POST)
        entry_id = document_name
        data = {
                'title': request.POST.get('title'),
            }
        db = firestore.client()
        db.collection("journalPrompt").document(entry_id).update(data)
        messages.success(request, 'journalPrompt updated successfully.')
        return redirect('journalPrompt_view')
    else:
        messages.error(request, 'Error updating journalPrompt . Please check your input.')
        return redirect('journalPrompt_view')
    
def update_majorAssessment(request, document_name):
    if request.method == 'POST':
        form = MajorAssessmentForm(request.POST)
        entry_id = document_name
        data = {
                'description': request.POST.get('description'),
                'numberOfQuestions': request.POST.get('numberOfQuestions'),
                'order': request.POST.get('order'),
                'title': request.POST.get('title'),
            }
        db = firestore.client()
        db.collection("majorAssessment").document(entry_id).update(data)
        messages.success(request, 'majorAssessment updated successfully.')
        return redirect('majorAssessment_view')
    else:
        messages.error(request, 'Error updating majorAssessment . Please check your input.')
        return redirect('majorAssessment_view')
    
def update_psychomarkers(request, document_name):
    if request.method == 'POST':
        form = PsychomarkersForm(request.POST)
        entry_id = document_name
        db = firestore.client()
       
        
        user_path=request.POST.get('user')

        user_ref=get_document_reference(user_path)
        
        data = {
                'time': request.POST.get('time'),
                'user':user_ref,
                'depression': request.POST.get('depression'),
            }
        db = firestore.client()
        db.collection("psychomarkers").document(entry_id).update(data)
        messages.success(request, 'psychomarkers updated successfully.')
        return redirect('psychomarkers_view')
    else:
        messages.error(request, 'Error updating psychomarkers . Please check your input.')
        return redirect('psychomarkers_view')
    
def update_scenarios(request, document_name):
    if request.method == 'POST':
        form = ScenariosForm(request.POST)
        entry_id = document_name

        db = firestore.client()
       
        
        majorAssessment_path=request.POST.get('majorAssessment')
        suggestedActivity_path=request.POST.get('suggestedActivity')
        suggestedBite_path=request.POST.get('suggestedBite')
        suggestedJournal=request.POST.get('suggestedJournal')
        SuggestBitesFromBank_path=request.POST.get('SuggestBitesFromBank')

        majorAssessment_ref=get_document_reference(majorAssessment_path)
        suggestedActivity_ref=get_document_reference(suggestedActivity_path)
        suggestedBite_ref=get_document_reference(suggestedBite_path)
        suggestedJournal_ref=get_document_reference(suggestedJournal)
        SuggestBitesFromBank_ref=get_document_reference(SuggestBitesFromBank_path)


        data = {
                'PositiveCorrectionAlternative': request.POST.get('PositiveCorrectionAlternative'),
                'actionreply': request.POST.get('actionreply'),
                'correction': request.POST.get('correction'),
                'positiveActionReply': request.POST.get('positiveActionReply'),
                'title': request.POST.get('title'),
                'majorAssessment': majorAssessment_ref,
                'suggestedActivity': suggestedActivity_ref,
                'type': request.POST.get('type'),

                'suggestedBite': suggestedBite_ref,
                'suggestedJournal': suggestedJournal_ref,
                'max': request.POST.get('max'),
                'feeling': request.POST.get('feeling'),
                'status': request.POST.get('status'),
                'order': request.POST.get('order'),
                'story': request.POST.get('story'),
                'InteractiveStatement': request.POST.get('InteractiveStatement'),
                'Recommendation': request.POST.get('Recommendation'),
                'SuggestBitesFromBank': SuggestBitesFromBank_ref,

            }
        db = firestore.client()
        db.collection("scenarios").document(entry_id).update(data)
        messages.success(request, 'scenarios updated successfully.')
        return redirect('scenarios_view')
    else:
        messages.error(request, 'Error updating scenarios . Please check your input.')
        return redirect('scenarios_view')
    
def update_shortBite(request, document_name):
    if request.method == 'POST':
        form = ShortBiteForm(request.POST)
        entry_id = document_name
        data = {
                'order': request.POST.get('order'),
                'title': request.POST.get('title'),
            }
        db = firestore.client()
        db.collection("shortBite").document(entry_id).update(data)
        messages.success(request, 'shortBite updated successfully.')
        return redirect('shortBite_view')
    else:
        messages.error(request, 'Error updating shortBite . Please check your input.')
        return redirect('shortBite_view')
    
def update_tags(request, document_name):
    if request.method == 'POST':
        form = TagsForm(request.POST)
        entry_id = document_name
        data = {
                'description': request.POST.get('description'),
                'image': request.POST.get('image'),
                'title': request.POST.get('title'),
            }
        db = firestore.client()
        db.collection("tags").document(entry_id).update(data)
        messages.success(request, 'tags updated successfully.')
        return redirect('tags_view')
    else:
        messages.error(request, 'Error updating tags  . Please check your input.')
        return redirect('tags_view')
    
def update_trivia(request, document_name):
    if request.method == 'POST':
        form = TriviaForm(request.POST)
        entry_id = document_name
        data = {
                'CBT_points': request.POST.get('CBT_points'),
                'answer_1': request.POST.get('answer_1'),
                'answer_2': request.POST.get('answer_2'),
                'answer_3': request.POST.get('answer_3'),
                'correct_answer': request.POST.get('correct_answer'),
                'description': request.POST.get('description'),
                'explanation': request.POST.get('explanation'),
                'image': request.POST.get('image'),
                'question': request.POST.get('question'),
                'learning_points': request.POST.get('learning_points'),
                'order': request.POST.get('order'),
                'topic': request.POST.get('topic'),
            }
        db = firestore.client()
        db.collection("trivia").document(entry_id).update(data)
        messages.success(request, 'trivia updated successfully.')
        return redirect('trivia_view')
    else:
        messages.error(request, 'Error updating trivia  . Please check your input.')
        return redirect('trivia_view')
    
def update_users(request, document_name):
    if request.method == 'POST':
        form = UsersForm(request.POST)
        entry_id = document_name
        data = {
                'email': request.POST.get('email'),
                'created_time': request.POST.get('created_time'),
                'gender': request.POST.get('gender'),
                'uid': request.POST.get('uid'),
                'display_name': request.POST.get('display_name'),
            }
        db = firestore.client()
        db.collection("users").document(entry_id).update(data)
        messages.success(request, 'users updated successfully.')
        return redirect('users_view')
    else:
        messages.error(request, 'Error updating users  . Please check your input.')
        return redirect('users_view')
    
def update_selfawarenessScenarios(request, document_name):
    if request.method == 'POST':
        form = SelfAwarenessScenariosForm(request.POST)
        entry_id = document_name
        db = firestore.client()

      
        
        journal_path = request.POST.get('journal')
        activity_path= request.POST.get('activity')
        inAppLinks_path= request.POST.get('inAppLink')
        wildCard_path= request.POST.get('wildcard')
        biteID_path= request.POST.get('biteID')
        normalBites_path= request.POST.get('normalBite')

      


        try:
            journal_ref = get_document_reference(journal_path)
            activity_ref = get_document_reference(activity_path)
            inAppLinks_ref = get_document_reference(inAppLinks_path)
            normalBites_ref = get_document_reference(normalBites_path)
            biteID_ref = get_document_reference(biteID_path)
            wildCard_ref = get_document_reference(wildCard_path)
           
           
           
        except ValueError as e:
            messages.error(request, f'Invalid document path: {e}')
            return redirect('selfawarenessScenarios_view')
        data = {
                'activity': activity_ref,
                'biteID': biteID_ref,
                'correction1from0to2': request.POST.get('correction1from0to2'),
                'correction2from3to5': request.POST.get('correction2from3to5'),
                'inAppLink': inAppLinks_ref,
                'interactiveStatement': request.POST.get('interactiveStatement'),
                'journal': journal_ref,
                'normalBite':normalBites_ref,
                'recommendation1': request.POST.get('recommendation1'),
                'recommendation2': request.POST.get('recommendation2'),
                'scenarioID': request.POST.get('scenarioID'),
                'story': request.POST.get('story'),
                'storyTitle': request.POST.get('storyTitle'),
                'wildcard': wildCard_ref,
            }
        
        db.collection("selfawarenessScenarios").document(entry_id).update(data)
        messages.success(request, 'selfawarenessScenarios updated successfully.')
        return redirect('selfawarenessScenarios_view')
    else:
        messages.error(request, 'Error updating selfawarenessScenarios  . Please check your input.')
        return redirect('selfawarenessScenarios_view')
    

def update_assets(request, document_name):
    if request.method == 'POST':
        entry_id = document_name
        data = {
                'assetsType': request.POST.get('assetsType'),
                'label': request.POST.get('label'),
                'name': request.POST.get('name'),
                'path': request.POST.get('path'),
            }
        db = firestore.client()
        db.collection("assets").document(entry_id).update(data)
        messages.success(request, 'assets updated successfully.')
        return redirect('assets_view')
    else:
        messages.error(request, 'Error updating assets  . Please check your input.')
        return redirect('assets_view')    
    
def update_nutrition(request, document_name):
    if request.method == 'POST':
        entry_id = document_name
        data = {
                'carbContent': request.POST.get('carbContent'),
                'name_ar': request.POST.get('name_ar'),
                'name_en': request.POST.get('name_en'),
                'portion': request.POST.get('portion'),
                'proteinContent': request.POST.get('proteinContent'),
                'fatContent': request.POST.get('fatContent'),
                'totalCalories': request.POST.get('totalCalories'),
                'weight': request.POST.get('weight'),
                'unit': request.POST.get('unit'),
                'Class': request.POST.get('Class'),
                'food_category':request.POST.get('food_category'),


            }
        db = firestore.client()
        db.collection("nutrition").document(entry_id).update(data)
        messages.success(request, 'nutrition updated successfully.')
        return redirect('nutrition_view')
    else:
        messages.error(request, 'Error updating nutrition . Please check your input.')
        return redirect('nutrition_view')    
    
def update_readBites(request, document_name):
    if request.method == 'POST':
        entry_id = document_name
        db = firestore.client()


        user_path=request.POST.get('user')
        bite_path = request.POST.get('bite')
        
        user_ref=get_document_reference(user_path)
        bite_ref=get_document_reference(bite_path)
        data = {
                'bite': bite_ref,
                'user': user_ref,
                'time': request.POST.get('time'),
            }
        db = firestore.client()
        db.collection("readBites").document(entry_id).update(data)
        messages.success(request, 'readBites updated successfully.')
        return redirect('readBites_view')
    else:
        messages.error(request, 'Error updating readBites  . Please check your input.')
        return redirect('readBites_view')    
    
def update_readStories(request, document_name):
    if request.method == 'POST':
        entry_id = document_name

        db = firestore.client()

        
        

        user_path=request.POST.get('user')
        senario_path = request.POST.get('senario')
        
        user_ref=get_document_reference(user_path)
        senario_ref=get_document_reference(senario_path)

        data = {
                'senario': senario_ref,
                'user': user_ref,
                'time': request.POST.get('time'),
            }
        db = firestore.client()
        db.collection("readStories").document(entry_id).update(data)
        messages.success(request, 'readStories updated successfully.')
        return redirect('readStories_view')
    else:
        messages.error(request, 'Error updating readStories  . Please check your input.')
        return redirect('readStories_view')    
    
def update_selfAwarnessBites(request, document_name):
    if request.method == 'POST':
        entry_id = document_name
        data = {
                'scenarioID': request.POST.get('scenarioID'),
                'selfawarenessBiteText': request.POST.get('selfawarenessBiteText'),
                'selfawarenessBiteTitle': request.POST.get('selfawarenessBiteTitle'),
                'tags': request.POST.get('tags'),
            }
        db = firestore.client()
        db.collection("selfAwarnessBites").document(entry_id).update(data)
        messages.success(request, 'selfAwarnessBites updated successfully.')
        return redirect('selfAwarnessBites_view')
    else:
        messages.error(request, 'Error updating selfAwarnessBites  . Please check your input.')
        return redirect('selfAwarnessBites_view')    
    
def update_selfawareness_collection(request, document_name):
    if request.method == 'POST':
        entry_id = document_name
        data = {
                'selfawareness_bite_text': request.POST.get('selfawareness_bite_text'),
                'tags': request.POST.get('tags'),
            }
        db = firestore.client()
        db.collection("selfawareness_collection").document(entry_id).update(data)
        messages.success(request, 'selfawareness_collection updated successfully.')
        return redirect('selfawareness_collection_view')
    else:
        messages.error(request, 'Error updating selfawareness_collection   . Please check your input.')
        return redirect('selfawareness_collection_view')    
    
def update_suggestedActivities(request, document_name):
    if request.method == 'POST':
        entry_id = document_name
        db = firestore.client()

        

        activity_path=request.POST.get('activity')
        user_path=request.POST.get('user')
        activity_ref=get_document_reference(activity_path)  
        user_ref=get_document_reference(user_path)

        data = {
                'activity': activity_ref,
                'state': request.POST.get('state'),
                'type': request.POST.get('type'),
                'user': user_ref,
                'time': request.POST.get('time'),

            }
        db.collection("suggestedActivities").document(entry_id).update(data)
        messages.success(request, 'suggestedActivities updated successfully.')
        return redirect('suggestedActivities_view')
    else:
        messages.error(request, 'Error updating suggestedActivities   . Please check your input.')
        return redirect('suggestedActivities_view')    
    
def update_suggestedBites(request, document_name):
    if request.method == 'POST':
        entry_id = document_name
        db = firestore.client()

    
        
        bite_path=request.POST.get('bite')
        user_path=request.POST.get('user')
        selfAwarnessBite_path=request.POST.get('selfAwarnessBite')

        bite_ref=get_document_reference(bite_path)
        user_ref=get_document_reference(user_path)
        selfAwarnessBite_ref=get_document_reference(selfAwarnessBite_path)
        


        data = {
                'bite': bite_ref,
                'state': request.POST.get('state'),
                'user': user_ref,
                'selfAwarnessBite': selfAwarnessBite_ref,
                'time': request.POST.get('time'),

            }
        db = firestore.client()
        db.collection("suggestedBites").document(entry_id).update(data)
        messages.success(request, 'suggestedBites updated successfully.')
        return redirect('suggestedBites_view')
    else:
        messages.error(request, 'Error updating  suggestedBites . Please check your input.')
        return redirect('suggestedBites_view')    
    
def update_suggestedInAppLinks(request, document_name):
    if request.method == 'POST':
        entry_id = document_name
        db = firestore.client()

       
        
        user_path=request.POST.get('user')
        inAppLinks_path = request.POST.get('inAppLink')
        
        user_ref=get_document_reference(user_path)
        inAppLink_ref=get_document_reference(inAppLinks_path)

        data = {
                'inAppLink': inAppLink_ref,
                'user': user_ref,
                'time': request.POST.get('time'),

            }
        db = firestore.client()
        db.collection("suggestedInAppLinks").document(entry_id).update(data)
        messages.success(request, 'suggestedInAppLinks updated successfully.')
        return redirect('suggestedInAppLinks_view')
    else:
        messages.error(request, 'Error updating suggestedInAppLinks  . Please check your input.')
        return redirect('suggestedInAppLinks_view')    
    
def update_suggestedJournals(request, document_name):
    if request.method == 'POST':
        entry_id = document_name

        db = firestore.client()

        

        user_path=request.POST.get('user')
        journal_path = request.POST.get('journal')
        
        user_ref=get_document_reference(user_path)
        journal_ref=get_document_reference(journal_path)

        data = {
                'journal': journal_ref,
                'user': user_ref,
                'time': request.POST.get('time'),

            }
        db = firestore.client()
        db.collection("suggestedJournals").document(entry_id).update(data)
        messages.success(request, 'suggestedJournals updated successfully.')
        return redirect('suggestedJournals_view')
    else:
        messages.error(request, 'Error updating suggestedJournals  . Please check your input.')
        return redirect('suggestedJournals_view')    
    
def update_suggestedSelfAwarnessBites(request, document_name):
    if request.method == 'POST':
        entry_id = document_name

        db = firestore.client()


        

        user_path=request.POST.get('user')
        selfAwarnessBite_path = request.POST.get('selfAwarnessBite')
        
        user_ref=get_document_reference(user_path)
        selfAwarnessBite_ref=get_document_reference(selfAwarnessBite_path)

        data = {
                'selfAwarnessBite': selfAwarnessBite_ref,
                'state': request.POST.get('state'),
                'user': user_ref,
                'time': request.POST.get('time'),

            }
        db = firestore.client()
        db.collection("suggestedSelfAwarnessBites").document(entry_id).update(data)
        messages.success(request, 'suggestedSelfAwarnessBites updated successfully.')
        return redirect('suggestedSelfAwarnessBites_view')
    else:
        messages.error(request, 'Error updating  suggestedSelfAwarnessBites . Please check your input.')
        return redirect('suggestedSelfAwarnessBites_view')    
    
def update_suggestedWildCards(request, document_name):
    if request.method == 'POST':
        entry_id = document_name
        db = firestore.client()

        

        user_path=request.POST.get('user')
        wildCard_path = request.POST.get('wildCard')
        
        user_ref=get_document_reference(user_path)
        wildCard_ref=get_document_reference(wildCard_path)

        data = {
                'wildCard':wildCard_ref ,
                'user': user_ref,
                'time': request.POST.get('time'),

            }
        db = firestore.client()
        db.collection("suggestedWildCards").document(entry_id).update(data)
        messages.success(request, 'suggestedWildCards updated successfully.')
        return redirect('suggestedWildCards_view')
    else:
        messages.error(request, 'Error updating suggestedWildCards  . Please check your input.')
        return redirect('suggestedWildCards_view')    
    
def update_testTrivia(request, document_name):
    if request.method == 'POST':
        entry_id = document_name
        data = {
                'answers': request.POST.get('answers'),
                'correct_answer': request.POST.get('correct_answer'),
                'correct_answer_index': request.POST.get('correct_answer_index'),
                'order': request.POST.get('order'),
                'question': request.POST.get('question'),

            }
        db = firestore.client()
        db.collection("testTrivia").document(entry_id).update(data)
        messages.success(request, 'testTrivia updated successfully.')
        return redirect('testTrivia_view')
    else:
        messages.error(request, 'Error updating testTrivia  . Please check your input.')
        return redirect('testTrivia_view')    
    
def update_wildCard(request, document_name):
    if request.method == 'POST':
        entry_id = document_name
        data = {
                'content': request.POST.get('content'),
            }
        db = firestore.client()
        db.collection("wildCard").document(entry_id).update(data)
        messages.success(request, 'wildCard updated successfully.')
        return redirect('wildCard_view')
    else:
        messages.error(request, 'Error updating wildCard  . Please check your input.')
        return redirect('wildCard_view')    
    
def update_selfladder(request, document_name):
    if request.method == 'POST':
        entry_id = document_name
        db = firestore.client()
        
        userId_path=request.POST.get('userID')
        userId_ref=get_document_reference(userId_path)



        data = {
                'userID': userId_ref,
                'type': request.POST.get('type'),
                'time': request.POST.get('time'),
            }
        db.collection("selfLadder").document(entry_id).update(data)
        messages.success(request, 'selfladder updated successfully.')
        return redirect('selfladder_view')
    else:
        messages.error(request, 'Error updating selfladder  . Please check your input.')
        return redirect('selfladder_view')    
    


def update_notifications(request, document_name):
    if request.method == 'POST':
        entry_id = document_name
        db = firestore.client()
        data = {
                'title': request.POST.get('title'),
                'body': request.POST.get('body'),
                'time':request.POST.get('time'),
            }
        db.collection("notifications").document(entry_id).update(data)
        messages.success(request, 'notifications updated successfully.')
        return redirect('notifications_view')
    else:
        messages.error(request, 'Error updating notification  . Please check your input.')
        return redirect('notifications_view')  


def patient_search(request):
    query = request.GET.get('q')
    if query:
        results = users.objects.filter(
            Q(display_name__icontains=query) | Q(email__icontains=query)
        )
    else:
        results = users.objects.none()

    return render(request, 'frontend/techcare_data/pations_search.html', {'results': results})

# def chartbiomarkers(request):
#     try:
#         # Fetch biomarkers data from Firestore
#         biomarkers_collection = db.collection("biomarkers")
#         biomarkers_documents = biomarkers_collection.stream()

#         # Prepare data for Chart.js
#         data_labels = []
#         data_values = []

#         for doc in biomarkers_documents:
#             # Ensure 'bloodGlucose' field exists in document
#             doc_data = doc.to_dict()
#             if 'time' in doc_data and 'bloodGlucose' in doc_data:
#                 timestamp = doc_data['time'].strftime("%Y-%m-%d %H:%M:%S")
#                 data_labels.append(timestamp)
#                 data_values.append(doc_data['bloodGlucose'])
#         # Prepare data in a format suitable for JavaScript (JSON)
#         data_for_chartjs = {
#             'labels': data_labels,
#             'datasets': [{
#                 'label': 'Blood Glucose Levels',
#                 'data': data_values,
#                 'borderColor': 'rgb(75, 192, 192)',
#                 'fill': False  # Line chart should not be filled
#             }]
#         }

#         return render(request, 'frontend/techcare_data/chart_biomarkers.html', {'data_for_chartjs': data_for_chartjs})

#     except Exception as e:
#         # Handle exceptions gracefully, e.g., log the error
#         import logging
#         logging.error(f"Error fetching biomarkers data: {str(e)}")
#         # Render an error page or return an HTTP response with error details
#         return render(request, 'error.html', {'error_message': str(e)})


def export_nutrition_data(request):
    # Fetch data from Firestore
    db = firestore.client()
    collection = db.collection("nutrition")
    documents = collection.stream()
    document_data = [{'name': doc.id, **doc.to_dict()} for doc in documents]

    # Create a new Excel workbook and select the active worksheet
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Nutrition Data"

    # Define the column headers
    headers = [
        "carbContent", "name_ar", "name_en", "portion",
        "proteinContent", "totalCalories", "weight","fatContent","nutItem","food_category","Class","unit"
    ]
    worksheet.append(headers)

    # Add rows to the worksheet
    for entry in document_data:
        row = [
            entry.get("carbContent"),
            entry.get("name_ar"),
            entry.get("name_en"),
            entry.get("portion"),
            entry.get("proteinContent"),
            entry.get("totalCalories"),
            entry.get("weight"),
            entry.get("fatContent"),
            entry.get("nutItem"),
            entry.get("food_category"),
            entry.get("Class"),
            entry.get("unit")

        ]
        worksheet.append(row)

    # Save the workbook to a BytesIO stream
    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)

    # Create the HTTP response with the Excel file
    response = HttpResponse(
        stream.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response['Content-Disposition'] = 'attachment; filename="nutrition_data.xlsx"'
    return response



def import_nutrition_data(request):
    if request.method == 'POST' and request.FILES.get('import_file'):
        file = request.FILES['import_file']
        if file.name.endswith('.xlsx') or file.name.endswith('.xls'):
            # Save the file temporarily
            file_name = default_storage.save(file.name, file)
            file_path = Path(settings.MEDIA_ROOT) / file_name

            # Process the file
            try:
                # Read the Excel file into a DataFrame
                df = pd.read_excel(file_path)

                # Initialize Firestore client
                db = firestore.Client()

                # Reference to the Firestore collection
                collection_ref = db.collection("nutrition")

                # Add each row from the DataFrame to Firestore
                for index, row in df.iterrows():
                    document_data = {
                        "carbContent": row.get('carbContent'),
                        "name_ar": row.get('name_ar'),
                        "name_en": row.get('name_en'),
                        "portion": row.get('portion'),
                        "proteinContent": row.get('proteinContent'),
                        "totalCalories": row.get('totalCalories'),
                        "weight": row.get('weight'),
                        "fatContent":row.get('fatContent'),
                        "nutItem":row.get('nutItem'),
                        "food_category":row.get('food_category'),
                        "Class":row.get('Class'),
                        "unit":row.get('unit')
                    }
                    # Add document to Firestore with auto-generated ID
                    collection_ref.add(document_data)

                messages.success(request, 'Data imported successfully!')
            except Exception as e:
                messages.error(request, f'Error importing data: {e}')
        else:
            messages.error(request, 'Invalid file format. Please upload an Excel file.')
        
        return redirect('nutrition_view')  # Redirect to the appropriate view

    return render(request, 'frontend/techcare_data/nutrition_table.html')



def handle_value(value):
    """Converts Firestore specific types to suitable formats for Excel."""
    if isinstance(value, dict):
        # Flatten nested dictionaries
        return {f"{key}_{subkey}": subvalue for key, val in value.items() for subkey, subvalue in (val.items() if isinstance(val, dict) else [(key, val)])}
    elif hasattr(value, 'id') and hasattr(value, 'path'):
        # Handle DocumentReference objects and prepend '/'
        return f'/{value.path}'  # Converts DocumentReference to its path with leading '/'
    elif isinstance(value, list):
        # Join list elements into a single string
        return ', '.join(str(v) for v in value)
    elif isinstance(value, (int, float, str, bool)):
        # Handle simple data types directly
        return value
    return str(value)  # Convert other types to string

def export_activities_data(request):
    # Fetch data from Firestore
    db = firestore.client()
    collection = db.collection("activities")
    documents = collection.stream()
    
    document_data = []
    for doc in documents:
        data = doc.to_dict()
        # Handle specific Firestore types and flatten data
        flat_data = {key: handle_value(value) for key, value in data.items()}
        document_data.append(flat_data)

    # Create a new Excel workbook and select the active worksheet
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Activities Data"

    # Define the column headers
    headers = list(document_data[0].keys()) if document_data else []
    worksheet.append(headers)

    # Add rows to the worksheet
    for entry in document_data:
        row = [entry.get(header, '') for header in headers]
        worksheet.append(row)

    # Save the workbook to a BytesIO stream
    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)

    # Create the HTTP response with the Excel file
    response = HttpResponse(
        stream.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response['Content-Disposition'] = 'attachment; filename="activities_data.xlsx"'
    return response



def import_activities_data(request):
    if request.method == 'POST' and request.FILES.get('import_file'):
        file = request.FILES['import_file']
        if file.name.endswith('.xlsx') or file.name.endswith('.xls'):
            # Save the file temporarily
            file_name = default_storage.save(file.name, file)
            file_path = Path(settings.MEDIA_ROOT) / file_name

            # Process the file
            try:
                # Read the Excel file into a DataFrame
                df = pd.read_excel(file_path)

                # Initialize Firestore client
                db = firestore.Client()

                # Reference to the Firestore collection
                collection_ref = db.collection("activities")

                # Add each row from the DataFrame to Firestore
                for index, row in df.iterrows():
                    document_data = {
                        "tags": row.get('tags'),
                        "description": row.get('description'),
                        "duration": row.get('duration'),
                        "title": row.get('title'),
                        "type": row.get('type'),
                        "track": row.get('track'),
                        "audiotrackId": row.get('audiotrackId'),
                        "audiotrackTitle": row.get('audiotrackTitle'),
                        "label": row.get('label'),
                    }
                    # Add document to Firestore with auto-generated ID
                    collection_ref.add(document_data)

                messages.success(request, 'Data imported successfully!')
            except Exception as e:
                messages.error(request, f'Error importing data: {e}')
        else:
            messages.error(request, 'Invalid file format. Please upload an Excel file.')
        
        return redirect('activities_view')  # Redirect to the appropriate view

    return render(request, 'frontend/techcare_data/activities_table.html')







def export_assessmentQuestion_data(request):
    # Fetch data from Firestore
    db = firestore.client()
    collection = db.collection("assessmentQuestion")
    documents = collection.stream()
    
    document_data = []
    for doc in documents:
        data = doc.to_dict()
        # Handle specific Firestore types and flatten data
        flat_data = {key: handle_value(value) for key, value in data.items()}
        document_data.append(flat_data)

    # Create a new Excel workbook and select the active worksheet
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "AssessmentQuestion Data"

    # Define the column headers
    headers = list(document_data[0].keys()) if document_data else []
    worksheet.append(headers)

    # Add rows to the worksheet
    for entry in document_data:
        row = [entry.get(header, '') for header in headers]
        worksheet.append(row)

    # Save the workbook to a BytesIO stream
    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)

    # Create the HTTP response with the Excel file
    response = HttpResponse(
        stream.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response['Content-Disposition'] = 'attachment; filename="assessmentQuestion_data.xlsx"'
    return response



def import_assessmentQuestion_data(request):
    if request.method == 'POST' and request.FILES.get('import_file'):
        file = request.FILES['import_file']
        if file.name.endswith('.xlsx') or file.name.endswith('.xls'):
            # Save the file temporarily
            file_name = default_storage.save(file.name, file)
            file_path = Path(settings.MEDIA_ROOT) / file_name

            # Process the file
            try:
                # Read the Excel file into a DataFrame
                df = pd.read_excel(file_path)

                # Initialize Firestore client
                db = firestore.Client()

                # Reference to the Firestore collection
                collection_ref = db.collection("assessmentQuestion")

                # Add each row from the DataFrame to Firestore
                for index, row in df.iterrows():
                    document_data = {
                        "question": row.get('question'),
                        "points": row.get('points'),
                        "ngrams": row.get('ngrams'),
                        "majorAssessment": row.get('majorAssessment'),
                        "order": row.get('order'),
                        "max": row.get('max'),
                        "assessmentType": row.get('assessmentType'),
                    }
                    # Add document to Firestore with auto-generated ID
                    collection_ref.add(document_data)

                messages.success(request, 'Data imported successfully!')
            except Exception as e:
                messages.error(request, f'Error importing data: {e}')
        else:
            messages.error(request, 'Invalid file format. Please upload an Excel file.')
        
        return redirect('assessmentQuestion_view')  # Redirect to the appropriate view

    return render(request, 'frontend/techcare_data/assessmentQuestion_table.html')









def export_assets_data(request):
    # Fetch data from Firestore
    db = firestore.client()
    collection = db.collection("assets")
    documents = collection.stream()
    
    document_data = []
    for doc in documents:
        data = doc.to_dict()
        # Handle specific Firestore types and flatten data
        flat_data = {key: handle_value(value) for key, value in data.items()}
        document_data.append(flat_data)

    # Create a new Excel workbook and select the active worksheet
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Assets Data"

    # Define the column headers
    headers = list(document_data[0].keys()) if document_data else []
    worksheet.append(headers)

    # Add rows to the worksheet
    for entry in document_data:
        row = [entry.get(header, '') for header in headers]
        worksheet.append(row)

    # Save the workbook to a BytesIO stream
    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)

    # Create the HTTP response with the Excel file
    response = HttpResponse(
        stream.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response['Content-Disposition'] = 'attachment; filename="assets.xlsx"'
    return response



def import_assets_data(request):
    if request.method == 'POST' and request.FILES.get('import_file'):
        file = request.FILES['import_file']
        if file.name.endswith('.xlsx') or file.name.endswith('.xls'):
            # Save the file temporarily
            file_name = default_storage.save(file.name, file)
            file_path = Path(settings.MEDIA_ROOT) / file_name

            # Process the file
            try:
                # Read the Excel file into a DataFrame
                df = pd.read_excel(file_path)

                # Initialize Firestore client
                db = firestore.Client()

                # Reference to the Firestore collection
                collection_ref = db.collection("assets")

                # Add each row from the DataFrame to Firestore
                for index, row in df.iterrows():
                    document_data = {
                        "assetsType": row.get('assetsType'),
                        "label": row.get('label'),
                        "name": row.get('name'),
                        "path": row.get('path'),
                    }
                    # Add document to Firestore with auto-generated ID
                    collection_ref.add(document_data)

                messages.success(request, 'Data imported successfully!')
            except Exception as e:
                messages.error(request, f'Error importing data: {e}')
        else:
            messages.error(request, 'Invalid file format. Please upload an Excel file.')
        
        return redirect('assets_view')  # Redirect to the appropriate view

    return render(request, 'frontend/techcare_data/assets_view.html')


def export_badges_data(request):
    # Fetch data from Firestore
    db = firestore.client()
    collection = db.collection("badges")
    documents = collection.stream()
    
    document_data = []
    for doc in documents:
        data = doc.to_dict()
        # Handle specific Firestore types and flatten data
        flat_data = {key: handle_value(value) for key, value in data.items()}
        document_data.append(flat_data)

    # Create a new Excel workbook and select the active worksheet
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Badges Data"

    # Define the column headers
    headers = list(document_data[0].keys()) if document_data else []
    worksheet.append(headers)

    # Add rows to the worksheet
    for entry in document_data:
        row = [entry.get(header, '') for header in headers]
        worksheet.append(row)

    # Save the workbook to a BytesIO stream
    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)

    # Create the HTTP response with the Excel file
    response = HttpResponse(
        stream.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response['Content-Disposition'] = 'attachment; filename="badges.xlsx"'
    return response



def import_badges_data(request):
    if request.method == 'POST' and request.FILES.get('import_file'):
        file = request.FILES['import_file']
        if file.name.endswith('.xlsx') or file.name.endswith('.xls'):
            # Save the file temporarily
            file_name = default_storage.save(file.name, file)
            file_path = Path(settings.MEDIA_ROOT) / file_name

            # Process the file
            try:
                # Read the Excel file into a DataFrame
                df = pd.read_excel(file_path)

                # Initialize Firestore client
                db = firestore.Client()

                # Reference to the Firestore collection
                collection_ref = db.collection("badges")

                # Add each row from the DataFrame to Firestore
                for index, row in df.iterrows():
                    document_data = {
                        "title": row.get('title'),
                    }
                    # Add document to Firestore with auto-generated ID
                    collection_ref.add(document_data)

                messages.success(request, 'Data imported successfully!')
            except Exception as e:
                messages.error(request, f'Error importing data: {e}')
        else:
            messages.error(request, 'Invalid file format. Please upload an Excel file.')
        
        return redirect('badges_view')  # Redirect to the appropriate view

    return render(request, 'frontend/techcare_data/badges_view.html')




def export_biomarkers_data(request):
    # Fetch data from Firestore
    db = firestore.client()
    collection = db.collection("biomarkers")
    documents = collection.stream()

    document_data = []
    for doc in documents:
        data = doc.to_dict()
        # Handle specific Firestore types and flatten data
        flat_data = {key: handle_value(value) for key, value in data.items()}
        document_data.append(flat_data)

    if not document_data:
        return HttpResponse("No data available to export.", content_type="text/plain")

    # Create a new Excel workbook and select the active worksheet
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Biomarkers Data"

    # Define the column headers
    headers = list(set(header for entry in document_data for header in entry.keys()))  # Ensure all headers are included
    worksheet.append(headers)

    # Add rows to the worksheet
    for entry in document_data:
        row = [entry.get(header, '') for header in headers]
        worksheet.append(row)

    # Save the workbook to a BytesIO stream
    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)

    # Create the HTTP response with the Excel file
    response = HttpResponse(
        stream.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response['Content-Disposition'] = 'attachment; filename="biomarkers.xlsx"'
    return response


def import_biomarkers_data(request):
    if request.method == 'POST' and request.FILES.get('import_file'):
        file = request.FILES['import_file']
        if file.name.endswith('.xlsx') or file.name.endswith('.xls'):
            # Save the file temporarily
            file_name = default_storage.save(file.name, file)
            file_path = Path(settings.MEDIA_ROOT) / file_name

            # Process the file
            try:
                # Read the Excel file into a DataFrame
                df = pd.read_excel(file_path)

                # Initialize Firestore client
                db = firestore.Client()

                # Reference to the Firestore collection
                collection_ref = db.collection("biomarkers")

                # Add each row from the DataFrame to Firestore
                for index, row in df.iterrows():
                    document_data = {
                        "bloodGlucose": row.get('bloodGlucose'),
                        "bloodGlucoseType": row.get('bloodGlucoseType'),
                        "dailyActivity": row.get('dailyActivity'),
                        "dailyCarbs": row.get('dailyCarbs'),
                        "weeklyActivity": row.get('weeklyActivity'),
                        "sleepQuality": row.get('sleepQuality'),
                        "time": row.get('time'),
                        "weight": row.get('weight'),
                        "FBS": row.get('FBS'),
                        "HBA1c": row.get('HBA1c'),
                        "user": row.get('user'),

                    }
                    # Add document to Firestore with auto-generated ID
                    collection_ref.add(document_data)

                messages.success(request, 'Data imported successfully!')
            except Exception as e:
                messages.error(request, f'Error importing data: {e}')
        else:
            messages.error(request, 'Invalid file format. Please upload an Excel file.')
        
        return redirect('biomarkers_view')  # Redirect to the appropriate view

    return render(request, 'frontend/techcare_data/biomarkers_view.html')





def export_bites_data(request):
    # Fetch data from Firestore
    db = firestore.client()
    collection = db.collection("bites")
    documents = collection.stream()

    document_data = []
    for doc in documents:
        data = doc.to_dict()
        # Handle specific Firestore types and flatten data
        flat_data = {key: handle_value(value) for key, value in data.items()}
        document_data.append(flat_data)

    if not document_data:
        return HttpResponse("No data available to export.", content_type="text/plain")

    # Create a new Excel workbook and select the active worksheet
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Bites Data"

    # Define the column headers
    headers = list(set(header for entry in document_data for header in entry.keys()))  # Ensure all headers are included
    worksheet.append(headers)

    # Add rows to the worksheet
    for entry in document_data:
        row = [entry.get(header, '') for header in headers]
        worksheet.append(row)

    # Save the workbook to a BytesIO stream
    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)

    # Create the HTTP response with the Excel file
    response = HttpResponse(
        stream.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response['Content-Disposition'] = 'attachment; filename="bites.xlsx"'
    return response


def import_bites_data(request):
    if request.method == 'POST' and request.FILES.get('import_file'):
        file = request.FILES['import_file']
        if file.name.endswith('.xlsx') or file.name.endswith('.xls'):
            # Save the file temporarily
            file_name = default_storage.save(file.name, file)
            file_path = Path(settings.MEDIA_ROOT) / file_name

           
            try:
               
                df = pd.read_excel(file_path)

               
                db = firestore.Client()

               
                collection_ref = db.collection("bites")

                # Check for headers and set to None if not present
                headers = ["CBT_points", "Learning_points", "categories", "content", "difficulty", "image", "next", "order", "scenarioID", "tags", "title"]
                for header in headers:
                    if header not in df.columns:
                        df[header] = None
                for index, row in df.iterrows():
                    
                    categories_title = row.get('categories')

                  
                    categories_ref = db.collection("categories")
                    categories_query = categories_ref.where('title', '==', categories_title).limit(1).stream()
                    categories_ref = None
                    for categories in categories_query:
                        categories_ref = categories.reference
                        break

                    if not categories_ref:
                        messages.error(request, f'Category "{categories_title}" not found in categories collection.')
                        continue
                    
                    
                    tag_title = row.get('tags')

                  
                    tags_ref = db.collection("tags")
                    tag_query = tags_ref.where('title', '==', tag_title).limit(1).stream()
                    tag_ref = None
                    for tag in tag_query:
                        tag_ref = tag.reference
                        break

                    if not tag_ref:
                        messages.error(request, f'Tag "{tag_title}" not found in tags collection.')
                        continue
                    
                    document_data = {
                        "CBT_points": row.get('CBT_points'),
                        "Learning_points": row.get('Learning_points'),
                        "categories": categories_ref,
                        "content": row.get('content'),
                        "difficulty": row.get('difficulty'),
                        "image": row.get('image'),
                        "next": row.get('next'),
                        "order": row.get('order'),
                        "scenarioID": row.get('scenarioID'),
                        "tags": tag_ref,
                        "title": row.get('title'),

                    }
                    # Add document to Firestore with auto-generated ID
                    collection_ref.add(document_data)

                messages.success(request, 'Data imported successfully!')
            except Exception as e:
                messages.error(request, f'Error importing data: {e}')
        else:
            messages.error(request, 'Invalid file format. Please upload an Excel file.')
        
        return redirect('bites_view')  # Redirect to the appropriate view

    return render(request, 'frontend/techcare_data/bites_view.html')




def export_categories_data(request):
    # Fetch data from Firestore
    db = firestore.client()
    collection = db.collection("categories")
    documents = collection.stream()

    document_data = []
    for doc in documents:
        data = doc.to_dict()
        # Handle specific Firestore types and flatten data
        flat_data = {key: handle_value(value) for key, value in data.items()}
        document_data.append(flat_data)

    if not document_data:
        return HttpResponse("No data available to export.", content_type="text/plain")

    # Create a new Excel workbook and select the active worksheet
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Categories Data"

    # Define the column headers
    headers = list(set(header for entry in document_data for header in entry.keys()))  # Ensure all headers are included
    worksheet.append(headers)

    # Add rows to the worksheet
    for entry in document_data:
        row = [entry.get(header, '') for header in headers]
        worksheet.append(row)

    # Save the workbook to a BytesIO stream
    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)

    # Create the HTTP response with the Excel file
    response = HttpResponse(
        stream.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response['Content-Disposition'] = 'attachment; filename="categories.xlsx"'
    return response


def import_categories_data(request):
    if request.method == 'POST' and request.FILES.get('import_file'):
        file = request.FILES['import_file']
        if file.name.endswith('.xlsx') or file.name.endswith('.xls'):
            # Save the file temporarily
            file_name = default_storage.save(file.name, file)
            file_path = Path(settings.MEDIA_ROOT) / file_name

            # Process the file
            try:
                # Read the Excel file into a DataFrame
                df = pd.read_excel(file_path)

                # Initialize Firestore client
                db = firestore.Client()

                # Reference to the Firestore collection
                collection_ref = db.collection("categories")

                # Add each row from the DataFrame to Firestore
                for index, row in df.iterrows():
                    document_data = {
                        "description": row.get('description'),
                        "title": row.get('title'),

                    }
                    # Add document to Firestore with auto-generated ID
                    collection_ref.add(document_data)

                messages.success(request, 'Data imported successfully!')
            except Exception as e:
                messages.error(request, f'Error importing data: {e}')
        else:
            messages.error(request, 'Invalid file format. Please upload an Excel file.')
        
        return redirect('categories_view')  # Redirect to the appropriate view

    return render(request, 'frontend/techcare_data/categories_view.html')



def export_feelings_data(request):
    # Fetch data from Firestore
    db = firestore.client()
    collection = db.collection("feelings")
    documents = collection.stream()

    document_data = []
    for doc in documents:
        data = doc.to_dict()
        # Handle specific Firestore types and flatten data
        flat_data = {key: handle_value(value) for key, value in data.items()}
        document_data.append(flat_data)

    if not document_data:
        return HttpResponse("No data available to export.", content_type="text/plain")

    # Create a new Excel workbook and select the active worksheet
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Feelings Data"

    # Define the column headers
    headers = list(set(header for entry in document_data for header in entry.keys()))  # Ensure all headers are included
    worksheet.append(headers)

    # Add rows to the worksheet
    for entry in document_data:
        row = [entry.get(header, '') for header in headers]
        worksheet.append(row)

    # Save the workbook to a BytesIO stream
    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)

    # Create the HTTP response with the Excel file
    response = HttpResponse(
        stream.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response['Content-Disposition'] = 'attachment; filename="feelings.xlsx"'
    return response


def import_feelings_data(request):
    if request.method == 'POST' and request.FILES.get('import_file'):
        file = request.FILES['import_file']
        if file.name.endswith('.xlsx') or file.name.endswith('.xls'):
            # Save the file temporarily
            file_name = default_storage.save(file.name, file)
            file_path = Path(settings.MEDIA_ROOT) / file_name

            # Process the file
            try:
                # Read the Excel file into a DataFrame
                df = pd.read_excel(file_path)

                # Initialize Firestore client
                db = firestore.Client()

                # Reference to the Firestore collection
                collection_ref = db.collection("feelings")

                # Add each row from the DataFrame to Firestore
                for index, row in df.iterrows():
                    document_data = {
                        "anger": row.get('anger'),
                        "fear": row.get('fear'),
                        "happiness": row.get('happiness'),
                        "joy": row.get('joy'),
                        "love": row.get('love'),
                        "sadness": row.get('sadness'),
                        "shame": row.get('shame'),
                        "strength": row.get('strength'),
                        "user": row.get('user'),
                        "time": row.get('time'),
                        

                    }
                    # Add document to Firestore with auto-generated ID
                    collection_ref.add(document_data)

                messages.success(request, 'Data imported successfully!')
            except Exception as e:
                messages.error(request, f'Error importing data: {e}')
        else:
            messages.error(request, 'Invalid file format. Please upload an Excel file.')
        
        return redirect('feelings_view')  # Redirect to the appropriate view

    return render(request, 'frontend/techcare_data/feelings_view.html')




def export_inAppLinks_data(request):
    # Fetch data from Firestore
    db = firestore.client()
    collection = db.collection("inAppLinks")
    documents = collection.stream()

    document_data = []
    for doc in documents:
        data = doc.to_dict()
        # Handle specific Firestore types and flatten data
        flat_data = {key: handle_value(value) for key, value in data.items()}
        document_data.append(flat_data)

    if not document_data:
        return HttpResponse("No data available to export.", content_type="text/plain")

    # Create a new Excel workbook and select the active worksheet
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "InAppLinks Data"

    # Define the column headers
    headers = list(set(header for entry in document_data for header in entry.keys()))  # Ensure all headers are included
    worksheet.append(headers)

    # Add rows to the worksheet
    for entry in document_data:
        row = [entry.get(header, '') for header in headers]
        worksheet.append(row)

    # Save the workbook to a BytesIO stream
    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)

    # Create the HTTP response with the Excel file
    response = HttpResponse(
        stream.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response['Content-Disposition'] = 'attachment; filename="inAppLinks.xlsx"'
    return response


def import_inAppLinks_data(request):
    if request.method == 'POST' and request.FILES.get('import_file'):
        file = request.FILES['import_file']
        if file.name.endswith('.xlsx') or file.name.endswith('.xls'):
            # Save the file temporarily
            file_name = default_storage.save(file.name, file)
            file_path = Path(settings.MEDIA_ROOT) / file_name

            # Process the file
            try:
                # Read the Excel file into a DataFrame
                df = pd.read_excel(file_path)

                # Initialize Firestore client
                db = firestore.Client()

                # Reference to the Firestore collection
                collection_ref = db.collection("inAppLinks")

                # Add each row from the DataFrame to Firestore
                for index, row in df.iterrows():
                    document_data = {
                        "description": row.get('description'),
                        "order": row.get('order'),
                        "type": row.get('type'),
                        "title": row.get('title'),
                        "image": row.get('image'),
                        "link": row.get('link'),
                        
                    }
                    # Add document to Firestore with auto-generated ID
                    collection_ref.add(document_data)

                messages.success(request, 'Data imported successfully!')
            except Exception as e:
                messages.error(request, f'Error importing data: {e}')
        else:
            messages.error(request, 'Invalid file format. Please upload an Excel file.')
        
        return redirect('inAppLinks_view')  # Redirect to the appropriate view

    return render(request, 'frontend/techcare_data/inAppLinks_view.html')






def export_inquiry_data(request):
    # Fetch data from Firestore
    db = firestore.client()
    collection = db.collection("inquiry")
    documents = collection.stream()

    document_data = []
    for doc in documents:
        data = doc.to_dict()
        # Handle specific Firestore types and flatten data
        flat_data = {key: handle_value(value) for key, value in data.items()}
        document_data.append(flat_data)

    if not document_data:
        return HttpResponse("No data available to export.", content_type="text/plain")

    # Create a new Excel workbook and select the active worksheet
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Inquiry Data"

    # Define the column headers
    headers = list(set(header for entry in document_data for header in entry.keys()))  # Ensure all headers are included
    worksheet.append(headers)

    # Add rows to the worksheet
    for entry in document_data:
        row = [entry.get(header, '') for header in headers]
        worksheet.append(row)

    # Save the workbook to a BytesIO stream
    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)

    # Create the HTTP response with the Excel file
    response = HttpResponse(
        stream.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response['Content-Disposition'] = 'attachment; filename="inquiry.xlsx"'
    return response


def import_inquiry_data(request):
    if request.method == 'POST' and request.FILES.get('import_file'):
        file = request.FILES['import_file']
        if file.name.endswith('.xlsx') or file.name.endswith('.xls'):
            # Save the file temporarily
            file_name = default_storage.save(file.name, file)
            file_path = Path(settings.MEDIA_ROOT) / file_name

            # Process the file
            try:
                # Read the Excel file into a DataFrame
                df = pd.read_excel(file_path)

                # Initialize Firestore client
                db = firestore.Client()

                # Reference to the Firestore collection
                collection_ref = db.collection("inquiry")

                # Add each row from the DataFrame to Firestore
                for index, row in df.iterrows():
                    document_data = {
                        "answer": row.get('answer'),
                        "question": row.get('question'),
                        "time": row.get('time'),
                        "topic": row.get('topic'),
                        "user": row.get('user'),                        
                    }
                    # Add document to Firestore with auto-generated ID
                    collection_ref.add(document_data)

                messages.success(request, 'Data imported successfully!')
            except Exception as e:
                messages.error(request, f'Error importing data: {e}')
        else:
            messages.error(request, 'Invalid file format. Please upload an Excel file.')
        
        return redirect('inquiry_view')  # Redirect to the appropriate view

    return render(request, 'frontend/techcare_data/inquiry_view.html')






def export_items_data(request):
    # Fetch data from Firestore
    db = firestore.client()
    collection = db.collection("items")
    documents = collection.stream()

    # Define the column headers and filtered data
    document_data = []
    headers = {'name'}
    
    for doc in documents:
        data = doc.to_dict()
        # Handle specific Firestore types and flatten data
        flat_data = {key: handle_value(value) for key, value in data.items()}
        
        # Extract only the 'name' field
        filtered_data = {key: value for key, value in flat_data.items() if key == 'name'}
        document_data.append(filtered_data)
    
    if not document_data:
        # No data available, return a file with only headers
        document_data = [{'name': 'No data available'}]
    
    # Create a new Excel workbook and select the active worksheet
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Items Data"

    # Append headers and data to the worksheet
    worksheet.append(['name'])  # Only 'name' header
    
    for data in document_data:
        worksheet.append([data.get('name', '')])  # Append only 'name' field data

    # Save the workbook to a BytesIO stream
    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)

    # Create the HTTP response with the Excel file
    response = HttpResponse(
        stream.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response['Content-Disposition'] = 'attachment; filename="items.xlsx"'
    return response


def import_items_data(request):
    if request.method == 'POST' and request.FILES.get('import_file'):
        file = request.FILES['import_file']
        if file.name.endswith('.xlsx') or file.name.endswith('.xls'):
            # Save the file temporarily
            file_name = default_storage.save(file.name, file)
            file_path = Path(settings.MEDIA_ROOT) / file_name

            # Process the file
            try:
                # Read the Excel file into a DataFrame
                df = pd.read_excel(file_path)

                # Initialize Firestore client
                db = firestore.Client()

                # Reference to the Firestore collection
                collection_ref = db.collection("items")

                # Add each row from the DataFrame to Firestore
                for index, row in df.iterrows():
                    document_data = {
                        "data": row.get('data'),
                        "name": row.get('name'),                      
                    }
                    # Add document to Firestore with auto-generated ID
                    collection_ref.add(document_data)

                messages.success(request, 'Data imported successfully!')
            except Exception as e:
                messages.error(request, f'Error importing data: {e}')
        else:
            messages.error(request, 'Invalid file format. Please upload an Excel file.')
        
        return redirect('items_view')  # Redirect to the appropriate view

    return render(request, 'frontend/techcare_data/items_view.html')





def export_journal_data(request):
    # Fetch data from Firestore
    db = firestore.client()
    collection = db.collection("journal")
    documents = collection.stream()

    document_data = []
    for doc in documents:
        data = doc.to_dict()
        # Handle specific Firestore types and flatten data
        flat_data = {key: handle_value(value) for key, value in data.items()}
        document_data.append(flat_data)

    if not document_data:
        return HttpResponse("No data available to export.", content_type="text/plain")

    # Create a new Excel workbook and select the active worksheet
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Journal Data"

    # Define the column headers
    headers = list(set(header for entry in document_data for header in entry.keys()))  # Ensure all headers are included
    worksheet.append(headers)

    # Add rows to the worksheet
    for entry in document_data:
        row = [entry.get(header, '') for header in headers]
        worksheet.append(row)

    # Save the workbook to a BytesIO stream
    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)

    # Create the HTTP response with the Excel file
    response = HttpResponse(
        stream.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response['Content-Disposition'] = 'attachment; filename="journal.xlsx"'
    return response


def import_journal_data(request):
    if request.method == 'POST' and request.FILES.get('import_file'):
        file = request.FILES['import_file']
        if file.name.endswith('.xlsx') or file.name.endswith('.xls'):
            # Save the file temporarily
            file_name = default_storage.save(file.name, file)
            file_path = Path(settings.MEDIA_ROOT) / file_name

            # Process the file
            try:
                # Read the Excel file into a DataFrame
                df = pd.read_excel(file_path)

                # Initialize Firestore client
                db = firestore.Client()

                # Reference to the Firestore collection
                collection_ref = db.collection("journal")

                # Add each row from the DataFrame to Firestore
                for index, row in df.iterrows():
                    document_data = {
                        "content": row.get('content'),
                        "feeling": row.get('feeling'), 
                        "time": row.get('time'), 
                        "title": row.get('title'),                     
                        "user": row.get('user'), 
                        "tags": row.get('tags'),
                        "description": row.get('description'),
                    }
                    # Add document to Firestore with auto-generated ID
                    collection_ref.add(document_data)

                messages.success(request, 'Data imported successfully!')
            except Exception as e:
                messages.error(request, f'Error importing data: {e}')
        else:
            messages.error(request, 'Invalid file format. Please upload an Excel file.')
        
        return redirect('journal_view')  # Redirect to the appropriate view

    return render(request, 'frontend/techcare_data/journal_view.html')




def export_journalPrompt_data(request):
    # Fetch data from Firestore
    db = firestore.client()
    collection = db.collection("journalPrompt")
    documents = collection.stream()

    document_data = []
    for doc in documents:
        data = doc.to_dict()
        # Handle specific Firestore types and flatten data
        flat_data = {key: handle_value(value) for key, value in data.items()}
        document_data.append(flat_data)

    if not document_data:
        return HttpResponse("No data available to export.", content_type="text/plain")

    # Create a new Excel workbook and select the active worksheet
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "JournalPrompt Data"

    # Define the column headers
    headers = list(set(header for entry in document_data for header in entry.keys()))  # Ensure all headers are included
    worksheet.append(headers)

    # Add rows to the worksheet
    for entry in document_data:
        row = [entry.get(header, '') for header in headers]
        worksheet.append(row)

    # Save the workbook to a BytesIO stream
    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)

    # Create the HTTP response with the Excel file
    response = HttpResponse(
        stream.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response['Content-Disposition'] = 'attachment; filename="journalPrompt.xlsx"'
    return response


def import_journalPrompt_data(request):
    if request.method == 'POST' and request.FILES.get('import_file'):
        file = request.FILES['import_file']
        if file.name.endswith('.xlsx') or file.name.endswith('.xls'):
            # Save the file temporarily
            file_name = default_storage.save(file.name, file)
            file_path = Path(settings.MEDIA_ROOT) / file_name

            # Process the file
            try:
                # Read the Excel file into a DataFrame
                df = pd.read_excel(file_path)

                # Initialize Firestore client
                db = firestore.Client()

                # Reference to the Firestore collection
                collection_ref = db.collection("journalPrompt")

                # Add each row from the DataFrame to Firestore
                for index, row in df.iterrows():
                    document_data = {
                        "scenarioID": row.get('scenarioID'),
                        "title": row.get('title'),                     
                    }
                    # Add document to Firestore with auto-generated ID
                    collection_ref.add(document_data)

                messages.success(request, 'Data imported successfully!')
            except Exception as e:
                messages.error(request, f'Error importing data: {e}')
        else:
            messages.error(request, 'Invalid file format. Please upload an Excel file.')
        
        return redirect('journalPrompt_view')  # Redirect to the appropriate view

    return render(request, 'frontend/techcare_data/journalPrompt_view.html')




def export_majorAssessment_data(request):
    # Fetch data from Firestore
    db = firestore.client()
    collection = db.collection("majorAssessment")
    documents = collection.stream()

    document_data = []
    for doc in documents:
        data = doc.to_dict()
        # Handle specific Firestore types and flatten data
        flat_data = {key: handle_value(value) for key, value in data.items()}
        document_data.append(flat_data)

    if not document_data:
        return HttpResponse("No data available to export.", content_type="text/plain")

    # Create a new Excel workbook and select the active worksheet
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "MajorAssessment Data"

    # Define the column headers
    headers = list(set(header for entry in document_data for header in entry.keys()))  # Ensure all headers are included
    worksheet.append(headers)

    # Add rows to the worksheet
    for entry in document_data:
        row = [entry.get(header, '') for header in headers]
        worksheet.append(row)

    # Save the workbook to a BytesIO stream
    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)

    # Create the HTTP response with the Excel file
    response = HttpResponse(
        stream.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response['Content-Disposition'] = 'attachment; filename="majorAssessment.xlsx"'
    return response


def import_majorAssessment_data(request):
    if request.method == 'POST' and request.FILES.get('import_file'):
        file = request.FILES['import_file']
        if file.name.endswith('.xlsx') or file.name.endswith('.xls'):
            # Save the file temporarily
            file_name = default_storage.save(file.name, file)
            file_path = Path(settings.MEDIA_ROOT) / file_name

            # Process the file
            try:
                # Read the Excel file into a DataFrame
                df = pd.read_excel(file_path)

                # Initialize Firestore client
                db = firestore.Client()

                # Reference to the Firestore collection
                collection_ref = db.collection("majorAssessment")

                # Add each row from the DataFrame to Firestore
                for index, row in df.iterrows():
                    document_data = {
                        "description": row.get('description'),
                        "title": row.get('title'),      
                        "numberOfQuestions": row.get('numberOfQuestions'),      
                        "order": row.get('order'),                     
                    }
                    # Add document to Firestore with auto-generated ID
                    collection_ref.add(document_data)

                messages.success(request, 'Data imported successfully!')
            except Exception as e:
                messages.error(request, f'Error importing data: {e}')
        else:
            messages.error(request, 'Invalid file format. Please upload an Excel file.')
        
        return redirect('majorAssessment_view')  # Redirect to the appropriate view

    return render(request, 'frontend/techcare_data/majorAssessment_view.html')


def export_psychomarkers_data(request):
    # Fetch data from Firestore
    db = firestore.client()
    collection = db.collection("psychomarkers")
    documents = collection.stream()

    document_data = []
    for doc in documents:
        data = doc.to_dict()
        # Handle specific Firestore types and flatten data
        flat_data = {key: handle_value(value) for key, value in data.items()}
        document_data.append(flat_data)

    if not document_data:
        return HttpResponse("No data available to export.", content_type="text/plain")

    # Create a new Excel workbook and select the active worksheet
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Psychomarkers Data"

    # Define the column headers
    headers = list(set(header for entry in document_data for header in entry.keys()))  # Ensure all headers are included
    worksheet.append(headers)

    # Add rows to the worksheet
    for entry in document_data:
        row = [entry.get(header, '') for header in headers]
        worksheet.append(row)

    # Save the workbook to a BytesIO stream
    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)

    # Create the HTTP response with the Excel file
    response = HttpResponse(
        stream.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response['Content-Disposition'] = 'attachment; filename="psychomarkers.xlsx"'
    return response


def import_psychomarkers_data(request):
    if request.method == 'POST' and request.FILES.get('import_file'):
        file = request.FILES['import_file']
        if file.name.endswith('.xlsx') or file.name.endswith('.xls'):
            # Save the file temporarily
            file_name = default_storage.save(file.name, file)
            file_path = Path(settings.MEDIA_ROOT) / file_name

            # Process the file
            try:
                # Read the Excel file into a DataFrame
                df = pd.read_excel(file_path)

                # Initialize Firestore client
                db = firestore.Client()

                # Reference to the Firestore collection
                collection_ref = db.collection("psychomarkers")

                # Add each row from the DataFrame to Firestore
                for index, row in df.iterrows():
                    document_data = {
                        "depression": row.get('depression'),
                        "time": row.get('time'),
                        "user": row.get('user'),
                        "anxiety": row.get('anxiety'),
                                     
                    }
                    # Add document to Firestore with auto-generated ID
                    collection_ref.add(document_data)

                messages.success(request, 'Data imported successfully!')
            except Exception as e:
                messages.error(request, f'Error importing data: {e}')
        else:
            messages.error(request, 'Invalid file format. Please upload an Excel file.')
        
        return redirect('psychomarkers_view')  # Redirect to the appropriate view

    return render(request, 'frontend/techcare_data/psychomarkers_view.html')







def export_readBites_data(request):
    # Fetch data from Firestore
    db = firestore.client()
    collection = db.collection("readBites")
    documents = collection.stream()

    document_data = []
    for doc in documents:
        data = doc.to_dict()
        # Handle specific Firestore types and flatten data
        flat_data = {key: handle_value(value) for key, value in data.items()}
        document_data.append(flat_data)

    if not document_data:
        return HttpResponse("No data available to export.", content_type="text/plain")

    # Create a new Excel workbook and select the active worksheet
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "ReadBites Data"

    # Define the column headers
    headers = list(set(header for entry in document_data for header in entry.keys()))  # Ensure all headers are included
    worksheet.append(headers)

    # Add rows to the worksheet
    for entry in document_data:
        row = [entry.get(header, '') for header in headers]
        worksheet.append(row)

    # Save the workbook to a BytesIO stream
    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)

    # Create the HTTP response with the Excel file
    response = HttpResponse(
        stream.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response['Content-Disposition'] = 'attachment; filename="readBites.xlsx"'
    return response


def import_readBites_data(request):
    if request.method == 'POST' and request.FILES.get('import_file'):
        file = request.FILES['import_file']
        if file.name.endswith('.xlsx') or file.name.endswith('.xls'):
            # Save the file temporarily
            file_name = default_storage.save(file.name, file)
            file_path = Path(settings.MEDIA_ROOT) / file_name

            # Process the file
            try:
                # Read the Excel file into a DataFrame
                df = pd.read_excel(file_path)

                # Initialize Firestore client
                db = firestore.Client()

                # Reference to the Firestore collection
                collection_ref = db.collection("readBites")

                # Add each row from the DataFrame to Firestore
                for index, row in df.iterrows():
                    document_data = {
                        "bite": row.get('bite'),
                        "time": row.get('time'),
                        "user": row.get('user'),                 
                    }
                    # Add document to Firestore with auto-generated ID
                    collection_ref.add(document_data)

                messages.success(request, 'Data imported successfully!')
            except Exception as e:
                messages.error(request, f'Error importing data: {e}')
        else:
            messages.error(request, 'Invalid file format. Please upload an Excel file.')
        
        return redirect('readBites_view')  # Redirect to the appropriate view

    return render(request, 'frontend/techcare_data/readBites_view.html')




def export_readStories_data(request):
    # Fetch data from Firestore
    db = firestore.client()
    collection = db.collection("readStories")
    documents = collection.stream()

    document_data = []
    for doc in documents:
        data = doc.to_dict()
        # Handle specific Firestore types and flatten data
        flat_data = {key: handle_value(value) for key, value in data.items()}
        document_data.append(flat_data)

    if not document_data:
        return HttpResponse("No data available to export.", content_type="text/plain")

    # Create a new Excel workbook and select the active worksheet
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "ReadStories Data"

    # Define the column headers
    headers = list(set(header for entry in document_data for header in entry.keys()))  # Ensure all headers are included
    worksheet.append(headers)

    # Add rows to the worksheet
    for entry in document_data:
        row = [entry.get(header, '') for header in headers]
        worksheet.append(row)

    # Save the workbook to a BytesIO stream
    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)

    # Create the HTTP response with the Excel file
    response = HttpResponse(
        stream.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response['Content-Disposition'] = 'attachment; filename="readStories.xlsx"'
    return response


def import_readStories_data(request):
    if request.method == 'POST' and request.FILES.get('import_file'):
        file = request.FILES['import_file']
        if file.name.endswith('.xlsx') or file.name.endswith('.xls'):
            # Save the file temporarily
            file_name = default_storage.save(file.name, file)
            file_path = Path(settings.MEDIA_ROOT) / file_name

            # Process the file
            try:
                # Read the Excel file into a DataFrame
                df = pd.read_excel(file_path)

                # Initialize Firestore client
                db = firestore.Client()

                # Reference to the Firestore collection
                collection_ref = db.collection("readStories")

                # Add each row from the DataFrame to Firestore
                for index, row in df.iterrows():
                    document_data = {
                        "senario": row.get('senario'),
                        "time": row.get('time'),
                        "user": row.get('user'),                 
                    }
                    # Add document to Firestore with auto-generated ID
                    collection_ref.add(document_data)

                messages.success(request, 'Data imported successfully!')
            except Exception as e:
                messages.error(request, f'Error importing data: {e}')
        else:
            messages.error(request, 'Invalid file format. Please upload an Excel file.')
        
        return redirect('readStories_view')  # Redirect to the appropriate view

    return render(request, 'frontend/techcare_data/readStories_view.html')




def export_scenarios_data(request):
    # Fetch data from Firestore
    db = firestore.client()
    collection = db.collection("scenarios")
    documents = collection.stream()

    document_data = []
    for doc in documents:
        data = doc.to_dict()
        # Handle specific Firestore types and flatten data
        flat_data = {key: handle_value(value) for key, value in data.items()}
        document_data.append(flat_data)

    if not document_data:
        return HttpResponse("No data available to export.", content_type="text/plain")

    # Create a new Excel workbook and select the active worksheet
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Scenarios Data"

    # Define the column headers
    headers = list(set(header for entry in document_data for header in entry.keys()))  # Ensure all headers are included
    worksheet.append(headers)

    # Add rows to the worksheet
    for entry in document_data:
        row = [entry.get(header, '') for header in headers]
        worksheet.append(row)

    # Save the workbook to a BytesIO stream
    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)

    # Create the HTTP response with the Excel file
    response = HttpResponse(
        stream.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response['Content-Disposition'] = 'attachment; filename="scenarios.xlsx"'
    return response


def import_scenarios_data(request):
    if request.method == 'POST' and request.FILES.get('import_file'):
        file = request.FILES['import_file']
        if file.name.endswith('.xlsx') or file.name.endswith('.xls'):
            # Save the file temporarily
            file_name = default_storage.save(file.name, file)
            file_path = Path(settings.MEDIA_ROOT) / file_name

            # Process the file
            try:
                # Read the Excel file into a DataFrame
                df = pd.read_excel(file_path)

                # Initialize Firestore client
                db = firestore.Client()

                # Reference to the Firestore collection
                collection_ref = db.collection("scenarios")

                # Add each row from the DataFrame to Firestore
                for index, row in df.iterrows():
                    document_data = {
                        "InteractiveStatement": row.get('InteractiveStatement'),
                        "PositiveCorrectionAlternative": row.get('PositiveCorrectionAlternative'),
                        "Recommendation": row.get('Recommendation'), 
                        "SuggestBitesFromBank": row.get('SuggestBitesFromBank'),                 
                        "actionreply": row.get('actionreply'),  
                        "correction": row.get('correction'),  
                        "feeling": row.get('feeling'),  
                        "majorAssessment": row.get('majorAssessment'),  
                        "max": row.get('max'),  
                        "order": row.get('order'),  
                        "positiveActionReply": row.get('positiveActionReply'),  
                        "status": row.get('status'),  
                        "story": row.get('storystory'),  
                        "suggestedActivity": row.get('suggestedActivity'),  
                        "suggestedBite": row.get('suggestedBite'),  
                        "suggestedJournal": row.get('suggestedJournal'),  
                        "title": row.get('title'),  
                        "type": row.get('type'),  
                    }
                    # Add document to Firestore with auto-generated ID
                    collection_ref.add(document_data)

                messages.success(request, 'Data imported successfully!')
            except Exception as e:
                messages.error(request, f'Error importing data: {e}')
        else:
            messages.error(request, 'Invalid file format. Please upload an Excel file.')
        
        return redirect('scenarios_view')  # Redirect to the appropriate view

    return render(request, 'frontend/techcare_data/scenarios_view.html')



def export_selfAwarnessBites_data(request):
    # Fetch data from Firestore
    db = firestore.client()
    collection = db.collection("selfAwarnessBites")
    documents = collection.stream()

    document_data = []
    for doc in documents:
        data = doc.to_dict()
        # Handle specific Firestore types and flatten data
        flat_data = {key: handle_value(value) for key, value in data.items()}
        document_data.append(flat_data)

    if not document_data:
        return HttpResponse("No data available to export.", content_type="text/plain")

    # Create a new Excel workbook and select the active worksheet
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "SelfAwarnessBites Data"

    # Define the column headers
    headers = list(set(header for entry in document_data for header in entry.keys()))  # Ensure all headers are included
    worksheet.append(headers)

    # Add rows to the worksheet
    for entry in document_data:
        row = [entry.get(header, '') for header in headers]
        worksheet.append(row)

    # Save the workbook to a BytesIO stream
    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)

    # Create the HTTP response with the Excel file
    response = HttpResponse(
        stream.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response['Content-Disposition'] = 'attachment; filename="selfAwarnessBites.xlsx"'
    return response


def import_selfAwarnessBites_data(request):
    if request.method == 'POST' and request.FILES.get('import_file'):
        file = request.FILES['import_file']
        if file.name.endswith('.xlsx') or file.name.endswith('.xls'):
            # Save the file temporarily
            file_name = default_storage.save(file.name, file)
            file_path = Path(settings.MEDIA_ROOT) / file_name

            # Process the file
            try:
                # Read the Excel file into a DataFrame
                df = pd.read_excel(file_path)

                # Initialize Firestore client
                db = firestore.Client()

                # Reference to the Firestore collection
                collection_ref = db.collection("selfAwarnessBites")

                # Add each row from the DataFrame to Firestore
                for index, row in df.iterrows():
                    document_data = {
                        "scenarioID": row.get('scenarioID'),
                        "selfawarenessBiteText": row.get('selfawarenessBiteText'),
                        "selfawarenessBiteTitle": row.get('selfawarenessBiteTitle'),           
                        "tags": row.get('tags'),                      
                    }
                    # Add document to Firestore with auto-generated ID
                    collection_ref.add(document_data)

                messages.success(request, 'Data imported successfully!')
            except Exception as e:
                messages.error(request, f'Error importing data: {e}')
        else:
            messages.error(request, 'Invalid file format. Please upload an Excel file.')
        
        return redirect('selfAwarnessBites_view')  # Redirect to the appropriate view

    return render(request, 'frontend/techcare_data/selfAwarnessBites_view.html')



def export_selfLadder_data(request):
    # Fetch data from Firestore
    db = firestore.client()
    collection = db.collection("selfLadder")
    documents = collection.stream()

    document_data = []
    for doc in documents:
        data = doc.to_dict()
        # Handle specific Firestore types and flatten data
        flat_data = {key: handle_value(value) for key, value in data.items()}
        document_data.append(flat_data)

    if not document_data:
        return HttpResponse("No data available to export.", content_type="text/plain")

    # Create a new Excel workbook and select the active worksheet
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "SelfLadder Data"

    # Define the column headers
    headers = list(set(header for entry in document_data for header in entry.keys()))  # Ensure all headers are included
    worksheet.append(headers)

    # Add rows to the worksheet
    for entry in document_data:
        row = [entry.get(header, '') for header in headers]
        worksheet.append(row)

    # Save the workbook to a BytesIO stream
    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)

    # Create the HTTP response with the Excel file
    response = HttpResponse(
        stream.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response['Content-Disposition'] = 'attachment; filename="selfLadder.xlsx"'
    return response


def import_selfLadder_data(request):
    if request.method == 'POST' and request.FILES.get('import_file'):
        file = request.FILES['import_file']
        if file.name.endswith('.xlsx') or file.name.endswith('.xls'):
            # Save the file temporarily
            file_name = default_storage.save(file.name, file)
            file_path = Path(settings.MEDIA_ROOT) / file_name

            # Process the file
            try:
                # Read the Excel file into a DataFrame
                df = pd.read_excel(file_path)

                # Initialize Firestore client
                db = firestore.Client()

                # Reference to the Firestore collection
                collection_ref = db.collection("selfLadder")

                # Add each row from the DataFrame to Firestore
                for index, row in df.iterrows():
                    document_data = {
                        "time": row.get('time'),
                        "type": row.get('type'),
                        "userID": row.get('userID'),
                        
                    }
                    # Add document to Firestore with auto-generated ID
                    collection_ref.add(document_data)

                messages.success(request, 'Data imported successfully!')
            except Exception as e:
                messages.error(request, f'Error importing data: {e}')
        else:
            messages.error(request, 'Invalid file format. Please upload an Excel file.')
        
        return redirect('selfladder_view')  # Redirect to the appropriate view

    return render(request, 'frontend/techcare_data/selfladder_view.html')




def export_selfawarenessScenarios_data(request):
    # Fetch data from Firestore
    db = firestore.client()
    collection = db.collection("selfawarenessScenarios")
    documents = collection.stream()

    document_data = []
    for doc in documents:
        data = doc.to_dict()
        # Handle specific Firestore types and flatten data
        flat_data = {key: handle_value(value) for key, value in data.items()}
        document_data.append(flat_data)

    if not document_data:
        return HttpResponse("No data available to export.", content_type="text/plain")

    # Create a new Excel workbook and select the active worksheet
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "SelfawarenessScenarios Data"

    # Define the column headers
    headers = list(set(header for entry in document_data for header in entry.keys()))  # Ensure all headers are included
    worksheet.append(headers)

    # Add rows to the worksheet
    for entry in document_data:
        row = [entry.get(header, '') for header in headers]
        worksheet.append(row)

    # Save the workbook to a BytesIO stream
    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)

    # Create the HTTP response with the Excel file
    response = HttpResponse(
        stream.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response['Content-Disposition'] = 'attachment; filename="selfawarenessScenarios.xlsx"'
    return response


def import_selfawarenessScenarios_data(request):
    if request.method == 'POST' and request.FILES.get('import_file'):
        file = request.FILES['import_file']
        if file.name.endswith('.xlsx') or file.name.endswith('.xls'):
            # Save the file temporarily
            file_name = default_storage.save(file.name, file)
            file_path = Path(settings.MEDIA_ROOT) / file_name

            # Process the file
            try:
                # Read the Excel file into a DataFrame
                df = pd.read_excel(file_path)

                # Initialize Firestore client
                db = firestore.Client()

                # Reference to the Firestore collection
                collection_ref = db.collection("selfawarenessScenarios")

                # Add each row from the DataFrame to Firestore
                for index, row in df.iterrows():
                    
                    activities_title = row.get('activity')

                  
                    activities_ref = db.collection("activities")
                    activities_query = activities_ref.where('audiotrackTitle', '==', activities_title).limit(1).stream()
                    activities_ref = None
                    for activities in activities_query:
                        activities_ref = activities.reference
                        break

                    if not activities_ref:
                        messages.error(request, f'activity "{activities_title}" not found in activities collection.')
                        continue
                    
                    inAppLink_title = row.get('inAppLink')

                  
                    inAppLink_ref = db.collection("inAppLinks")
                    inAppLink_query = inAppLink_ref.where('title', '==', inAppLink_title).limit(1).stream()
                    inAppLink_ref = None
                    for inAppLink in inAppLink_query:
                        inAppLink_ref = inAppLink.reference
                        break

                    if not inAppLink_ref:
                        messages.error(request, f'inAppLink "{inAppLink_title}" not found in inAppLink collection.')
                        continue
                    
                    
                    wildcard_title = row.get('wildcard')
                  
                    wildcard_ref = db.collection("wildCard")
                    wildcard_query = wildcard_ref.where('content', '==', wildcard_title).limit(1).stream()
                    wildcard_ref = None
                    for wildcard in wildcard_query:
                        wildcard_ref = wildcard.reference
                        break

                    if not wildcard_ref:
                        messages.error(request, f'wildcard "{wildcard_title}" not found in wildcard collection.')
                        continue
                    
                    
                    journal_title = row.get('journal')
                  
                    journal_ref = db.collection("journalPrompt")
                    journal_query = journal_ref.where('title', '==', journal_title).limit(1).stream()
                    journal_ref = None
                    for journal in journal_query:
                        journal_ref = journal.reference
                        break

                    if not journal_ref:
                        messages.error(request, f'journal "{journal_title}" not found in journalPrompt collection.')
                        continue
                    
                    normalBite_title = row.get('normalBite')
                  
                    normalBite_ref = db.collection("bites")
                    normalBite_query = normalBite_ref.where('title', '==', normalBite_title).limit(1).stream()
                    normalBite_ref = None
                    for normalBite in normalBite_query:
                        normalBite_ref = normalBite.reference
                        break

                    if not normalBite_ref:
                        messages.error(request, f'normalBite "{normalBite_title}" not found in bites collection.')
                        continue
                    
                    
                    selfAwarnessBites_title = row.get('biteID')
                  
                    selfAwarnessBites_ref = db.collection("selfAwarnessBites")
                    selfAwarnessBites_query = selfAwarnessBites_ref.where('selfawarenessBiteTitle', '==', selfAwarnessBites_title).limit(1).stream()
                    selfAwarnessBites_ref = None
                    for selfAwarnessBites in selfAwarnessBites_query:
                        selfAwarnessBites_ref = selfAwarnessBites.reference
                        break

                    if not selfAwarnessBites_ref:
                        messages.error(request, f'selfAwarnessBite "{selfAwarnessBites_title}" not found in selfAwarnessBites collection.')
                        continue
                    
                    document_data = {
                        "activity": activities_ref,
                        "biteID":selfAwarnessBites_ref,
                        "correction1from0to2": row.get('correction1from0to2'),
                        "correction2from3to5": row.get('correction2from3to5'),
                        "inAppLink": inAppLink_ref,
                        "interactiveStatement": row.get('interactiveStatement'),
                        "journal": journal_ref,
                        "normalBite": normalBite_ref,
                        "recommendation1": row.get('recommendation1'),
                        "recommendation2": row.get('recommendation2'),
                        "scenarioID": row.get('scenarioID'),
                        "story": row.get('story'),
                        "storyTitle": row.get('storyTitle'),
                        "wildcard": wildcard_ref,
                        
                    }
                    # Add document to Firestore with auto-generated ID
                    collection_ref.add(document_data)

                messages.success(request, 'Data imported successfully!')
            except Exception as e:
                messages.error(request, f'Error importing data: {e}')
        else:
            messages.error(request, 'Invalid file format. Please upload an Excel file.')
        
        return redirect('selfawarenessScenarios_view')  # Redirect to the appropriate view

    return render(request, 'frontend/techcare_data/selfawarenessScenarios_view.html')




def export_selfawareness_collection_data(request):
    # Fetch data from Firestore
    db = firestore.client()
    collection = db.collection("selfawareness_collection")
    documents = collection.stream()

    document_data = []
    for doc in documents:
        data = doc.to_dict()
        # Handle specific Firestore types and flatten data
        flat_data = {key: handle_value(value) for key, value in data.items()}
        document_data.append(flat_data)

    if not document_data:
        return HttpResponse("No data available to export.", content_type="text/plain")

    # Create a new Excel workbook and select the active worksheet
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Selfawareness_collection Data"

    # Define the column headers
    headers = list(set(header for entry in document_data for header in entry.keys()))  # Ensure all headers are included
    worksheet.append(headers)

    # Add rows to the worksheet
    for entry in document_data:
        row = [entry.get(header, '') for header in headers]
        worksheet.append(row)

    # Save the workbook to a BytesIO stream
    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)

    # Create the HTTP response with the Excel file
    response = HttpResponse(
        stream.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response['Content-Disposition'] = 'attachment; filename="selfawareness_collection.xlsx"'
    return response


def import_selfawareness_collection_data(request):
    if request.method == 'POST' and request.FILES.get('import_file'):
        file = request.FILES['import_file']
        if file.name.endswith('.xlsx') or file.name.endswith('.xls'):
            # Save the file temporarily
            file_name = default_storage.save(file.name, file)
            file_path = Path(settings.MEDIA_ROOT) / file_name

            # Process the file
            try:
                # Read the Excel file into a DataFrame
                df = pd.read_excel(file_path)

                # Initialize Firestore client
                db = firestore.Client()

                # Reference to the Firestore collection
                collection_ref = db.collection("selfawareness_collection")

                # Add each row from the DataFrame to Firestore
                for index, row in df.iterrows():
                    document_data = {
                        "scenario_id": row.get('scenario_id'),
                        "selfawareness_bite_text": row.get('selfawareness_bite_text'),
                        "selfawareness_bite_title": row.get('selfawareness_bite_title'),
                        "tags": row.get('tags'),
                    }
                    # Add document to Firestore with auto-generated ID
                    collection_ref.add(document_data)

                messages.success(request, 'Data imported successfully!')
            except Exception as e:
                messages.error(request, f'Error importing data: {e}')
        else:
            messages.error(request, 'Invalid file format. Please upload an Excel file.')
        
        return redirect('selfawareness_collection_view')  # Redirect to the appropriate view

    return render(request, 'frontend/techcare_data/selfawareness_collection_view.html')





def export_shortBite_data(request):
    # Fetch data from Firestore
    db = firestore.client()
    collection = db.collection("shortBite")
    documents = collection.stream()

    document_data = []
    for doc in documents:
        data = doc.to_dict()
        # Handle specific Firestore types and flatten data
        flat_data = {key: handle_value(value) for key, value in data.items()}
        document_data.append(flat_data)

    if not document_data:
        return HttpResponse("No data available to export.", content_type="text/plain")

    # Create a new Excel workbook and select the active worksheet
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "ShortBite Data"

    # Define the column headers
    headers = list(set(header for entry in document_data for header in entry.keys()))  # Ensure all headers are included
    worksheet.append(headers)

    # Add rows to the worksheet
    for entry in document_data:
        row = [entry.get(header, '') for header in headers]
        worksheet.append(row)

    # Save the workbook to a BytesIO stream
    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)

    # Create the HTTP response with the Excel file
    response = HttpResponse(
        stream.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response['Content-Disposition'] = 'attachment; filename="shortBite.xlsx"'
    return response


def import_shortBite_data(request):
    if request.method == 'POST' and request.FILES.get('import_file'):
        file = request.FILES['import_file']
        if file.name.endswith('.xlsx') or file.name.endswith('.xls'):
            # Save the file temporarily
            file_name = default_storage.save(file.name, file)
            file_path = Path(settings.MEDIA_ROOT) / file_name

            # Process the file
            try:
                # Read the Excel file into a DataFrame
                df = pd.read_excel(file_path)

                # Initialize Firestore client
                db = firestore.Client()

                # Reference to the Firestore collection
                collection_ref = db.collection("shortBite")

                # Add each row from the DataFrame to Firestore
                for index, row in df.iterrows():
                    document_data = {
                        "order": row.get('order'),
                        "title": row.get('title'),
                    }
                    # Add document to Firestore with auto-generated ID
                    collection_ref.add(document_data)

                messages.success(request, 'Data imported successfully!')
            except Exception as e:
                messages.error(request, f'Error importing data: {e}')
        else:
            messages.error(request, 'Invalid file format. Please upload an Excel file.')
        
        return redirect('shortBite_view')  # Redirect to the appropriate view

    return render(request, 'frontend/techcare_data/shortBite_view.html')
    





def export_suggestedActivities_data(request):
    # Fetch data from Firestore
    db = firestore.client()
    collection = db.collection("suggestedActivities")
    documents = collection.stream()

    document_data = []
    for doc in documents:
        data = doc.to_dict()
        # Handle specific Firestore types and flatten data
        flat_data = {key: handle_value(value) for key, value in data.items()}
        document_data.append(flat_data)

    if not document_data:
        return HttpResponse("No data available to export.", content_type="text/plain")

    # Create a new Excel workbook and select the active worksheet
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "SuggestedActivities Data"

    # Define the column headers
    headers = list(set(header for entry in document_data for header in entry.keys()))  # Ensure all headers are included
    worksheet.append(headers)

    # Add rows to the worksheet
    for entry in document_data:
        row = [entry.get(header, '') for header in headers]
        worksheet.append(row)

    # Save the workbook to a BytesIO stream
    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)

    # Create the HTTP response with the Excel file
    response = HttpResponse(
        stream.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response['Content-Disposition'] = 'attachment; filename="suggestedActivities.xlsx"'
    return response


def import_suggestedActivities_data(request):
    if request.method == 'POST' and request.FILES.get('import_file'):
        file = request.FILES['import_file']
        if file.name.endswith('.xlsx') or file.name.endswith('.xls'):
            # Save the file temporarily
            file_name = default_storage.save(file.name, file)
            file_path = Path(settings.MEDIA_ROOT) / file_name

            # Process the file
            try:
                # Read the Excel file into a DataFrame
                df = pd.read_excel(file_path)

                # Initialize Firestore client
                db = firestore.Client()

                # Reference to the Firestore collection
                collection_ref = db.collection("suggestedActivities")

                # Add each row from the DataFrame to Firestore
                for index, row in df.iterrows():
                    document_data = {
                        "activity": row.get('activity'),
                        "time": row.get('time'),
                        "type": row.get('type'),
                        "user": row.get('user'),
                    }
                    # Add document to Firestore with auto-generated ID
                    collection_ref.add(document_data)

                messages.success(request, 'Data imported successfully!')
            except Exception as e:
                messages.error(request, f'Error importing data: {e}')
        else:
            messages.error(request, 'Invalid file format. Please upload an Excel file.')
        
        return redirect('suggestedActivities_view')  # Redirect to the appropriate view

    return render(request, 'frontend/techcare_data/suggestedActivities_view.html')




def export_suggestedBites_data(request):
    # Fetch data from Firestore
    db = firestore.client()
    collection = db.collection("suggestedBites")
    documents = collection.stream()

    document_data = []
    for doc in documents:
        data = doc.to_dict()
        # Handle specific Firestore types and flatten data
        flat_data = {key: handle_value(value) for key, value in data.items()}
        document_data.append(flat_data)

    if not document_data:
        return HttpResponse("No data available to export.", content_type="text/plain")

    # Create a new Excel workbook and select the active worksheet
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "SuggestedBites Data"

    # Define the column headers
    headers = list(set(header for entry in document_data for header in entry.keys()))  # Ensure all headers are included
    worksheet.append(headers)

    # Add rows to the worksheet
    for entry in document_data:
        row = [entry.get(header, '') for header in headers]
        worksheet.append(row)

    # Save the workbook to a BytesIO stream
    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)

    # Create the HTTP response with the Excel file
    response = HttpResponse(
        stream.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response['Content-Disposition'] = 'attachment; filename="suggestedBites.xlsx"'
    return response


def import_suggestedBites_data(request):
    if request.method == 'POST' and request.FILES.get('import_file'):
        file = request.FILES['import_file']
        if file.name.endswith('.xlsx') or file.name.endswith('.xls'):
            # Save the file temporarily
            file_name = default_storage.save(file.name, file)
            file_path = Path(settings.MEDIA_ROOT) / file_name

            # Process the file
            try:
                # Read the Excel file into a DataFrame
                df = pd.read_excel(file_path)

                # Initialize Firestore client
                db = firestore.Client()

                # Reference to the Firestore collection
                collection_ref = db.collection("suggestedBites")

                # Add each row from the DataFrame to Firestore
                for index, row in df.iterrows():
                    document_data = {
                        "bite": row.get('bite'),
                        "state": row.get('state'),
                        "time": row.get('time'),
                        "user": row.get('user'),
                    }
                    # Add document to Firestore with auto-generated ID
                    collection_ref.add(document_data)

                messages.success(request, 'Data imported successfully!')
            except Exception as e:
                messages.error(request, f'Error importing data: {e}')
        else:
            messages.error(request, 'Invalid file format. Please upload an Excel file.')
        
        return redirect('suggestedBites_view')  # Redirect to the appropriate view

    return render(request, 'frontend/techcare_data/suggestedBites_view.html')





def export_suggestedInAppLinks_data(request):
    # Fetch data from Firestore
    db = firestore.client()
    collection = db.collection("suggestedInAppLinks")
    documents = collection.stream()

    document_data = []
    for doc in documents:
        data = doc.to_dict()
        # Handle specific Firestore types and flatten data
        flat_data = {key: handle_value(value) for key, value in data.items()}
        document_data.append(flat_data)

    if not document_data:
        return HttpResponse("No data available to export.", content_type="text/plain")

    # Create a new Excel workbook and select the active worksheet
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "SuggestedInAppLinks Data"

    # Define the column headers
    headers = list(set(header for entry in document_data for header in entry.keys()))  # Ensure all headers are included
    worksheet.append(headers)

    # Add rows to the worksheet
    for entry in document_data:
        row = [entry.get(header, '') for header in headers]
        worksheet.append(row)

    # Save the workbook to a BytesIO stream
    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)

    # Create the HTTP response with the Excel file
    response = HttpResponse(
        stream.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response['Content-Disposition'] = 'attachment; filename="suggestedInAppLinks.xlsx"'
    return response


def import_suggestedInAppLinks_data(request):
    if request.method == 'POST' and request.FILES.get('import_file'):
        file = request.FILES['import_file']
        if file.name.endswith('.xlsx') or file.name.endswith('.xls'):
            # Save the file temporarily
            file_name = default_storage.save(file.name, file)
            file_path = Path(settings.MEDIA_ROOT) / file_name

            # Process the file
            try:
                # Read the Excel file into a DataFrame
                df = pd.read_excel(file_path)

                # Initialize Firestore client
                db = firestore.Client()

                # Reference to the Firestore collection
                collection_ref = db.collection("suggestedInAppLinks")

                # Add each row from the DataFrame to Firestore
                for index, row in df.iterrows():
                    document_data = {
                        "inAppLink": row.get('inAppLink'),
                        "time": row.get('time'),
                        "user": row.get('user'),
                    }
                    # Add document to Firestore with auto-generated ID
                    collection_ref.add(document_data)

                messages.success(request, 'Data imported successfully!')
            except Exception as e:
                messages.error(request, f'Error importing data: {e}')
        else:
            messages.error(request, 'Invalid file format. Please upload an Excel file.')
        
        return redirect('suggestedInAppLinks_view')  # Redirect to the appropriate view

    return render(request, 'frontend/techcare_data/suggestedInAppLinks_view.html')





def export_suggestedJournals_data(request):
    # Fetch data from Firestore
    db = firestore.client()
    collection = db.collection("suggestedJournals")
    documents = collection.stream()

    document_data = []
    for doc in documents:
        data = doc.to_dict()
        # Handle specific Firestore types and flatten data
        flat_data = {key: handle_value(value) for key, value in data.items()}
        document_data.append(flat_data)

    if not document_data:
        return HttpResponse("No data available to export.", content_type="text/plain")

    # Create a new Excel workbook and select the active worksheet
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "SuggestedJournals Data"

    # Define the column headers
    headers = list(set(header for entry in document_data for header in entry.keys()))  # Ensure all headers are included
    worksheet.append(headers)

    # Add rows to the worksheet
    for entry in document_data:
        row = [entry.get(header, '') for header in headers]
        worksheet.append(row)

    # Save the workbook to a BytesIO stream
    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)

    # Create the HTTP response with the Excel file
    response = HttpResponse(
        stream.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response['Content-Disposition'] = 'attachment; filename="suggestedJournals.xlsx"'
    return response


def import_suggestedJournals_data(request):
    if request.method == 'POST' and request.FILES.get('import_file'):
        file = request.FILES['import_file']
        if file.name.endswith('.xlsx') or file.name.endswith('.xls'):
            # Save the file temporarily
            file_name = default_storage.save(file.name, file)
            file_path = Path(settings.MEDIA_ROOT) / file_name

            # Process the file
            try:
                # Read the Excel file into a DataFrame
                df = pd.read_excel(file_path)

                # Initialize Firestore client
                db = firestore.Client()

                # Reference to the Firestore collection
                collection_ref = db.collection("suggestedJournals")

                # Add each row from the DataFrame to Firestore
                for index, row in df.iterrows():
                    document_data = {
                        "journal": row.get('journal'),
                        "state": row.get('state'),
                        "time": row.get('time'),
                        "user": row.get('user'),
                    }
                    # Add document to Firestore with auto-generated ID
                    collection_ref.add(document_data)

                messages.success(request, 'Data imported successfully!')
            except Exception as e:
                messages.error(request, f'Error importing data: {e}')
        else:
            messages.error(request, 'Invalid file format. Please upload an Excel file.')
        
        return redirect('suggestedJournals_view')  # Redirect to the appropriate view

    return render(request, 'frontend/techcare_data/suggestedJournals_view.html')



def export_suggestedSelfAwarnessBites_data(request):
    # Fetch data from Firestore
    db = firestore.client()
    collection = db.collection("suggestedSelfAwarnessBites")
    documents = collection.stream()

    document_data = []
    for doc in documents:
        data = doc.to_dict()
        # Handle specific Firestore types and flatten data
        flat_data = {key: handle_value(value) for key, value in data.items()}
        document_data.append(flat_data)

    if not document_data:
        return HttpResponse("No data available to export.", content_type="text/plain")

    # Create a new Excel workbook and select the active worksheet
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "SuggestedSelfAwarnessBites Data"

    # Define the column headers
    headers = list(set(header for entry in document_data for header in entry.keys()))  # Ensure all headers are included
    worksheet.append(headers)

    # Add rows to the worksheet
    for entry in document_data:
        row = [entry.get(header, '') for header in headers]
        worksheet.append(row)

    # Save the workbook to a BytesIO stream
    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)

    # Create the HTTP response with the Excel file
    response = HttpResponse(
        stream.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response['Content-Disposition'] = 'attachment; filename="suggestedSelfAwarnessBites.xlsx"'
    return response


def import_suggestedSelfAwarnessBites_data(request):
    if request.method == 'POST' and request.FILES.get('import_file'):
        file = request.FILES['import_file']
        if file.name.endswith('.xlsx') or file.name.endswith('.xls'):
            # Save the file temporarily
            file_name = default_storage.save(file.name, file)
            file_path = Path(settings.MEDIA_ROOT) / file_name

            # Process the file
            try:
                # Read the Excel file into a DataFrame
                df = pd.read_excel(file_path)

                # Initialize Firestore client
                db = firestore.Client()

                # Reference to the Firestore collection
                collection_ref = db.collection("suggestedSelfAwarnessBites")

                # Add each row from the DataFrame to Firestore
                for index, row in df.iterrows():
                    document_data = {
                        "journal": row.get('journal'),
                        "state": row.get('state'),
                        "time": row.get('time'),
                        "user": row.get('user'),
                    }
                    # Add document to Firestore with auto-generated ID
                    collection_ref.add(document_data)

                messages.success(request, 'Data imported successfully!')
            except Exception as e:
                messages.error(request, f'Error importing data: {e}')
        else:
            messages.error(request, 'Invalid file format. Please upload an Excel file.')
        
        return redirect('suggestedSelfAwarnessBites_view')  # Redirect to the appropriate view

    return render(request, 'frontend/techcare_data/suggestedSelfAwarnessBites_view.html')




def export_suggestedWildCards_data(request):
    # Fetch data from Firestore
    db = firestore.client()
    collection = db.collection("suggestedWildCards")
    documents = collection.stream()

    document_data = []
    for doc in documents:
        data = doc.to_dict()
        # Handle specific Firestore types and flatten data
        flat_data = {key: handle_value(value) for key, value in data.items()}
        document_data.append(flat_data)

    if not document_data:
        return HttpResponse("No data available to export.", content_type="text/plain")

    # Create a new Excel workbook and select the active worksheet
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "SuggestedWildCards Data"

    # Define the column headers
    headers = list(set(header for entry in document_data for header in entry.keys()))  # Ensure all headers are included
    worksheet.append(headers)

    # Add rows to the worksheet
    for entry in document_data:
        row = [entry.get(header, '') for header in headers]
        worksheet.append(row)

    # Save the workbook to a BytesIO stream
    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)

    # Create the HTTP response with the Excel file
    response = HttpResponse(
        stream.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response['Content-Disposition'] = 'attachment; filename="suggestedWildCards.xlsx"'
    return response


def import_suggestedWildCards_data(request):
    if request.method == 'POST' and request.FILES.get('import_file'):
        file = request.FILES['import_file']
        if file.name.endswith('.xlsx') or file.name.endswith('.xls'):
            # Save the file temporarily
            file_name = default_storage.save(file.name, file)
            file_path = Path(settings.MEDIA_ROOT) / file_name

            # Process the file
            try:
                # Read the Excel file into a DataFrame
                df = pd.read_excel(file_path)

                # Initialize Firestore client
                db = firestore.Client()

                # Reference to the Firestore collection
                collection_ref = db.collection("suggestedWildCards")

                # Add each row from the DataFrame to Firestore
                for index, row in df.iterrows():
                    document_data = {
                        "wildCard": row.get('wildCard'),
                        "time": row.get('time'),
                        "user": row.get('user'),
                    }
                    # Add document to Firestore with auto-generated ID
                    collection_ref.add(document_data)

                messages.success(request, 'Data imported successfully!')
            except Exception as e:
                messages.error(request, f'Error importing data: {e}')
        else:
            messages.error(request, 'Invalid file format. Please upload an Excel file.')
        
        return redirect('suggestedWildCards_view')  # Redirect to the appropriate view

    return render(request, 'frontend/techcare_data/suggestedWildCards_view.html')


def export_trivia_data(request):
    # Fetch data from Firestore
    db = firestore.client()
    collection = db.collection("trivia")
    documents = collection.stream()

    document_data = []
    for doc in documents:
        data = doc.to_dict()
        # Handle specific Firestore types and flatten data
        flat_data = {key: handle_value(value) for key, value in data.items()}
        document_data.append(flat_data)

    if not document_data:
        return HttpResponse("No data available to export.", content_type="text/plain")

    # Create a new Excel workbook and select the active worksheet
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Trivia Data"

    # Define the column headers
    headers = list(set(header for entry in document_data for header in entry.keys()))  # Ensure all headers are included
    worksheet.append(headers)

    # Add rows to the worksheet
    for entry in document_data:
        row = [entry.get(header, '') for header in headers]
        worksheet.append(row)

    # Save the workbook to a BytesIO stream
    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)

    # Create the HTTP response with the Excel file
    response = HttpResponse(
        stream.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response['Content-Disposition'] = 'attachment; filename="trivia.xlsx"'
    return response


def import_trivia_data(request):
    if request.method == 'POST' and request.FILES.get('import_file'):
        file = request.FILES['import_file']
        if file.name.endswith('.xlsx') or file.name.endswith('.xls'):
            # Save the file temporarily
            file_name = default_storage.save(file.name, file)
            file_path = Path(settings.MEDIA_ROOT) / file_name

            # Process the file
            try:
                # Read the Excel file into a DataFrame
                df = pd.read_excel(file_path)

                # Initialize Firestore client
                db = firestore.Client()

                # Reference to the Firestore collection
                collection_ref = db.collection("trivia")

                # Add each row from the DataFrame to Firestore
                for index, row in df.iterrows():
                    document_data = {
                        "answer_1": row.get('answer_1'),
                        "answer_2": row.get('answer_2'),
                        "answer_3": row.get('answer_3'),
                        "answer_4": row.get('answer_4'),
                        "correct_answer": row.get('correct_answer'),
                        "question": row.get('question'),
                        "question_ID": row.get('question_ID'),
                        
                    }
                    # Add document to Firestore with auto-generated ID
                    collection_ref.add(document_data)

                messages.success(request, 'Data imported successfully!')
            except Exception as e:
                messages.error(request, f'Error importing data: {e}')
        else:
            messages.error(request, 'Invalid file format. Please upload an Excel file.')
        
        return redirect('trivia_view')  # Redirect to the appropriate view

    return render(request, 'frontend/techcare_data/suggestedWildCards_view.html')

def nutrition_delete_selected(request):
    document_ids = request.POST.getlist('documents')
    
    db = firestore.client()
    for document_id in document_ids:
        db.collection('nutrition').document(document_id).delete()
    messages.success(request, "Selected documents have been successfully deleted.")
    return redirect(('nutrition_view'))


def activities_delete_selected(request):
    document_ids = request.POST.getlist('documents')
    
    db = firestore.client()
    for document_id in document_ids:
        db.collection('activities').document(document_id).delete()
    messages.success(request, "Selected documents have been successfully deleted.")
    return redirect(('activities_view'))

def assessmentQuestion_delete_selected(request):
    document_ids = request.POST.getlist('documents')
    
    db = firestore.client()
    for document_id in document_ids:
        db.collection('assessmentQuestion').document(document_id).delete()
    messages.success(request, "Selected documents have been successfully deleted.")
    return redirect(('assessmentQuestion_view'))

def assets_delete_selected(request):
    document_ids = request.POST.getlist('documents')
    
    db = firestore.client()
    for document_id in document_ids:
        db.collection('assets').document(document_id).delete()
    messages.success(request, "Selected documents have been successfully deleted.")
    return redirect(('assets_view'))



def biomarkers_delete_selected(request):
    document_ids = request.POST.getlist('documents')
    
    db = firestore.client()
    for document_id in document_ids:
        db.collection('biomarkers').document(document_id).delete()
    messages.success(request, "Selected documents have been successfully deleted.")
    return redirect(('biomarkers_view'))

def bites_delete_selected(request):
    document_ids = request.POST.getlist('documents')
    
    db = firestore.client()
    for document_id in document_ids:
        db.collection('bites').document(document_id).delete()
    messages.success(request, "Selected documents have been successfully deleted.")
    return redirect(('bites_view'))

def categories_delete_selected(request):
    document_ids = request.POST.getlist('documents')
    
    db = firestore.client()
    for document_id in document_ids:
        db.collection('categories').document(document_id).delete()
    messages.success(request, "Selected documents have been successfully deleted.")
    return redirect(('categories_view'))

def inAppLinks_delete_selected(request):
    document_ids = request.POST.getlist('documents')
    
    db = firestore.client()
    for document_id in document_ids:
        db.collection('inAppLinks').document(document_id).delete()
    messages.success(request, "Selected documents have been successfully deleted.")
    return redirect(('inAppLinks_view'))

def inquiry_delete_selected(request):
    document_ids = request.POST.getlist('documents')
    
    db = firestore.client()
    for document_id in document_ids:
        db.collection('inquiry').document(document_id).delete()
    messages.success(request, "Selected documents have been successfully deleted.")
    return redirect(('inquiry_view'))

def journalPrompt_delete_selected(request):
    document_ids = request.POST.getlist('documents')
    
    db = firestore.client()
    for document_id in document_ids:
        db.collection('journalPrompt').document(document_id).delete()
    messages.success(request, "Selected documents have been successfully deleted.")
    return redirect(('journalPrompt_view'))

def majorAssessment_delete_selected(request):
    document_ids = request.POST.getlist('documents')
    
    db = firestore.client()
    for document_id in document_ids:
        db.collection('majorAssessment').document(document_id).delete()
    messages.success(request, "Selected documents have been successfully deleted.")
    return redirect(('majorAssessment_view'))


def psychomarkers_delete_selected(request):
    document_ids = request.POST.getlist('documents')
    
    db = firestore.client()
    for document_id in document_ids:
        db.collection('psychomarkers').document(document_id).delete()
    messages.success(request, "Selected documents have been successfully deleted.")
    return redirect(('psychomarkers_view'))

def scenarios_delete_selected(request):
    document_ids = request.POST.getlist('documents')
    
    db = firestore.client()
    for document_id in document_ids:
        db.collection('scenarios').document(document_id).delete()
    messages.success(request, "Selected documents have been successfully deleted.")
    return redirect(('scenarios_view'))

def selfAwarnessBites_delete_selected(request):
    document_ids = request.POST.getlist('documents')
    
    db = firestore.client()
    for document_id in document_ids:
        db.collection('selfAwarnessBites').document(document_id).delete()
    messages.success(request, "Selected documents have been successfully deleted.")
    return redirect(('selfAwarnessBites_view'))

def selfawarenessScenarios_delete_selected(request):
    document_ids = request.POST.getlist('documents')
    
    db = firestore.client()
    for document_id in document_ids:
        db.collection('selfawarenessScenarios').document(document_id).delete()
    messages.success(request, "Selected documents have been successfully deleted.")
    return redirect(('selfawarenessScenarios_view'))

def selfawarenessCollection_delete_selected(request):
    document_ids = request.POST.getlist('documents')
    
    db = firestore.client()
    for document_id in document_ids:
        db.collection('selfawarenessCollection').document(document_id).delete()
    messages.success(request, "Selected documents have been successfully deleted.")
    return redirect(('selfawarenessCollection_view'))

def notifications_delete_selected(request):
    document_ids = request.POST.getlist('documents')
    
    db = firestore.client()
    for document_id in document_ids:
        db.collection('notifications').document(document_id).delete()
    messages.success(request, "Selected documents have been successfully deleted.")
    return redirect(('notifications_view'))

def shortBite_delete_selected(request):
    document_ids = request.POST.getlist('documents')
    
    db = firestore.client()
    for document_id in document_ids:
        db.collection('shortBite').document(document_id).delete()
    messages.success(request, "Selected documents have been successfully deleted.")
    return redirect(('shortBite_view'))

def tags_delete_selected(request):
    document_ids = request.POST.getlist('documents')
    
    db = firestore.client()
    for document_id in document_ids:
        db.collection('tags').document(document_id).delete()
    messages.success(request, "Selected documents have been successfully deleted.")
    return redirect(('tags_view'))


def testTrivia_delete_selected(request):
    document_ids = request.POST.getlist('documents')
    
    db = firestore.client()
    for document_id in document_ids:
        db.collection('testTrivia').document(document_id).delete()
    messages.success(request, "Selected documents have been successfully deleted.")
    return redirect(('testTrivia_view'))


def trivia_delete_selected(request):
    document_ids = request.POST.getlist('documents')
    
    db = firestore.client()
    for document_id in document_ids:
        db.collection('trivia').document(document_id).delete()
    messages.success(request, "Selected documents have been successfully deleted.")
    return redirect(('trivia_view'))

def users_delete_selected(request):
    document_ids = request.POST.getlist('documents')
    
    db = firestore.client()
    for document_id in document_ids:
        db.collection('users').document(document_id).delete()
    messages.success(request, "Selected documents have been successfully deleted.")
    return redirect(('users_view'))

def wildCard_delete_selected(request):
    document_ids = request.POST.getlist('documents')
    
    db = firestore.client()
    for document_id in document_ids:
        db.collection('wildCard').document(document_id).delete()
    messages.success(request, "Selected documents have been successfully deleted.")
    return redirect(('wildCard_view'))
